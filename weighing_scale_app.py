# Truck Scale Weighing System - Main Application
# This module contains the main application logic and UI

import tkinter as tk
from tkinter import ttk, simpledialog, filedialog
import threading
import time
import json
import os
import tempfile
import subprocess
import re
from datetime import datetime, timedelta
import csv
import sys
import logging
import shutil
import hashlib
import getpass
import atexit
import base64
import secrets
import io
import sqlite3
import serial
from PIL import Image, ImageTk

# Import camera manager
from camera_manager import CameraManager

# Constants for feature availability
try:
    import win32print
    PRINTING_ENABLED = True
except ImportError:
    PRINTING_ENABLED = False

try:
    import reportlab
    PDF_PRINTING_ENABLED = True
except ImportError:
    PDF_PRINTING_ENABLED = False

# Import refactored modules
from config import (
    PROGRAM_DATA_DIR, CONFIG_FILE, DB_FILE, ADMIN_PASSWORD_HASH, ADMIN_USERNAME,
    ACTIVATION_CODE, EXPIRY_DAYS_INITIAL_TRIAL, EXPIRY_DAYS_ON_ACTIVATION,
    REQUIRED_WEIGHT_STABILITY, MAX_WEIGHT_DEVIATION, DEFAULT_BAUD_RATE,
    DEFAULT_READ_INTERVAL_MS, PREDEFINED_REGEXES, AVAILABLE_PRINT_PLACEHOLDERS,
    WINDOW_TITLE, DEFAULT_WINDOW_GEOMETRY, MIN_WINDOW_SIZE, DEFAULT_CONFIG,
    DEFAULT_BASE_WEIGHT, DEFAULT_BASE_PRICE, DEFAULT_INCREMENT_WEIGHT, DEFAULT_INCREMENT_PRICE
)
from messagebox import CustomMessageBox
from database import DatabaseManager
from serial_manager import SerialManager
from ui_components import UIHelper, TabManager, FormValidator, ProgressDialog, ConfirmDialog
from pdf_print_manager import PDFPrintManager


# Part 2: TruckScaleApp Class Initialization & Core Application Methods
class TruckScaleApp:
    def __init__(self, root: tk.Tk, managers: dict = None):
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry(DEFAULT_WINDOW_GEOMETRY)
        self.root.minsize(MIN_WINDOW_SIZE[0], MIN_WINDOW_SIZE[1])

        self.msg_box = CustomMessageBox(self.root)
        
        # Initialize managers - use passed managers if available, otherwise create new ones
        if managers:
            self.db_manager = managers.get('database', DatabaseManager(self.msg_box))
            self.serial_manager = managers.get('serial', SerialManager(self.msg_box))
            self.pdf_manager = managers.get('pdf_print', PDFPrintManager(self.msg_box))
        else:
            self.db_manager = DatabaseManager(self.msg_box)
            self.serial_manager = SerialManager(self.msg_box)
            self.pdf_manager = PDFPrintManager(self.msg_box)
        
        # --- Login and User State ---
        self.is_logged_in = False
        self.logged_in_user = ""
        self.current_user_role = ""
        self.login_timestamp = None
        self.admin_password = ADMIN_PASSWORD_HASH

        # Icon is now loaded at startup for immediate visual representation
        self.root.state('zoomed')

        self.style = ttk.Style()
        self.setup_ttk_styles()

        # Initialize weight_value to display real-time received weight
        self.weight_value = tk.StringVar(value="0.00")
        self.weight_source_status = tk.StringVar(value="No Source")
        
        self.recalled_record_no_var = tk.StringVar(value="")
        self.recalled_weight_var = tk.StringVar(value="")
        self.recalled_weight_date_var = tk.StringVar(value="")
        self.recalled_weight_time_var = tk.StringVar(value="")
        
        self.auto_time_var = tk.StringVar(value="N/A")
        # auto_weight_display_var will now reflect the real-time weight, similar to weight_value
        self.auto_weight_display_var = tk.StringVar(value="0.00") 

        # Use predefined regexes from config
        self.predefined_regexes = PREDEFINED_REGEXES
        self.data_format_regex_var = tk.StringVar()

        # Use available print placeholders from config
        self.available_print_placeholders = AVAILABLE_PRINT_PLACEHOLDERS
        
        # --- ONE WAY TEMPLATE ---
        self.default_print_template_one_way = (
            "Advantechnique\n"
            "advantechnique@gmail.com\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "    TICKET NO : {ticket_no}\n"
            "COMPANY       : {company}\n"
            "TRUCK PLATE   : {truck_plate}\n"
            "PRODUCT       : {product}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "GROSS WEIGHT  : {gross_weight} KG\n"
            "GROSS DATE    : {gross_date}\n"
            "GROSS TIME    : {gross_time}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "UNIT PRICE    : {unit_price} PER KG\n"
            "TOTAL PRICE   : {total_price}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "DATE PRINTED  : {date_printed}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "OPERATOR      : {logged_in_user}\n"
        )

        # --- TWO WAY TEMPLATE ---
        self.default_print_template_two_way = (
            "Advantechnique\n"
            "advantechnique@gmail.com\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "    TICKET NO : {ticket_no}\n"
            "COMPANY       : {company}\n"
            "TRUCK PLATE   : {truck_plate}\n"
            "PRODUCT       : {product}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "GROSS WEIGHT  : {gross_weight} KG\n"
            "GROSS DATE    : {gross_date}\n"
            "GROSS TIME    : {gross_time}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "TARE WEIGHT   : {tare_weight} KG\n"
            "TARE DATE     : {tare_date}\n"
            "TARE TIME     : {tare_time}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "NET WEIGHT    : {net_weight} KG\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "UNIT PRICE    : {unit_price} PER KG\n"
            "TOTAL PRICE   : {total_price}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "DATE PRINTED  : {date_printed}\n"
            "\n"  # Blank line
            "\n"  # Another blank line
            "OPERATOR      : {logged_in_user}\n"
        )
        
        self.one_way_print_template_var = tk.StringVar(value=self.default_print_template_one_way)
        # Initialize two_way_print_template_var with its default content
        self.two_way_print_template_var = tk.StringVar(value=self.default_print_template_two_way) 
        
        self.print_template_var = tk.StringVar() 
        self.current_template_editor_mode_var = tk.StringVar(value="ONE_WAY") 
        
        self.print_include_barcode_var = tk.BooleanVar()

        self.available_entry_fields = [
            ("Company", "company"),
            ("Truck Plate", "truck_plate"),
            ("Product", "product"),
            ("Designation", "designation"),
            ("Sender", "sender"),
            ("Origin", "origin"),
            ("Destination", "destination"),
            ("Driver", "driver"),
            ("Total Price", "total_price")
        ]
        self.default_entry_fields = [
            ("Company", "company"),
            ("Truck Plate", "truck_plate"),
            ("Product", "product"),
            ("Designation", "designation"),
            ("Sender", "sender"),
            ("Origin", "origin"),
            ("Destination", "destination"),
            ("Driver", "driver"),
            ("Total Price", "total_price")
        ]
        self.selected_entry_fields: list[tuple[str, str]] = []

        self.config = self.load_config()
        
        self.decimal_places = int(self.config.get("decimal_places", 0))
        # Initialize tkinter variables before using them
        self.read_loop_interval_ms = tk.IntVar(value=int(self.config.get("read_loop_interval_ms", 50)))
        self.max_weight_deviation = tk.DoubleVar(value=float(self.config.get("max_weight_deviation", 10.0)))
        self.print_line_spacing = tk.IntVar(value=self.config.get("print_line_spacing", 0))
        self.print_encoding_var = tk.StringVar(value=self.config.get("print_encoding", "utf-8"))
        
        # Load print templates from config, or use defaults
        # The variables will be populated from the database later
        self.one_way_print_template_var.set("")
        self.two_way_print_template_var.set("")
        
        self.print_include_barcode_var.set(self.config.get("print_include_barcode", False))

        self.monospace_fonts = ["Courier New", "Consolas", "Lucida Console", "Monaco", "monospace"] 
        self.print_template_font_family_var = tk.StringVar(value=self.config.get("print_template_font_family", "Courier New"))
        self.print_template_font_size_var = tk.IntVar(value=self.config.get("print_template_font_size", 12))
        self.print_template_font_bold_var = tk.BooleanVar(value=self.config.get("print_template_font_bold", True))

        self.report_search_query_var = tk.StringVar()
        # Initialize with current month date range instead of 30 days
        month_start, month_end = self._get_current_month_range()
        self.report_from_date_var = tk.StringVar(value=month_start.strftime("%Y-%m-%d"))
        self.report_to_date_var = tk.StringVar(value=month_end.strftime("%Y-%m-%d"))
        
        self.search_query_var = tk.StringVar()
        # Initialize with current month date range instead of 30 days
        self.from_date_var = tk.StringVar(value=month_start.strftime("%Y-%m-%d"))
        self.to_date_var = tk.StringVar(value=month_end.strftime("%Y-%m-%d"))

        self.big_display_port_var = tk.StringVar(value=self.config.get("big_display_port", ""))
        self.big_display_baud_var = tk.StringVar(value=self.config.get("big_display_baud", "9600"))
        self.big_display_data_format_var = tk.StringVar(value=self.config.get("big_display_data_format", "{weight:.2f} {tare_date} {tare_time}\\r\\n"))
        self.big_display_auto_connect_var = tk.BooleanVar(value=self.config.get("big_display_auto_connect", True))
        self.big_display_status_var = tk.StringVar(value="Big Display Status: Disconnected")
        self.big_display_connected = False
        self.big_display_serial_port = None

        self.report_tree = None
        self.all_report_transactions = []

        # Initialize emulator variables
        self._emulator_port_entry_var = tk.StringVar(value=self.config.get("emulator_port", ""))
        self._emulator_baud_entry_var = tk.StringVar(value=self.config.get("emulator_baud", "9600"))
        self._emulator_weight_to_send_var = tk.StringVar(value=self.config.get("emulator_default_weight", "123456"))
        self._emulator_send_interval_var = tk.StringVar(value=self.config.get("emulator_send_interval", "0.1"))
        self._emulator_auto_connect_var = tk.BooleanVar(value=self.config.get("emulator_auto_connect", False))
        self._emulator_status_var = tk.StringVar(value="Emulator Status: Disconnected")
        self._emulator_serial_port = None
        self._emulator_sending_data = False
        self.serial_running = False

        self.port_entry_var = tk.StringVar(value=self.config.get("port", ""))
        self.baud_entry_var = tk.StringVar(value=self.config.get("baud", "9600"))
        self.data_format_regex_var.set(self.config.get("data_format_regex", r"ww(-?\d+)"))
        self.auto_connect_var = tk.BooleanVar(value=self.config.get("auto_connect", False))
        self.auto_detect_regex_enabled_var = tk.BooleanVar(value=self.config.get("auto_detect_regex_enabled", False))

        # NEW: Printer selection and copies
        self.selected_printer_var = tk.StringVar(value=self.config.get("selected_printer", ""))
        self.print_copies_var = tk.IntVar(value=self.config.get("print_copies", 1))
        # OPTIMIZATION: Defer printer enumeration (can be slow on some systems)
        self.available_printers = []
        self._printers_loaded = False
        
        # NEW: Starting ticket number setting
        self.starting_ticket_number_var = tk.IntVar(value=self.config.get("starting_ticket_number", 1000))
        
        # Camera variables
        self.camera_device_var = tk.StringVar(value=self.config.get("camera_device", ""))
        self.camera_status_var = tk.StringVar(value="Camera Status: Disconnected")
        self.camera_manager = CameraManager(self.msg_box)
        self.camera_connected = False
        
        # New camera toggle variables
        self.camera_use_image_var = tk.BooleanVar(value=self.config.get("camera_use_image", False))
        self.camera_mirror_mode_var = tk.BooleanVar(value=self.config.get("camera_mirror_mode", False))
        self.camera_uploaded_image_path = self.config.get("camera_uploaded_image_path", "")
        self.camera_uploaded_image = None  # PIL Image object

        # NEW: PDF Page size and orientation
        self.pdf_page_sizes = ["A6"]
        self.pdf_page_size_var = tk.StringVar(value=self.config.get("pdf_page_size", "A6"))
        self.pdf_orientation_var = tk.StringVar(value=self.config.get("pdf_orientation", "Portrait"))

        self.weight_option_var = tk.StringVar(value=self.config.get("weight_option", "TWO_WAY"))
        
        # Initialize price computation enabled variable early to prevent AttributeError
        self.price_computation_enabled_var = tk.BooleanVar(value=self.config.get("price_computation_enabled", True))

        loaded_entry_fields = self.config.get("entry_form_fields", [])
        if loaded_entry_fields:
            valid_entry_keys = {key for _, key in self.available_entry_fields}
            self.selected_entry_fields = [
                (display, key) for display, key in loaded_entry_fields
                if key in valid_entry_keys
            ]
            if not self.selected_entry_fields:
                self.selected_entry_fields = list(self.default_entry_fields)
        else:
            self.selected_entry_fields = list(self.default_entry_fields)

        self.current_edit_transaction_id = None

        self.is_app_activated = False
        self.activation_status_var = tk.StringVar()
        self.activation_code_entry_var = tk.StringVar()
        
        # New variables for changing admin password
        self.new_admin_password_var = tk.StringVar()
        self.new_admin_password_confirm_var = tk.StringVar()

        self.registered_pdf_font = False
        self.pdf_font_name = "MonospaceFont"

        # --- New PDF Print Settings ---
        self.pdf_fonts = ["Helvetica", "Times-Roman", "Courier", "MonospaceFont"]
        self.pdf_print_template_font_family_var = tk.StringVar(value=self.config.get("pdf_print_template_font_family", "Helvetica"))
        self.pdf_print_template_font_size_var = tk.IntVar(value=self.config.get("pdf_print_template_font_size", 10))
        self.pdf_print_template_font_bold_var = tk.BooleanVar(value=self.config.get("pdf_print_template_font_bold", True))
        
        # The variables for the rulers and preview content
        self.pdf_preview_text = None
        self._pdf_preview_content = ""
        self._pdf_preview_frame = None

        self.rulers_canvas = None
        self.rulers_frame = None
        self.rulers_text = None
        self.ruler_x_scale = 1
        self.ruler_y_scale = 1
        self.ruler_unit_var = tk.StringVar(value="inch")
        
        # --- PERFORMANCE: Defer non-critical initialization ---
        # Track which tabs have been built (lazy loading)
        self._tabs_built = {
            "Entry Form": False,
            "Pending Records": False,
            "Completed Records": False,
            "Reports": False,
            "Settings": False
        }
        
        self._initialize_activation_state_internal()
        self.init_database()
        
        # NEW: Load templates from the database after the database is initialized
        self._load_print_templates_from_db()
        
        # Build the main GUI in background (disabled until login succeeds)
        self.main_notebook = None  # Will be created by build_gui()
        self.build_gui()
        self.root.attributes('-disabled', True)
        
        # Call the login screen first (on top of disabled GUI)
        self.show_login_screen()

        self._update_action_button_states()
        if "initial_trial_start_date" in self.config and not self.config.get("has_saved_initial_config", False):
            self.save_config_on_startup()
            self.config["has_saved_initial_config"] = True
            self.config = self.load_config()

        # OPTIMIZATION: Defer background tasks until after login
        self._deferred_startup_tasks_scheduled = False
        
        self.print_template_text = None 
        self.print_preview_canvas = None
        
        self.pdf_preview_canvas = None

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def _generate_coordinate_based_label(self, data):
        """
        Generate coordinate-based label content for PDF printing with proper spacing.
        """
        # Extract data with defaults
        ticket_no = data.get('ticket_no', '')
        company = data.get('company', '')
        truck_plate = data.get('truck_plate', '')
        product = data.get('product', '')
        gross_weight = data.get('gross_weight', '')
        gross_date = data.get('gross_date', '')
        gross_time = data.get('gross_time', '')
        tare_weight = data.get('tare_weight', 'N/A')
        tare_date = data.get('tare_date', 'None')
        tare_time = data.get('tare_time', 'None')
        net_weight = data.get('net_weight', 'N/A')
        date_printed = data.get('date_printed', '')
        
        # Create coordinate-based layout with proper spacing
        lines = []
        
        # Header section: choose per-transaction-type header template if available
        default_tpl = "Company: {company}\nTicket No: {ticket_no}\nDate: {date_printed}\n"
        weight_type_raw = data.get('weight_type', '') or data.get('transaction_type', '')
        wt = str(weight_type_raw).lower()
        if 'two' in wt or '2-way' in wt or 'two-way' in wt or '2way' in wt:
            header_tpl = self.config.get("ticket_header_two_way", self.config.get("ticket_header_template", default_tpl))
        else:
            header_tpl = self.config.get("ticket_header_one_way", self.config.get("ticket_header_template", default_tpl))
        mapping = {
            'ticket_no': ticket_no,
            'company': company,
            'truck_plate': truck_plate,
            'product': product,
            'gross_weight': gross_weight,
            'gross_date': gross_date,
            'gross_time': gross_time,
            'tare_weight': tare_weight,
            'tare_date': tare_date,
            'tare_time': tare_time,
            'net_weight': net_weight,
            'date_printed': date_printed
        }
        class _SafeDict(dict):
            def __missing__(self, key):
                return ''

        try:
            formatted_header = header_tpl.format_map(_SafeDict(mapping))
            for hl in formatted_header.splitlines():
                lines.append(hl)
        except Exception:
            # Fallback to a simple default header on formatting errors
            lines.append(f"COMPANY       : {company}")
            lines.append(f"TICKET NO     : {ticket_no}")
        lines.append("")  # Empty line
        lines.append("")  # Empty line
        
        # Main content with proper spacing
        lines.append(f"    TICKET NO : {ticket_no}")
        lines.append(f"COMPANY       : {company}")
        lines.append(f"TRUCK PLATE   : {truck_plate}")
        lines.append(f"PRODUCT       : {product}")
        lines.append("")  # Empty line
        lines.append("")  # Empty line
        
        # Weight and date info
        lines.append(f"GROSS WEIGHT  : {gross_weight} KG")
        lines.append(f"GROSS DATE    : {gross_date}")
        lines.append(f"GROSS TIME    : {gross_time}")
        lines.append("")  # Empty line
        lines.append("")  # Empty line
        
        # Tare information
        lines.append(f"TARE WEIGHT   : {tare_weight} KG")
        lines.append(f"TARE DATE     : {tare_date}")
        lines.append(f"TARE TIME     : {tare_time}")
        lines.append("")  # Empty line
        lines.append("")  # Empty line
        
        # Net weight
        lines.append(f"NET WEIGHT    : {net_weight} KG")
        lines.append("")  # Empty line
        lines.append("")  # Empty line
        
        # Final info
        lines.append(f"DATE PRINTED  : {date_printed}")
        lines.append("")  # Empty line
        lines.append("")  # Empty line
        
        # Operator signature
        lines.append("OPERATOR      : ___________________")
        
        return "\n".join(lines)
    
    def _add_placeholder_dialog(self):
        """Opens a dialog to let the user define a new placeholder."""
        # Create a simple dialog window
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New Placeholder")
        dialog.geometry("300x150")
        dialog.transient(self.root)  # Stay on top of main window
        dialog.grab_set()  # Prevent interaction with main window

        # Label and entry for the placeholder name
        tk.Label(dialog, text="Placeholder Name (e.g., {new_field}):").pack(pady=5)
        name_entry = tk.Entry(dialog, width=30)
        name_entry.pack(pady=5)

        # Label and entry for the description (optional)
        tk.Label(dialog, text="Description (optional):").pack(pady=5)
        desc_entry = tk.Entry(dialog, width=30)
        desc_entry.pack(pady=5)

        # OK and Cancel buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)

        def on_ok():
            placeholder_name = name_entry.get().strip()
            if not placeholder_name or not placeholder_name.startswith('{') or not placeholder_name.endswith('}'):
                self.msg_box.showwarning("Invalid Input", "Please enter a valid placeholder in the format {name}.")
                return
            # Add to the common placeholders list (this is a simplification)
            # In practice, you might store this in a config file or database
            self.available_print_placeholders.append((placeholder_name[1:-1], placeholder_name))
            self._populate_placeholders_list()  # Refresh the listbox
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        ttk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side=tk.RIGHT, padx=5)

        # Focus on the entry field
        name_entry.focus()

    def _remove_placeholder(self):
        """Removes the currently selected placeholder from the listbox."""
        selected_indices = self.placeholder_listbox.curselection()
        if not selected_indices:
            self.msg_box.showwarning("No Selection", "Please select a placeholder to remove.")
            return

        selected_text = self.placeholder_listbox.get(selected_indices[0])
        # Extract the actual placeholder name (without brackets)
        placeholder_match = re.search(r"\[(.*)\]", selected_text)
        if placeholder_match:
            placeholder_name = placeholder_match.group(1)
            # Remove from the available placeholders list
            self.available_print_placeholders = [p for p in self.available_print_placeholders if p[0] != placeholder_name]
            self._populate_placeholders_list()  # Refresh the listbox
            self.msg_box.showinfo("Removed", f"Placeholder '{placeholder_name}' removed.")

    def _move_placeholder_up(self):
        """Moves the selected placeholder up in the listbox."""
        selected_indices = self.placeholder_listbox.curselection()
        if not selected_indices or selected_indices[0] == 0:
            self.msg_box.showwarning("Cannot Move", "Please select a placeholder that is not at the top.")
            return

        index = selected_indices[0]
        # Get the item at the current index and the one above it
        item_to_move = self.placeholder_listbox.get(index)
        item_above = self.placeholder_listbox.get(index - 1)
        # Remove both items
        self.placeholder_listbox.delete(index)
        self.placeholder_listbox.delete(index - 1)
        # Insert them in the new order
        self.placeholder_listbox.insert(index - 1, item_to_move)
        self.placeholder_listbox.insert(index, item_above)
        # Select the moved item
        self.placeholder_listbox.selection_clear(0, tk.END)
        self.placeholder_listbox.selection_set(index - 1)

    def _move_placeholder_down(self):
        """Moves the selected placeholder down in the listbox."""
        selected_indices = self.placeholder_listbox.curselection()
        if not selected_indices or selected_indices[0] == self.placeholder_listbox.size() - 1:
            self.msg_box.showwarning("Cannot Move", "Please select a placeholder that is not at the bottom.")
            return

        index = selected_indices[0]
        # Get the item at the current index and the one below it
        item_to_move = self.placeholder_listbox.get(index)
        item_below = self.placeholder_listbox.get(index + 1)
        # Remove both items
        self.placeholder_listbox.delete(index)
        self.placeholder_listbox.delete(index + 1)
        # Insert them in the new order
        self.placeholder_listbox.insert(index, item_below)
        self.placeholder_listbox.insert(index + 1, item_to_move)
        # Select the moved item
        self.placeholder_listbox.selection_clear(0, tk.END)
        self.placeholder_listbox.selection_set(index + 1)
        
# Part 3: Core Application Methods

    def _update_weight_display_and_status(self, weight_str: str, source_status: str, status_color: str, clear_recalled: bool = True):
        """
        Helper function to update the main weight display and its status text/color.
        """
        self.weight_value.set(weight_str)
        self.weight_source_status.set(source_status)
        self.weight_source_status_label.config(foreground=status_color)
        self.auto_weight_display_var.set(weight_str) # Keep this line to update the auto display

        if clear_recalled:
            self._reset_recalled_details()

    def show_login_screen(self):
        """
        Displays a modern, visually pleasing login screen with robust error handling and validation.
        The main window is disabled in the background.
        """
        self.login_window = tk.Toplevel(self.root)
        self.login_window.title("Truck Scale Weighing System - Login")
        self.login_window.transient(self.root)
        self.login_window.resizable(False, False)
        self.login_window.grab_set()

        # Set a close protocol that closes the application if the login window is closed
        self.login_window.protocol("WM_DELETE_WINDOW", self.root.destroy)

        # Define modern color scheme
        PRIMARY_COLOR = "#1e88e5"  # Modern blue
        SECONDARY_COLOR = "#e0e0e0"  # Medium gray (darkened from #f5f5f5)
        TEXT_COLOR = "#212121"  # Dark text
        ERROR_COLOR = "#d32f2f"  # Red for errors
        SUCCESS_COLOR = "#388e3c"  # Green for success
        BORDER_COLOR = "#bdbdbd"  # Border gray

        # Configure login window background
        self.login_window.configure(bg=SECONDARY_COLOR)

        # Create main container with padding
        main_container = tk.Frame(self.login_window, bg=SECONDARY_COLOR)
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # Header section with title and application name
        header_frame = tk.Frame(main_container, bg=PRIMARY_COLOR, height=140)
        header_frame.pack(fill=tk.X, padx=10, pady=(15, 5))  # Add vertical padding
        header_frame.pack_propagate(False)

        title_label = tk.Label(
            header_frame,
            text="TRUCK SCALE",
            font=("Segoe UI", 28, "bold"),
            bg=PRIMARY_COLOR,
            fg="white"
        )
        title_label.pack(pady=(15, 5))

        subtitle_label = tk.Label(
            header_frame,
            text="Weighing System - User Authentication",
            font=("Segoe UI", 11, "normal"),
            bg=PRIMARY_COLOR,
            fg="#E8F4F8"  # Light cyan for better contrast
        )
        subtitle_label.pack(pady=(5, 20))

        # Main login content frame
        content_frame = tk.Frame(main_container, bg=SECONDARY_COLOR)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=35)

        # Username label and dropdown
        username_label = tk.Label(
            content_frame,
            text="Username",
            font=("Segoe UI", 11, "bold"),
            bg=SECONDARY_COLOR,
            fg=TEXT_COLOR
        )
        username_label.pack(anchor="w", pady=(0, 5))

        # Fetch available users from database
        try:
            available_users = []
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT username FROM users ORDER BY username")
                users = cursor.fetchall()
                available_users = [user[0] for user in users]
        except Exception as e:
            logging.error(f"Error fetching users: {e}")
            available_users = ["admin"]  # Fallback to admin

        username_var = tk.StringVar()
        username_dropdown = ttk.Combobox(
            content_frame,
            textvariable=username_var,
            values=available_users,
            font=("Segoe UI", 10),
            width=30,
            state="readonly"
        )
        username_dropdown.pack(fill=tk.X, pady=(0, 20))
        
        # Set first user as default if available
        if available_users:
            username_dropdown.set(available_users[0])

        # Password label
        password_label = tk.Label(
            content_frame,
            text="Password",
            font=("Segoe UI", 11, "bold"),
            bg=SECONDARY_COLOR,
            fg=TEXT_COLOR
        )
        password_label.pack(anchor="w", pady=(0, 5))

        # Password entry with better styling
        password_entry = tk.Entry(
            content_frame,
            show="●",
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            bd=1,
            bg="white",
            fg=TEXT_COLOR,
            insertbackground=PRIMARY_COLOR
        )
        password_entry.pack(fill=tk.X, pady=(0, 5))
        password_entry.focus_set()

        # Add subtle border effect to password entry
        def on_password_focus_in(event):
            password_entry.configure(relief=tk.SOLID, bd=2)
        
        def on_password_focus_out(event):
            password_entry.configure(relief=tk.FLAT, bd=1)

        password_entry.bind("<FocusIn>", on_password_focus_in)
        password_entry.bind("<FocusOut>", on_password_focus_out)

        # Status/Error message display
        status_var = tk.StringVar(value="")
        status_label = tk.Label(
            content_frame,
            textvariable=status_var,
            font=("Segoe UI", 9),
            bg=SECONDARY_COLOR,
            fg=ERROR_COLOR,
            wraplength=300,
            justify=tk.LEFT
        )
        status_label.pack(anchor="w", pady=(0, 15), fill=tk.X)

        # Login attempt counter for security
        login_attempts = {"count": 0, "max": 5}
        max_attempts_warning = tk.Label(
            content_frame,
            text="",
            font=("Segoe UI", 9),
            bg=SECONDARY_COLOR,
            fg=ERROR_COLOR
        )
        max_attempts_warning.pack(anchor="w", pady=(0, 15))

        def perform_login():
            """Perform login with robust validation and error handling"""
            # Check if account is locked
            if login_attempts["count"] >= login_attempts["max"]:
                status_var.set("Account temporarily locked. Too many failed attempts.")
                max_attempts_warning.config(text=f"Attempts remaining: 0/{login_attempts['max']}")
                password_entry.config(state=tk.DISABLED)
                username_dropdown.config(state=tk.DISABLED)
                login_button.config(state=tk.DISABLED)
                return

            username = username_var.get().strip()
            password = password_entry.get()

            # Input validation
            if not username:
                status_var.set("⚠ Please select a username")
                password_entry.delete(0, tk.END)
                password_entry.focus_set()
                return

            if not password:
                status_var.set("⚠ Please enter your password")
                password_entry.focus_set()
                return

            if len(password) < 4:
                status_var.set("⚠ Password appears to be invalid")
                password_entry.delete(0, tk.END)
                password_entry.focus_set()
                return

            # Hash the password
            try:
                password_hash = hashlib.sha256(password.encode()).hexdigest()
            except Exception as e:
                logging.error(f"Error hashing password: {e}")
                status_var.set("⚠ Error processing password. Try again.")
                password_entry.delete(0, tk.END)
                password_entry.focus_set()
                return

            # Verify credentials against database
            try:
                with sqlite3.connect(DB_FILE) as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT password_hash, role FROM users WHERE username = ? LIMIT 1",
                        (username,)
                    )
                    result = cursor.fetchone()

                    if result and result[0] == password_hash:
                        # Successful login - clear status and proceed
                        status_var.set("✓ Login successful! Please wait...")
                        status_label.config(fg=SUCCESS_COLOR)
                        login_button.config(state=tk.DISABLED)
                        password_entry.config(state=tk.DISABLED)
                        username_dropdown.config(state=tk.DISABLED)

                        # Store login information
                        self.is_logged_in = True
                        self.logged_in_user = username
                        self.current_user_role = result[1]  # Store user's role
                        self.login_timestamp = datetime.now()

                        # Update admin password if logged in as admin
                        if username == ADMIN_USERNAME:
                            self.admin_password = password_hash

                        # Schedule window destruction and app re-enabling
                        self.login_window.after(500, self._finalize_login)

                    else:
                        # Failed login
                        login_attempts["count"] += 1
                        remaining = login_attempts["max"] - login_attempts["count"]
                        
                        if remaining <= 0:
                            status_var.set("✗ Account locked due to too many failed attempts")
                            max_attempts_warning.config(text=f"Attempts remaining: 0/{login_attempts['max']}")
                            password_entry.config(state=tk.DISABLED)
                            username_dropdown.config(state=tk.DISABLED)
                            login_button.config(state=tk.DISABLED)
                        else:
                            status_var.set(f"✗ Invalid username or password")
                            max_attempts_warning.config(text=f"Attempts remaining: {remaining}/{login_attempts['max']}")
                        
                        password_entry.delete(0, tk.END)
                        password_entry.focus_set()

            except sqlite3.Error as e:
                logging.error(f"Database error during login: {e}")
                status_var.set("⚠ Database connection error. Try again.")
                password_entry.delete(0, tk.END)
                password_entry.focus_set()
            except Exception as e:
                logging.error(f"Unexpected error during login: {e}")
                status_var.set("⚠ An unexpected error occurred. Try again.")
                password_entry.delete(0, tk.END)
                password_entry.focus_set()

        # Button frame with proper spacing
        button_frame = tk.Frame(content_frame, bg=SECONDARY_COLOR)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        login_button = tk.Button(
            button_frame,
            text="LOGIN",
            font=("Segoe UI", 11, "bold"),
            bg=PRIMARY_COLOR,
            fg="white",
            relief=tk.FLAT,
            padx=30,
            pady=10,
            cursor="hand2",
            command=perform_login,
            activebackground="#1565c0"
        )
        login_button.pack(side=tk.LEFT, fill=tk.X, expand=True)

        exit_button = tk.Button(
            button_frame,
            text="EXIT",
            font=("Segoe UI", 11, "bold"),
            bg="#757575",
            fg="white",
            relief=tk.FLAT,
            padx=30,
            pady=10,
            cursor="hand2",
            command=self.root.destroy,
            activebackground="#616161"
        )
        exit_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

        # Bind Enter key to login
        password_entry.bind('<Return>', lambda event: perform_login())
        username_dropdown.bind('<Return>', lambda event: perform_login())

        # Center the login window on screen
        self.login_window.update_idletasks()
        screen_width = self.login_window.winfo_screenwidth()
        screen_height = self.login_window.winfo_screenheight()
        window_width = 450
        window_height = 480
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.login_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    def _finalize_login(self):
        """Complete login process and enable main application"""
        try:
            self.login_window.destroy()
            
            # Re-enable the main window after successful login
            self.root.attributes('-disabled', False)
            
            # Set window to fullscreen after login
            self.root.state('zoomed')
            
            # Apply role-based access control immediately after login
            self._apply_role_based_access_control()
            
            # OPTIMIZATION: Schedule deferred startup tasks after login
            if not self._deferred_startup_tasks_scheduled:
                self._deferred_startup_tasks_scheduled = True
                self.root.after(100, self._schedule_deferred_startup_tasks)
            
            # Go to Entry Form tab
            if self.main_notebook:
                self.main_notebook.select(0)
        except Exception as e:
            logging.error(f"Error finalizing login: {e}")
            self.msg_box.showerror("Error", f"Error completing login: {e}")

    def logout(self):
        """
        Logs the current user out, disables the main GUI, and shows the login screen again.
        """
        self.is_logged_in = False
        self.logged_in_user = ""
        self.current_user_role = ""
        self.login_timestamp = None
        
        # Update price computation settings to disabled state
        self._update_price_settings_state()
        
        self.root.attributes('-disabled', True)
        self.show_login_screen()

    def _schedule_deferred_startup_tasks(self):
        """
        OPTIMIZATION: Schedule tasks that don't need to happen immediately after login.
        This improves perceived startup performance by deferring heavy operations.
        """
        # Icon already loaded at startup
        
        # Load printers in background
        if not self._printers_loaded:
            self._load_available_printers()
            self._printers_loaded = True
        
        # Attempt auto-connections after a short delay
        self.root.after(500, self._attempt_auto_connections)
        
        # Start auto-refresh UI elements in background thread
        threading.Thread(target=self._auto_refresh_ui_elements, daemon=True).start()

    def _set_app_icon_immediately(self):
        """
        Set app icon IMMEDIATELY at startup (critical timing).
        Must be called right after tk.Tk() creation and before any other UI operations.
        Dynamically locates the icon in the assets folder relative to the application.
        Sets icon for external (taskbar/shortcut), internal (buttons/dialogs), and tk window.
        """
        try:
            icon_filename = "app_icon.ico"
            # Try multiple possible locations for the icon
            possible_paths = [
                # Hardcoded absolute path (most reliable for development)
                r"C:\Users\Administrator\Desktop\weighing_scale_project\assets\app_icon.ico",
                # Path when running from source
                os.path.join(os.path.dirname(__file__), "assets", icon_filename),
                # Fallback: absolute path for current user
                os.path.join(os.path.expanduser("~"), "Desktop", "weighing_scale_project", "assets", icon_filename),
                # Path when running from PyInstaller executable
                os.path.join(os.path.dirname(sys.executable), "assets", icon_filename),
            ]
            
            icon_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    icon_path = path
                    print(f"[ICON] Found at: {icon_path}")
                    break
            
            if not icon_path:
                print(f"[ICON] Warning: Icon not found in any location")
                return
            
            # Store icon path for external use
            self.icon_path = icon_path
            
            # Try to set window icon for tkinter (external display - taskbar, title bar)
            try:
                # Use the absolute path directly
                self.root.iconbitmap(default=icon_path)
                print(f"[ICON] Successfully set via iconbitmap: {icon_path}")
            except tk.TclError as e:
                print(f"[ICON] iconbitmap failed: {e}, trying iconphoto...")
                try:
                    # PhotoImage can handle ICO files in some cases
                    self.app_icon_photo = tk.PhotoImage(file=icon_path)
                    self.root.iconphoto(True, self.app_icon_photo)
                    print(f"[ICON] Successfully set via iconphoto")
                except Exception as e2:
                    print(f"[ICON] iconphoto also failed: {e2}")
                    # Try to load PNG or convert ICO - as fallback
                    self._try_alternative_icon_format()
            except Exception as e:
                print(f"[ICON] Unexpected error setting icon: {e}")
                self._try_alternative_icon_format()
                
        except Exception as e:
            print(f"[ICON] Fatal error in _set_app_icon_immediately: {e}")
            import traceback
            traceback.print_exc()

    def _try_alternative_icon_format(self):
        """
        Try to load icon in alternative formats if ICO fails.
        """
        try:
            assets_path = r"C:\Users\Administrator\Desktop\weighing_scale_project\assets"
            # Try to find PNG or other formats
            for filename in ["app_icon.png", "icon.png", "app_icon.gif", "icon.gif"]:
                alt_path = os.path.join(assets_path, filename)
                if os.path.exists(alt_path):
                    try:
                        self.app_icon_photo = tk.PhotoImage(file=alt_path)
                        self.root.iconphoto(True, self.app_icon_photo)
                        print(f"[ICON] Loaded alternative format: {alt_path}")
                        return
                    except Exception as e:
                        print(f"[ICON] Failed to load {filename}: {e}")
        except Exception as e:
            print(f"[ICON] Error trying alternative formats: {e}")

    def get_icon_path(self):
        """
        Get the icon path for external use (shortcuts, external references, taskbar).
        Returns the full path to the app icon, or None if not found.
        """
        return self.icon_path

    def get_icon_photo(self):
        """
        Get the internal PhotoImage icon for use in dialogs and buttons.
        Returns the PhotoImage object, or None if not available.
        """
        return self.app_icon_photo

    def _load_available_printers(self):
        """
        OPTIMIZATION: Load available printers asynchronously (deferred from startup).
        Printer enumeration can be slow on some systems.
        """
        try:
            if sys.platform == "win32" and PRINTING_ENABLED:
                self.available_printers = [printer[2] for printer in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL)]
                
                # Auto-select default printer if none is configured
                if not self.selected_printer_var.get():
                    try:
                        self.selected_printer_var.set(win32print.GetDefaultPrinter())
                    except Exception:
                        pass
        except Exception:
            pass

    def setup_ttk_styles(self):
        default_font = ("Helvetica", 10)
        bold_font = ("Helvetica", 10, "bold")
        heading_font = ("Helvetica", 11, "bold")

        self.style.theme_use("clam")

        self.style.configure("TFrame", background="#f0f0f0")
        self.style.configure("TLabel", font=default_font, background="#f0f0f0", foreground="#333333")
        self.style.configure("TEntry", font=default_font, fieldbackground="white", foreground="#333333")
        self.style.configure("TCombobox", font=default_font, fieldbackground="white", foreground="#333333")
        self.style.configure("TSpinbox", font=default_font, fieldbackground="white", foreground="#333333")
        self.style.configure("TCheckbutton", font=default_font, background="#f0f0f0", foreground="#333333")
        self.style.configure("TRadiobutton",
                             font=default_font,
                             background="#f0f0f0",
                             foreground="#333333")

        self.style.configure("TButton",
                             font=bold_font,
                             background="#4CAF50",
                             foreground="white",
                             relief="flat",
                             padding=6)
        self.style.map("TButton",
                       background=[('active', '#45a049'), ('disabled', '#cccccc')],
                       foreground=[('disabled', '#666666')])

        self.style.configure("TLabelframe", background="#f0f0f0", relief="solid", borderwidth=1, borderradius=5)
        self.style.configure("TLabelframe.Label", font=heading_font, background="#f0f0f0", foreground="#2c3e50")

        self.style.configure("TNotebook", background="#e0e0e0", borderwidth=0)
        self.style.configure("TNotebook.Tab",
                             font=bold_font,
                             background="#d0d0d0",
                             foreground="#555555",
                             padding=[10, 5])
        self.style.map("TNotebook.Tab",
                       background=[('selected', '#e0e0e0'), ('active', '#c0c0c0')],
                       foreground=[('selected', '#333333')])
        
        self.style.configure("Treeview.Heading",
                             font=heading_font,
                             background="#34495e",
                             foreground="white",
                             relief="flat",
                             padding=[5, 5])
        self.style.map("Treeview.Heading",
                       background=[('active', '#FFFF00')],
                       foreground=[('active', 'black')])
        self.style.configure("Treeview",
                             font=default_font,
                             background="white",
                             foreground="#333333",
                             fieldbackground="white",
                             rowheight=25)
        self.style.map('Treeview',
                       background=[('selected', '#3498db')],
                       foreground=[('selected', 'white')])

        self.style.configure("Vertical.TScrollbar",
                             background="#cccccc",
                             troughcolor="#f0f0f0",
                             bordercolor="#cccccc",
                             arrowcolor="#666666")
        self.style.map("Vertical.TScrollbar",
                       background=[('active', '#aaaaaa')],
                       arrowcolor=[('active', '#333333')])

    def init_database(self):
        """
        Initializes the database using the DatabaseManager.
        """
        try:
            self.db_manager.init_database()
        except Exception as e:
            self.msg_box.showerror("Database Error", f"Failed to initialize database: {e}. Application will exit.")
            self.root.destroy()

    def _load_print_templates_from_db(self):
        """
        Loads the print templates from the database into StringVar variables.
        Uses defaults if no templates are found in database.
        """
        try:
            templates = self.db_manager.load_print_templates()
            # Use database templates if available, otherwise use defaults
            one_way_template = templates.get("ONE_WAY", self.default_print_template_one_way)
            two_way_template = templates.get("TWO_WAY", self.default_print_template_two_way)
            
            self.one_way_print_template_var.set(one_way_template)
            self.two_way_print_template_var.set(two_way_template)
        except Exception as e:
            logging.error(f"Error loading print templates: {e}")
            # Use defaults on error
            self.one_way_print_template_var.set(self.default_print_template_one_way)
            self.two_way_print_template_var.set(self.default_print_template_two_way)

    def load_config(self) -> dict:
        """
        Loads application configuration from JSON file.
        Also loads admin password from config if it's set.
        """
        import json
        
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    config = json.load(f)
            else:
                config = DEFAULT_CONFIG.copy()
                
            # Ensure all required keys exist
            for key, value in DEFAULT_CONFIG.items():
                if key not in config:
                    config[key] = value
                    
            return config
        except Exception as e:
            logging.error(f"Error loading config: {e}")
            return DEFAULT_CONFIG.copy()

    def save_config(self):
        """
        Saves the current application configuration to the JSON file.
        Note: The print templates are now saved separately to the database.
        """
        if not hasattr(self, 'port_entry'):
            return

        config_to_save = {
            "port": self.port_entry_var.get(),
            "baud": self.baud_entry_var.get(),
            "auto_connect": self.auto_connect_var.get(),
            "data_format_regex": self.data_format_regex_var.get(),
            "auto_detect_regex_enabled": self.auto_detect_regex_enabled_var.get(),
            "decimal_places": self.decimal_places_var.get(),
            "read_loop_interval_ms": self.read_loop_interval_ms.get(),
            "emulator_port": self._emulator_port_entry_var.get(),
            "emulator_baud": self._emulator_baud_entry_var.get(),
            "emulator_default_weight": self._emulator_weight_to_send_var.get(),
            "emulator_send_interval": self._emulator_send_interval_var.get(),
            "big_display_port": self.big_display_port_var.get(),
            "big_display_baud": self.big_display_baud_var.get(),
            "big_display_data_format": self.big_display_data_format_var.get(),
            "big_display_auto_connect": self.big_display_auto_connect_var.get(),
            "print_include_barcode": self.print_include_barcode_var.get(),
            "print_line_spacing": self.print_line_spacing.get(),
            "print_encoding": self.print_encoding_var.get(),
            "print_template_font_family": self.print_template_font_family_var.get(),
            "print_template_font_size": self.print_template_font_size_var.get(),
            "print_template_font_bold": self.print_template_font_bold_var.get(),
            "pdf_print_template_font_family": self.pdf_print_template_font_family_var.get(),
            "pdf_print_template_font_size": self.pdf_print_template_font_size_var.get(),
            "pdf_print_template_font_bold": self.pdf_print_template_font_bold_var.get(),
            "entry_form_fields": self.selected_entry_fields,
            "initial_trial_start_date": self.config.get("initial_trial_start_date"),
            "expiry_date": self.config.get("expiry_date"),
            "is_activated": self.config.get("is_activated", False),
            "has_saved_initial_config": self.config.get("has_saved_initial_config", False),
            "admin_password_hash": self.admin_password,
            "selected_printer": self.selected_printer_var.get(),
            "print_copies": self.print_copies_var.get(),
            "pdf_page_size": self.pdf_page_size_var.get(),
            "pdf_orientation": self.pdf_orientation_var.get(),
            "starting_ticket_number": self.starting_ticket_number_var.get(),
            "camera_device": self.camera_device_var.get(),
            "camera_auto_connect": self.camera_auto_connect_var.get(),
            "camera_use_image": self.camera_use_image_var.get(),
            "camera_mirror_mode": self.camera_mirror_mode_var.get(),
            "camera_uploaded_image_path": self.camera_uploaded_image_path,
            "weight_option": self.weight_option_var.get(),
            "ticket_header_font_size_one_way": self.config.get("ticket_header_font_size_one_way", "10"),
            "ticket_page_size_one_way": self.config.get("ticket_page_size_one_way", "A6"),
            "ticket_page_width_mm_one_way": self.config.get("ticket_page_width_mm_one_way", "105"),
            "ticket_page_height_mm_one_way": self.config.get("ticket_page_height_mm_one_way", "148"),
            "ticket_header_font_size_two_way": self.config.get("ticket_header_font_size_two_way", "10"),
            "ticket_page_size_two_way": self.config.get("ticket_page_size_two_way", "A6"),
            "ticket_page_width_mm_two_way": self.config.get("ticket_page_width_mm_two_way", "105"),
            "ticket_page_height_mm_two_way": self.config.get("ticket_page_height_mm_two_way", "148"),
            "one_way_optional_fields": self.config.get("one_way_optional_fields", []),
            "two_way_optional_fields": self.config.get("two_way_optional_fields", []),
            "pending_column_widths": self.config.get("pending_column_widths", {}),
            "transaction_column_widths": self.config.get("transaction_column_widths", {}),
            "report_column_widths": self.config.get("report_column_widths", {})
        }
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump(config_to_save, f, indent=4)

            try:
                new_decimal_places = int(self.decimal_places_var.get())
                if 0 <= new_decimal_places <= 3:
                    self.decimal_places = new_decimal_places
                else:
                    raise ValueError("Decimal places must be between 0 and 3.")
            except ValueError:
                self.decimal_places = 0
                self.msg_box.showwarning("Invalid Input", "Invalid decimal places value. It must be a number between 0 and 3. Defaulting to 0.")

        except Exception as e:
            self.msg_box.showerror("Save Error", f"Failed to save settings: {e}")

    def save_config_on_startup(self):
        """
        Saves default configuration settings to the config file if it's a new installation.
        This prevents the app from resetting trial dates on every startup.
        """
        try:
            current_config = {}
            if os.path.exists(CONFIG_FILE):
                try:
                    with open(CONFIG_FILE, "r") as f:
                        current_config = json.load(f)
                except json.JSONDecodeError:
                    current_config = {}

            current_config["initial_trial_start_date"] = self.config.get("initial_trial_start_date")
            current_config["expiry_date"] = self.config.get("expiry_date")
            current_config["is_activated"] = self.config.get("is_activated", False)
            current_config["has_saved_initial_config"] = False
            
            # Ensure all default keys are present in the new config file
            current_config.setdefault("print_include_barcode", False)
            current_config.setdefault("big_display_port", "")
            current_config.setdefault("big_display_baud", "9600")
            current_config.setdefault("big_display_data_format", "{weight:.2f} {tare_date} {tare_time}\\r\\n")
            current_config.setdefault("data_format_regex", r"(\d+\.\d+)")
            current_config.setdefault("auto_detect_regex_enabled", False)
            current_config.setdefault("print_encoding", "utf-8")
            current_config.setdefault("print_template_font_family", "Courier New")
            current_config.setdefault("print_template_font_size", 10)
            current_config.setdefault("print_template_font_bold", False)
            current_config.setdefault("pdf_print_template_font_family", "Helvetica")
            current_config.setdefault("pdf_print_template_font_size", 13)
            current_config.setdefault("pdf_print_template_font_bold", False)
            current_config.setdefault("read_loop_interval_ms", 10) # Set default to 10
            current_config.setdefault("admin_password_hash", self.admin_password)
            # NEW: Set default values for new print settings on startup
            current_config.setdefault("selected_printer", "")
            current_config.setdefault("print_copies", 1)
            current_config.setdefault("pdf_page_size", "Half Letter") # Change default to Half Letter
            current_config.setdefault("pdf_orientation", "Portrait")
            # NEW: Set default values for ticket format settings
            current_config.setdefault("ticket_header_font_size_one_way", "13")
            current_config.setdefault("ticket_page_size_one_way", "Custom")
            current_config.setdefault("ticket_page_width_mm_one_way", "105")
            current_config.setdefault("ticket_page_height_mm_one_way", "148")
            current_config.setdefault("ticket_header_font_size_two_way", "13")
            current_config.setdefault("ticket_page_size_two_way", "Custom")
            current_config.setdefault("ticket_page_width_mm_two_way", "105")
            current_config.setdefault("ticket_page_height_mm_two_way", "148")
            # NEW: Set default values for optional fields (initially empty)
            current_config.setdefault("one_way_optional_fields", [])
            current_config.setdefault("two_way_optional_fields", [])


            with open(CONFIG_FILE, "w") as f:
                json.dump(current_config, f, indent=4)
            self.config.update(current_config)
        except Exception as e:
            # Handle the case where saving fails, but don't crash the app
            pass

    def load_master_data_lists(self):
        """
        Loads the master data lists (companies, trucks, etc.) into the respective listboxes.
        """
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                if hasattr(self, 'master_entries'):
                    for category, data in self.master_entries.items():
                        listbox = data['listbox']
                        listbox.delete(0, tk.END)
                        cursor.execute(f"SELECT name FROM {category} ORDER BY name COLLATE NOCASE ASC")
                        for row in cursor.fetchall():
                            listbox.insert(tk.END, row[0])
        except sqlite3.Error as e:
            # Don't show an error here, as the master data is a secondary feature.
            # Errors will be handled by the main application flow if the database is critical.
            pass

    def populate_master_entry(self, event: tk.Event, category: str):
        """
        Populates the entry field with the selected value from a master data listbox.
        """
        listbox = self.master_entries[category]['listbox']
        selection = listbox.curselection()
        if selection:
            selected_value = listbox.get(selection[0])
            # self.master_entries[category]['entry_var'].set(selected_value)

    def add_master(self, category: str):
        """
        Adds a new entry to a master data table.
        """
        value = self.master_entries[category]['entry_var'].get().strip()
        if not value:
            self.msg_box.showerror("Error", "Cannot add empty value.")
            return
        
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                # Check for case-insensitive duplicates
                cursor.execute(f"SELECT name FROM {category} WHERE name = ? COLLATE NOCASE", (value,))
                if cursor.fetchone() is not None:
                    self.msg_box.showerror("Error", f"'{value}' already exists (case-insensitively) in {category}.")
                    return

                conn.execute(f"INSERT INTO {category} (name) VALUES (?)", (value,))
                conn.commit()
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to add to {category}: {e}")
        except Exception as e:
            self.msg_box.showerror("Error", f"An unexpected error occurred while adding to {category}: {e}")
        finally:
            self.load_master_data_lists()

    def delete_master(self, category: str):
        """
        Deletes an entry from a master data table, but only if it's not referenced in existing transactions.
        Only allowed for admin and operator roles.
        """
        # Check user permissions - only admin and operator can delete
        if self.current_user_role not in ['admin', 'operator']:
            self.msg_box.showerror("Access Denied", "Only administrators and operators can delete master data entries.")
            return
        
        listbox = self.master_entries[category]['listbox']
        selection = listbox.curselection()
        if not selection:
            self.msg_box.showerror("Error", f"No item selected to delete from {category}.")
            return

        value_to_delete = listbox.get(selection[0])

        referenced_in_transactions = False
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                # Map master data tables to transaction columns for checking references
                column_map = {
                    "companies": "company", "trucks": "truck_plate", "products": "product",
                    "drivers": "driver", "origins": "origin", "destinations": "destination",
                    "designations": "designation", "senders": "sender"
                }
                if category in column_map:
                    col_name = column_map[category]
                    cursor.execute(f"SELECT COUNT(*) FROM transactions WHERE {col_name} = ?", (value_to_delete,))
                    if cursor.fetchone()[0] > 0:
                        referenced_in_transactions = True

                if referenced_in_transactions:
                    self.msg_box.showerror("Deletion Error", f"Cannot delete '{value_to_delete}' from {category} because it is referenced by existing transactions. Please update or delete the related transactions first.")
                    return

                confirmed = self.msg_box.askyesno("Confirm Delete", f"Are you sure you want to delete '{value_to_delete}' from {category}? This cannot be undone.")
                if confirmed:
                    try:
                        with sqlite3.connect(DB_FILE) as conn:
                            conn.execute(f"DELETE FROM {category} WHERE name = ?", (value_to_delete,))
                            conn.commit()
                        self.msg_box.showinfo("Success", f"Deleted '{value_to_delete}' from {category}.")
                    except sqlite3.Error as e:
                        self.msg_box.showerror("Database Error", f"Failed to delete from {category}: {e}")
                    except Exception as e:
                        self.msg_box.showerror("Error", f"An unexpected error occurred while deleting from {category}: {e}")
                    finally:
                        self.load_master_data_lists()

        except Exception as e:
            self.msg_box.showerror("Error", f"An unexpected error occurred: {e}")

# Part 4: Activation, Button States, and Auto-Connections

    def _initialize_activation_state_internal(self):
        """
        Initializes the application's activation state based on the config file.
        This sets up the trial period or checks for permanent activation.
        """
        today = datetime.now().date()
        
        expiry_str = "N/A" 

        if self.config.get("is_activated", False):
            self.is_app_activated = True
            expiry_str = self.config.get("expiry_date", "N/A")
            self.activation_status_var.set(f"Application Activated. Expires: {expiry_str}")
            return

        initial_trial_start_date_str = self.config.get("initial_trial_start_date")
        expiry_date_str = self.config.get("expiry_date")

        if not initial_trial_start_date_str:
            self.config["initial_trial_start_date"] = today.strftime("%Y-%m-%d")
            self.config["expiry_date"] = (today + timedelta(days=EXPIRY_DAYS_INITIAL_TRIAL)).strftime("%Y-%m-%d")
            self.config["is_activated"] = False
            self.config["has_saved_initial_config"] = False
            expiry_date_str = self.config["expiry_date"]

        try:
            expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
            if today <= expiry_date:
                self.is_app_activated = True
                self.activation_status_var.set(f"Trial Active. Expires: {expiry_date_str}")
            else:
                self.is_app_activated = False
                self.activation_status_var.set(f"Trial Expired on {expiry_date_str}. Please activate.")
        except (ValueError, TypeError):
            self.is_app_activated = False
            self.activation_status_var.set("Activation Status Corrupted. Please activate.")

    def _attempt_activation(self):
        """
        Handles the activation process based on a hardcoded code.
        """
        entered_code = self.activation_code_entry_var.get()
        if entered_code == ACTIVATION_CODE:
            self.is_app_activated = True
            new_expiry_date = (datetime.now().date() + timedelta(days=EXPIRY_DAYS_ON_ACTIVATION)).strftime("%Y-%m-%d")
            self.config["expiry_date"] = new_expiry_date
            self.config["is_activated"] = True
            self.save_config()
            self.msg_box.showinfo("Activation", "Application activated successfully! Save functionality unlocked.")
            self._update_action_button_states()
        else:
            self.msg_box.showerror("Activation Failed", "Invalid activation code. Please try again.")
            self.activation_code_entry_var.set("")
        self.activation_code_entry_var.set("")
    
    def _update_action_button_states(self):
        """
        Manages the state (enabled/disabled) of various buttons and UI elements
        based on application state, such as login status and app activation.
        """
        if hasattr(self, 'save_button'):
            is_activated = self.is_app_activated
            
            # Check if mandatory fields are filled
            company_filled = self.entry_vars.get('company', tk.StringVar()).get().strip() != ""
            truck_plate_filled = self.entry_vars.get('truck_plate', tk.StringVar()).get().strip() != ""
            
            can_save = is_activated and company_filled and truck_plate_filled

            self.save_button.config(state=tk.NORMAL if can_save else tk.DISABLED)
            
            if not is_activated:
                if hasattr(self, '_emulator_start_send_button'):
                    self._emulator_start_send_button.config(state=tk.DISABLED)
                    self._emulator_stop_send_button.config(state=tk.DISABLED)
                self.weight_source_status.set("EXPIRED - Activate App")
                self.weight_source_status_label.config(foreground="red")

        if hasattr(self, '_emulator_start_send_button'):
            if self._emulator_serial_port and self._emulator_serial_port.is_open:
                self._emulator_start_send_button.config(state=tk.NORMAL if not self._emulator_sending_data and self.is_app_activated else tk.DISABLED)
                self._emulator_stop_send_button.config(state=tk.NORMAL if self._emulator_sending_data else tk.DISABLED)
            else:
                self._emulator_start_send_button.config(state=tk.DISABLED)
                self._emulator_stop_send_button.config(state=tk.DISABLED)

    def get_serial_ports(self) -> list[str]:
        """
        Returns a list of available serial port names.
        """
        ports = [port.device for port in serial.tools.list_ports.comports()]
        return ports

    def _attempt_auto_connections(self):
        """
        Attempts to automatically connect to serial ports (main scale, emulator, big display)
        if configured to do so on startup.
        """
        if self.config.get("auto_connect", False) and self.config.get("port") and self.config.get("baud"):
            self.start_serial_reading(
                self.config["port"],
                int(self.config["baud"]),
                self.config.get("data_format_regex", r"(\d+\.\d+)"),
            )

        if self.config.get("emulator_auto_connect", False) and \
           self._emulator_port_entry_var.get() and self._emulator_baud_entry_var.get():
            self._emulator_connect_serial(
                self._emulator_port_entry_var.get(),
                int(self.config.get("emulator_baud"))
            )

        if self.big_display_auto_connect_var.get() and \
           self.big_display_port_var.get() and self.big_display_baud_var.get():
            self._connect_big_display()

        # Auto-connect camera if enabled
        if self.camera_auto_connect_var.get() and self.camera_manager:
            selected_camera = self.camera_device_var.get()
            if selected_camera:
                try:
                    # Extract camera index from selection (e.g., "Camera 0" -> 0)
                    camera_index = int(selected_camera.split()[-1])
                    if self.camera_manager.connect_camera(camera_index):
                        self.camera_connected = True
                        self.camera_status_var.set("Camera Status: Connected")
                        if hasattr(self, 'camera_connect_button'):
                            self.camera_connect_button.config(text="Disconnect Camera")
                        
                        # Set mirror mode
                        self.camera_manager.set_mirror_mode(self.camera_mirror_mode_var.get())
                        
                        # Start capture
                        self.camera_manager.start_capture()
                        logging.info(f"Camera auto-connected: {selected_camera}")
                    else:
                        logging.warning(f"Failed to auto-connect camera: {selected_camera}")
                except (ValueError, IndexError) as e:
                    logging.error(f"Invalid camera selection for auto-connect: {selected_camera}")

# Part 5: GUI, Serial Communication, and Data Handling

    def build_gui(self):
        """
        Builds the main graphical user interface with all the tabs.
        This is called after a successful login.
        """
        self.main_notebook = ttk.Notebook(self.root)
        self.main_notebook.pack(expand=True, fill="both")
        
        entry_tab = ttk.Frame(self.main_notebook)
        pending_tab = ttk.Frame(self.main_notebook)
        transaction_tab = ttk.Frame(self.main_notebook)
        report_tab = ttk.Frame(self.main_notebook)
        settings_container_tab = ttk.Frame(self.main_notebook)

        self.main_notebook.add(entry_tab, text="Entry Form")
        self.main_notebook.add(pending_tab, text="Pending Records")
        self.main_notebook.add(transaction_tab, text="Completed Records")
        self.main_notebook.add(report_tab, text="Reports")
        self.main_notebook.add(settings_container_tab, text="Settings")

        self.build_entry_tab(entry_tab)
        self.build_pending_tab(pending_tab)
        self.build_completed_records_tab(transaction_tab)
        self.build_report_tab(report_tab)
        self.build_settings_container_tab(settings_container_tab)

        self.main_notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        
        # Apply role-based access control
        self._apply_role_based_access_control()

    def _update_price_settings_state(self):
        """Update price computation settings enabled state based on current user role."""
        if not hasattr(self, 'price_settings_frame'):
            return
            
        is_privileged = hasattr(self, 'current_user_role') and self.current_user_role in ['admin', 'operator']
        
        # Update all price computation settings widgets
        for widget in self.price_settings_frame.winfo_children():
            if isinstance(widget, ttk.Checkbutton) or isinstance(widget, ttk.Entry) or isinstance(widget, ttk.Button):
                widget.config(state="normal" if is_privileged else "disabled")

    def _apply_role_based_access_control(self):
        """
        Applies role-based access control to tabs based on the logged-in user's role.
        - 'user': Can access all tabs EXCEPT Settings
        - 'operator': Can access all tabs EXCEPT Settings sub-tabs (Comm Ports, Activation, Print & Ticket Settings)
        - 'admin': Can access ALL tabs and all functions
        """
        if not hasattr(self, 'main_notebook') or not hasattr(self, 'current_user_role'):
            return
        
        user_role = self.current_user_role
        
        # Update price computation settings state
        self._update_price_settings_state()
        
        # Get the Settings tab index
        settings_tab_index = None
        for i, tab_id in enumerate(self.main_notebook.tabs()):
            if self.main_notebook.tab(tab_id, "text") == "Settings":
                settings_tab_index = i
                break
        
        if user_role == "user":
            # Users cannot access the Settings tab
            if settings_tab_index is not None:
                self.main_notebook.tab(settings_tab_index, state="disabled")
        
        elif user_role == "operator":
            # Operators can access Settings tab but not Comm Ports, Activation, and Print & Ticket Settings sub-tabs
            if settings_tab_index is not None:
                self.main_notebook.tab(settings_tab_index, state="normal")
            
            # CRITICAL: Hide restricted tabs within Settings
            # Wait a moment to ensure settings_notebook exists
            if hasattr(self, 'settings_notebook') and self.settings_notebook is not None:
                tabs_to_hide = ["Comm Ports", "Activation", "Ticket Format", "Print & Ticket Settings", "Admin Settings"]
                for tab_id in self.settings_notebook.tabs():
                    tab_text = self.settings_notebook.tab(tab_id, "text")
                    if tab_text in tabs_to_hide:
                        self.settings_notebook.tab(tab_id, state="disabled")
                    else:
                        self.settings_notebook.tab(tab_id, state="normal")
            else:
                # If settings_notebook not ready, schedule this to run again
                self.root.after(100, self._apply_role_based_access_control)
        
        elif user_role == "admin":
            # Admins have full access to all tabs
            if settings_tab_index is not None:
                self.main_notebook.tab(settings_tab_index, state="normal")
            
            if hasattr(self, 'settings_notebook') and self.settings_notebook is not None:
                for tab_id in self.settings_notebook.tabs():
                    self.settings_notebook.tab(tab_id, state="normal")

    def _reset_recalled_details(self):
        """
        Clears the displayed details of a recalled transaction.
        """
        self.recalled_record_no_var.set("")
        self.recalled_weight_var.set("")
        self.recalled_weight_date_var.set("")
        self.recalled_weight_time_var.set("")


    def start_serial_reading(self, port: str, baud: int, data_format_regex: str):
        """
        Starts the serial connection and a background thread to read from the scale.
        """
        if self._emulator_sending_data:
            self.stop_serial_connection()
            self.msg_box.showinfo("Info", "Emulator sending stopped to allow Main Scale connection.")

        if self.serial_running:
            self.stop_serial_connection()

        try:
            self.serial_port = serial.Serial(port, baudrate=baud, timeout=0.01)
            self.serial_running = True
            self.serial_thread = threading.Thread(
                target=self.read_serial,
                daemon=True
            )
            self.serial_thread.start()
            self.connect_button.config(state=tk.DISABLED)
            self.stop_button.config(state=tk.NORMAL)
            if not "EXPIRED" in self.weight_source_status.get():
                # On successful connection, set the display to 0.00 and status to Live Scale
                self.root.after(0, lambda: self._update_weight_display_and_status("0.00", "Live Scale", "green"))
        except serial.SerialException as e:
            self.msg_box.showerror("Serial Error", f"Could not open main serial port {port}: {e}")
            self.connect_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.root.after(0, lambda: self._update_weight_display_and_status("ERROR", "Error", "red"))
        except Exception as e:
            self.msg_box.showerror("Serial Error", f"An unexpected error occurred: {e}")
            self.connect_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.root.after(0, lambda: self._update_weight_display_and_status("ERROR", "Error", "red"))

    def read_serial(self):
        """
        The main loop for reading data from the serial port.
        It handles data parsing, stability checking, and UI updates.
        """
        current_stable_weight = None
        stability_count = 0

        while self.serial_running:
            data_format_regex = self.data_format_regex_var.get()
            current_time_str = datetime.now().strftime("%I:%M:%S %p")
            self.root.after(0, lambda: self.auto_time_var.set(current_time_str))

            try:
                if not self.serial_port or not self.serial_port.is_open:
                    self.root.after(0, lambda: self._update_weight_display_and_status("DISCONNECTED", "Disconnected", "gray"))
                    time.sleep(2)
                    continue

                line = self.serial_port.readline().decode(errors='ignore').strip()
                self.last_serial_line = line

                if line:
                    match = re.search(data_format_regex, line)
                    if match:
                        self.regex_fail_count = 0
                        try:
                            extracted_weight = float(match.group(1))

                            format_string = f"%.{self.decimal_places}f"
                            # Update the main weight display (weight_value) with the real-time received weight
                            self.root.after(0, lambda w=extracted_weight: self.weight_value.set(format_string % w))
                            # Also update the auto_weight_display_var for consistency
                            self.root.after(0, lambda w=extracted_weight: self.auto_weight_display_var.set(format_string % w))

                            if "Loaded Gross" not in self.weight_source_status.get() and \
                               "Editing ID" not in self.weight_source_status.get() and \
                               "EXPIRED" not in self.weight_source_status.get():
                                self.root.after(0, lambda: self.weight_source_status.set("Live Scale"))
                                self.root.after(0, lambda: self.weight_source_status_label.config(foreground="green"))

                            # Check for weight stability
                            if current_stable_weight is None or abs(current_stable_weight - extracted_weight) >= 0.01:
                                current_stable_weight = extracted_weight
                                stability_count = 1
                            else:
                                stability_count += 1

                            # Send stable weight to big display
                            if stability_count >= REQUIRED_WEIGHT_STABILITY:
                                self._send_weight_to_big_display(extracted_weight, "N/A", "N/A")

                        except (ValueError, IndexError):
                            self.regex_fail_count += 1
                    else:
                        self.regex_fail_count += 1
                        
                        if self.regex_fail_count > 3:
                            self.root.after(0, lambda: self._update_weight_display_and_status("NO MATCH", "No Match", "orange"))

                        if self.regex_fail_count > 10:
                            # Auto-detect if regex is enabled and failing
                            if self.auto_detect_regex_enabled_var.get():
                                new_regex = self._auto_detect_regex(line)
                                if new_regex and new_regex != data_format_regex:
                                    self.root.after(0, self._update_regex_from_detection, new_regex)
                            self.regex_fail_count = 0

                else:
                    if current_stable_weight is not None:
                        stability_count = 0
                        current_stable_weight = None

            except serial.SerialException as e:
                self.root.after(0, lambda: self._update_weight_display_and_status("PORT ERR", "Port Error", "red"))
                self.serial_running = False
                self.serial_port = None
                self.root.after(0, lambda: self.connect_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.stop_button.config(state=tk.DISABLED))
                break
            except Exception as e:
                self.root.after(0, lambda: self.big_display_status_var.set("Big Display Status: Error"))

            # Modified: Use the dynamic read_loop_interval_ms
            time.sleep(self.read_loop_interval_ms.get() / 1000.0)

    def stop_serial_connection(self):
        """
        Stops the main serial connection and cleans up the thread and port.
        """
        self.serial_running = False
        if self.serial_thread and self.serial_thread.is_alive():
            self.serial_thread.join(timeout=1)
        if self.serial_port and self.serial_port.is_open:
            self.serial_port.close()

        format_string = f"%.{self.decimal_places}f"
        if not "EXPIRED" in self.weight_source_status.get():
            self.weight_value.set(format_string % 0.00)
            self.weight_source_status.set("No Source")
            self.weight_source_status_label.config(foreground="gray")
        self._reset_recalled_details()
        self.connect_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)

    def _start_main_serial_connection_from_settings(self):
        """
        Initiates the main scale connection from the settings tab.
        """
        port = self.port_entry_var.get()
        baud_rate = self.baud_entry_var.get()
        data_regex = self.data_format_regex_var.get()

        if not port or not baud_rate:
            self.msg_box.showerror("Missing Info", "Please select a COM Port and Baud Rate for the Main Scale.")
            return

        try:
            int(baud_rate)
            re.compile(data_regex)
        except ValueError:
            self.msg_box.showerror("Invalid Input", "Baud Rate must be a number.")
            return
        except re.error as e:
            self.msg_box.showerror("Invalid Regex", f"The provided regex is invalid: {e}")
            return

        self.start_serial_reading(port, int(baud_rate), data_regex)
        
    def _auto_detect_regex(self, line: str) -> str | None:
        """
        Attempts to automatically detect a matching regex from the predefined list
        for a given line of serial data.
        """
        for regex_pattern in self.predefined_regexes:
            if regex_pattern == "Custom":
                continue
            try:
                match = re.search(regex_pattern, line)
                if match and match.group(1):
                    float(match.group(1))
                    return regex_pattern
            except (ValueError, IndexError):
                continue
        return None

    def _update_regex_from_detection(self, new_regex: str):
        """
        Updates the regex variable and notifies the user.
        """
        self.data_format_regex_var.set(new_regex)
        self.msg_box.showinfo(
            "Regex Auto-Detected",
            f"A new data format has been detected and applied:\n\n{new_regex}\n\n"
            "This has been updated in the Settings tab."
        )

    def _manual_detect_regex_action(self):
        """
        A user-triggered action to manually detect the regex from the last received line.
        """
        if not self.serial_running:
            self.msg_box.showerror("Error", "Serial connection is not active. Cannot detect regex.")
            return
        if not self.last_serial_line:
            self.msg_box.showinfo("Info", "No data has been received from the scale yet. Please wait.")
            return

        new_regex = self._auto_detect_regex(self.last_serial_line)
        if new_regex:
            self._update_regex_from_detection(new_regex)
        else:
            self.msg_box.showwarning(
                "Detection Failed",
                f"Could not automatically determine a data format for the last received line:\n\n'{self.last_serial_line}'"
            )
    
    def build_big_display_settings_sub_tab(self, tab: ttk.Frame):
        """Builds the Big Display settings sub-tab."""
        big_display_frame = ttk.LabelFrame(tab, text="Big Display Settings", padding="10")
        big_display_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5, ipady=5)
        
        # Big Display Port Selection
        ttk.Label(big_display_frame, text="Big Display COM Port:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.big_display_port_entry = ttk.Combobox(big_display_frame, values=self.get_serial_ports(), textvariable=self.big_display_port_var, style="TCombobox")
        self.big_display_port_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        
        # Big Display Baud Rate
        ttk.Label(big_display_frame, text="Baud Rate:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.big_display_baud_entry = ttk.Combobox(big_display_frame, values=["9600", "19200", "38400", "57600", "115200"], textvariable=self.big_display_baud_var, style="TCombobox")
        self.big_display_baud_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
        
        # Buttons
        button_frame = ttk.Frame(big_display_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        self.big_display_connect_button = ttk.Button(button_frame, text="Connect", command=self._connect_big_display)
        self.big_display_connect_button.pack(side=tk.LEFT, padx=5)
        
        self.big_display_disconnect_button = ttk.Button(button_frame, text="Disconnect", command=self._disconnect_big_display)
        self.big_display_disconnect_button.pack(side=tk.LEFT, padx=5)
        
        # Status Label
        self.big_display_status_label = ttk.Label(big_display_frame, textvariable=self.big_display_status_var, style="TLabel")
        self.big_display_status_label.grid(row=3, column=0, columnspan=2, pady=5)
        
    def _on_tab_changed(self, event: tk.Event):
        """
        Handles actions to perform when a user switches tabs in the main notebook.
        """
        selected_tab_widget = event.widget
        selected_tab_text = selected_tab_widget.tab(selected_tab_widget.select(), "text")

        if selected_tab_text == "Pending Records":
            # Load all pending transactions without date filters
            self.load_pending(skip_date_filter=True)
        elif selected_tab_text == "Completed Records":
            # Load all completed transactions without date filters
            self.load_transactions(skip_date_filter=True)
        elif selected_tab_text == "Entry Form":
            self.refresh_master_data()
            self.rebuild_entry_form_tab()
            self._update_action_button_states()
        elif selected_tab_text == "Reports":
            # Load all transactions without date filters
            self._search_report_transactions(skip_date_filter=True)
        elif selected_tab_text == "Settings":
            if hasattr(self, 'settings_notebook'):
                current_internal_tab_widget = self.settings_notebook.tab(self.settings_notebook.select(), "text")
                self._on_internal_settings_tab_changed(event=event)
# Part 6: Entry Form Methods

    def rebuild_entry_form_tab(self):
        """
        Rebuilds the Entry Form tab, useful for reflecting changes in the
        customizable entry fields.
        """
        try:
            notebook = self.main_notebook
            if not notebook:
                return
            
            # Try to find and index the Entry Form tab
            entry_tab_index = notebook.index("Entry Form")
            entry_tab_frame = notebook.winfo_children()[entry_tab_index]

            for widget in entry_tab_frame.winfo_children():
                widget.destroy()

            self.build_entry_tab(entry_tab_frame)
        except tk.TclError:
            # Tab not yet created, skip rebuild
            logging.debug("Entry Form tab not yet created, skipping rebuild")
        except Exception as e:
            logging.error(f"Error rebuilding entry form tab: {e}")

    def build_entry_tab(self, tab: ttk.Frame):
        """
        Builds the main entry form for capturing weights and transaction data.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        main_entry_frame = ttk.Frame(tab)
        main_entry_frame.pack(padx=10, pady=10, fill="both", expand=True)

        main_entry_frame.grid_columnconfigure(0, weight=1, uniform="entry_cols")
        main_entry_frame.grid_columnconfigure(1, weight=1, uniform="entry_cols")
        main_entry_frame.grid_rowconfigure(0, weight=1)  # Top row with transaction details and camera
        main_entry_frame.grid_rowconfigure(1, weight=0)  # Bottom buttons - minimal space

        col_left_container = ttk.Frame(main_entry_frame)
        col_left_container.grid(row=0, column=0, sticky="nsew", padx=(5, 2), pady=5)
        col_left_container.grid_rowconfigure(0, weight=0)  # Weight option
        col_left_container.grid_rowconfigure(1, weight=1)  # Transaction details
        col_left_container.grid_rowconfigure(2, weight=0)  # Live scale weight title
        col_left_container.grid_rowconfigure(3, weight=0)  # Live scale weight display
        col_left_container.grid_rowconfigure(4, weight=0)  # Live scale weight status - minimal
        col_left_container.grid_columnconfigure(0, weight=1)

        weight_option_frame = ttk.LabelFrame(col_left_container, text="WEIGHT OPTION", padding="10")
        weight_option_frame.grid(row=0, column=0, sticky="ew", padx=(5, 2), pady=5)
        
        self.one_way_radio = ttk.Radiobutton(weight_option_frame, text="ONE WAY", variable=self.weight_option_var, value="ONE_WAY", 
                                      command=lambda: self.save_config(), style="TRadiobutton")
        self.one_way_radio.pack(side=tk.LEFT, padx=10, pady=5)
        
        self.two_way_radio = ttk.Radiobutton(weight_option_frame, text="TWO WAY", variable=self.weight_option_var, value="TWO_WAY",
                                              command=lambda: self.save_config(), style="TRadiobutton")
        self.two_way_radio.pack(side=tk.LEFT, padx=10, pady=5)

        transaction_details_frame = ttk.LabelFrame(col_left_container, text="TRANSACTION DETAILS", padding="10")
        transaction_details_frame.grid(row=1, column=0, sticky="nsew", padx=(5, 2), pady=5)
        transaction_details_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(transaction_details_frame, text="TICKET #:", style="TLabel").grid(row=0, column=0, sticky="e", padx=5, pady=2)
        self.ticket_no_label = tk.Label(transaction_details_frame, text="0", bg="#FFFFE0", relief="sunken", bd=1, font=("Helvetica", 10), anchor='w')
        self.ticket_no_label.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        self.ticket_no_label.config(text=str(self.get_next_ticket_no()))


        self.entry_vars = {}
        self.entry_combos = {}
        
        last_field_row = 1
        for i, (display_name, internal_key) in enumerate(self.selected_entry_fields):
            ttk.Label(transaction_details_frame, text=display_name.upper() + ":", style="TLabel").grid(row=i+last_field_row, column=0, sticky="e", padx=5, pady=2)
            
            # Special handling for Total Price field
            if internal_key == "total_price":
                # Total Price is read-only and displays computed value
                var = tk.StringVar(value="0")
                
                # Create a frame to hold the total price label and toggle button
                total_price_frame = ttk.Frame(transaction_details_frame)
                total_price_frame.grid(row=i+last_field_row, column=1, sticky="ew", padx=5, pady=2)
                total_price_frame.grid_columnconfigure(0, weight=1)
                
                label = ttk.Label(total_price_frame, textvariable=var, font=("Courier", 10, "bold"), 
                                foreground="green", relief="sunken", anchor="w")
                label.grid(row=0, column=0, sticky="ew", padx=(0, 5))
                
                # Add toggle button for computation enable/disable
                computation_toggle_btn = ttk.Checkbutton(total_price_frame, text="Enable", 
                                                         variable=self.price_computation_enabled_var,
                                                         command=self._on_price_computation_toggled)
                computation_toggle_btn.grid(row=0, column=1, sticky="e")
                
                self.entry_vars[internal_key] = var
                # Store reference to label for updates
                self.total_price_entry_label = label
            else:
                var = tk.StringVar()
                # New: Auto-capitalize all entry form inputs for consistency
                var.trace_add("write", lambda name, index, mode, var=var: var.set(var.get().upper()))
                
                if internal_key in ["company", "truck_plate"]:
                    var.trace_add("write", lambda name, index, mode, sv=self: sv._update_action_button_states())

                combo = ttk.Combobox(transaction_details_frame, textvariable=var, state="normal", style="TCombobox")
                combo.grid(row=i+last_field_row, column=1, sticky="ew", padx=5, pady=2)
                self.entry_vars[internal_key] = var
                self.entry_combos[internal_key] = combo
            
        for i in range(last_field_row + len(self.selected_entry_fields)):
            transaction_details_frame.grid_rowconfigure(i, weight=1)


        col_right_container = ttk.Frame(main_entry_frame)
        col_right_container.grid(row=0, column=1, sticky="nsew", padx=(2, 5), pady=5)
        col_right_container.grid_rowconfigure(0, weight=0)  # Recalled details - no expansion
        col_right_container.grid_rowconfigure(1, weight=1)  # Camera feed - gets most space
        col_right_container.grid_columnconfigure(0, weight=1)

        recalled_info_frame = ttk.LabelFrame(col_right_container, text="📋 RECALLED TRANSACTION DETAILS", padding="8")
        recalled_info_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        recalled_info_frame.grid_columnconfigure(1, weight=1)

        # Compact styling with smaller fonts and reduced spacing
        recalled_label_font = ("Helvetica", 10, "bold")
        recalled_value_font = ("Courier New", 10, "bold")
        
        # Define custom colors for better visual appeal
        header_bg = "#E8F4F8"  # Light blue
        value_bg = "#F0F8FF"   # Alice blue
        border_color = "#4A90E2"  # Professional blue

        row_offset = 0
        
        # Record Number
        ttk.Label(recalled_info_frame, text="RECORD NO:", font=recalled_label_font, style="TLabel", foreground="#1F4E78").grid(row=row_offset, column=0, sticky="e", padx=8, pady=4)
        record_label = tk.Label(recalled_info_frame, textvariable=self.recalled_record_no_var, bg=value_bg, relief="solid", bd=2, font=recalled_value_font, anchor='w', fg="#2E5090")
        record_label.grid(row=row_offset, column=1, sticky="ew", padx=8, pady=4)
        row_offset += 1

        # Separator line
        ttk.Separator(recalled_info_frame, orient=tk.HORIZONTAL).grid(row=row_offset, column=0, columnspan=2, sticky="ew", padx=0, pady=2)
        row_offset += 1

        # Recalled Weight
        ttk.Label(recalled_info_frame, text="RECALLED WEIGHT:", font=recalled_label_font, style="TLabel", foreground="#1F4E78").grid(row=row_offset, column=0, sticky="e", padx=8, pady=4)
        weight_label = tk.Label(recalled_info_frame, textvariable=self.recalled_weight_var, bg=value_bg, relief="solid", bd=2, font=recalled_value_font, anchor='w', fg="#2E5090")
        weight_label.grid(row=row_offset, column=1, sticky="ew", padx=8, pady=4)
        row_offset += 1

        # Date and Time on same row with better layout
        ttk.Label(recalled_info_frame, text="RECALLED DATE:", font=recalled_label_font, style="TLabel", foreground="#1F4E78").grid(row=row_offset, column=0, sticky="e", padx=8, pady=4)
        date_label = tk.Label(recalled_info_frame, textvariable=self.recalled_weight_date_var, bg=value_bg, relief="solid", bd=2, font=recalled_value_font, anchor='w', fg="#2E5090")
        date_label.grid(row=row_offset, column=1, sticky="ew", padx=8, pady=4)
        row_offset += 1

        ttk.Label(recalled_info_frame, text="RECALLED TIME:", font=recalled_label_font, style="TLabel", foreground="#1F4E78").grid(row=row_offset, column=0, sticky="e", padx=8, pady=4)
        time_label = tk.Label(recalled_info_frame, textvariable=self.recalled_weight_time_var, bg=value_bg, relief="solid", bd=2, font=recalled_value_font, anchor='w', fg="#2E5090")
        time_label.grid(row=row_offset, column=1, sticky="ew", padx=8, pady=4)
        row_offset += 1

        # Remove expandable frame - let the frame size naturally fit its content

        # Camera Feed Frame - positioned under recalled transaction details, aligned with Save button column
        self.camera_feed_frame = ttk.LabelFrame(col_right_container, text="📷 CAMERA FEED")
        self.camera_feed_frame.grid(row=1, column=0, sticky="nsew")
        self.camera_feed_frame.grid_columnconfigure(0, weight=1)
        self.camera_feed_frame.grid_rowconfigure(0, weight=1)
        
        # Camera display label - fully covers the area with no padding, system theme background
        self.camera_label = tk.Label(self.camera_feed_frame, bg="white", relief="sunken", bd=2)
        self.camera_label.grid(row=0, column=0, sticky="nsew")
        
        # Set camera manager callback
        if self.camera_manager:
            self.camera_manager.set_frame_callback(self._update_camera_frame)
        
        # Initialize camera list
        self._refresh_camera_list()


        # Live Scale Weight - Status moved outside, enlarged for 6 digits
        # Add title label above the weight display
        ttk.Label(col_left_container, text="LIVE SCALE WEIGHT", font=("Helvetica", 10, "bold"), style="TLabel").grid(row=2, column=0, sticky="ew", padx=5, pady=(2, 0))
        
        captured_weight_display_frame = ttk.LabelFrame(col_left_container, text="", padding="2")
        captured_weight_display_frame.grid(row=3, column=0, sticky="ew", padx=5, pady=2)
        captured_weight_display_frame.grid_columnconfigure(0, weight=1)

        # The main weight display - enlarged font for up to 6 digits
        self.captured_weight_label = tk.Label(captured_weight_display_frame, textvariable=self.weight_value,
                                                font=("Helvetica", 36, "bold"), bg="#ADD8E6", fg="navy",
                                                relief="sunken", bd=2, anchor="center")
        self.captured_weight_label.pack(expand=False, fill="both", padx=2, pady=2)

        # Status label moved outside the box - positioned below
        self.weight_source_status_label = ttk.Label(col_left_container, textvariable=self.weight_source_status,
                                                    font=("Helvetica", 8, "italic"), foreground="gray")
        self.weight_source_status_label.grid(row=4, column=0, sticky="e", padx=5, pady=(0, 2))

        # Update the button states based on current connection status
        self._update_big_display_button_states()
        
        bottom_button_frame = ttk.Frame(main_entry_frame, padding="5")
        bottom_button_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
        
        action_buttons_inner_frame = ttk.Frame(bottom_button_frame)
        action_buttons_inner_frame.pack(expand=True, fill="x")

        action_buttons_inner_frame.grid_columnconfigure(0, weight=1)
        action_buttons_inner_frame.grid_columnconfigure(1, weight=1)

        self.new_button = ttk.Button(action_buttons_inner_frame, text="New", command=self._reset_entry_form_for_new_entry, style="TButton")
        self.new_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        self.save_button = ttk.Button(action_buttons_inner_frame, text="Save", command=self._handle_save_button_click, style="TButton")
        self.save_button.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self._sync_camera_feed_width()

        self.refresh_master_data()
        self._update_action_button_states()

    def _update_transaction_price_display(self, *args):
        """
        Update transaction price display when weight value changes.
        """
        try:
            weight_str = self.weight_value.get()
            
            if weight_str and self.price_computation_enabled_var.get():
                weight_kg = float(weight_str)
                unit_price, total_price = self.calculate_price(weight_kg)
                total_price_str = f"{total_price:.0f}"
                
                # Update customize entry form display
                if hasattr(self, 'transaction_total_price_var'):
                    self.transaction_total_price_var.set(total_price_str)
                
                # Update entry form total price field if it exists
                if hasattr(self, 'total_price_entry_label') and hasattr(self, 'entry_vars') and 'total_price' in self.entry_vars:
                    self.entry_vars['total_price'].set(total_price_str)
            else:
                total_price_str = "0"
                
                # Update customize entry form display
                if hasattr(self, 'transaction_total_price_var'):
                    self.transaction_total_price_var.set(total_price_str)
                
                # Update entry form total price field if it exists
                if hasattr(self, 'total_price_entry_label') and hasattr(self, 'entry_vars') and 'total_price' in self.entry_vars:
                    self.entry_vars['total_price'].set(total_price_str)
                    
        except (ValueError, AttributeError):
            total_price_str = "0"
            
            # Update customize entry form display
            if hasattr(self, 'transaction_total_price_var'):
                self.transaction_total_price_var.set(total_price_str)
            
            # Update entry form total price field if it exists
            if hasattr(self, 'total_price_entry_label') and hasattr(self, 'entry_vars') and 'total_price' in self.entry_vars:
                self.entry_vars['total_price'].set(total_price_str)

    def _update_price_display(self, *args):
        """
        Update price display when weight value changes.
        """
        try:
            weight_str = self.weight_value.get()
            if weight_str:
                weight_kg = float(weight_str)
                unit_price, total_price = self.calculate_price(weight_kg)
                self.unit_price_var.set(f"{unit_price:.2f}")
                self.total_price_var.set(f"{total_price:.0f}")
            else:
                self.unit_price_var.set("0.00")
                self.total_price_var.set("0")
        except ValueError:
            self.unit_price_var.set("0.00")
            self.total_price_var.set("0")

    def _flash_weight_display(self, color: str = "green"):
        """
        Flashes the weight display background to indicate a success or error state.
        """
        self.captured_weight_label.config(bg=color)
        self.root.after(200, lambda: self.captured_weight_label.config(bg="#ADD8E6"))

    def _reset_entry_form_for_new_entry(self):
        """
        Resets all entry form fields for a new transaction.
        """
        self.current_edit_transaction_id = None
        self.clear_entry_form_fields_only()
        self.ticket_no_label.config(text=str(self.get_next_ticket_no()))

        if not "EXPIRED" in self.weight_source_status.get():
            self._update_weight_display_and_status(f"%.{self.decimal_places}f" % 0.00, "No Source", "gray")
        else:
            self.weight_value.set(f"%.{self.decimal_places}f" % 0.00)
        
        self.weight_option_var.set("TWO_WAY")

        self.refresh_master_data()
        self._update_action_button_states()

    def clear_entry_form_fields_only(self):
        """
        Clears the content of all entry form fields.
        """
        for var in self.entry_vars.values():
            var.set("")
        self._reset_recalled_details()

    def clear_entry_form_fields_except_truck_plate(self):
        """
        Clears all entry form fields except for the truck plate.
        """
        for _, internal_key in self.selected_entry_fields:
            if internal_key != 'truck_plate':
                if internal_key in self.entry_vars:
                    self.entry_vars[internal_key].set("")
        self._reset_recalled_details()

    def _load_transaction_for_editing(self, trans_id: int):
        """Loads an existing transaction into the entry form for editing."""
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("""SELECT company, truck_plate, product, designation, sender, origin, destination, driver,
                                gross_weight, tare_weight, net_weight, gross_date, gross_time, tare_date, tare_time,
                                weight_type, ticket_no, status
                                FROM transactions WHERE id = ?""", (trans_id,))
                transaction_details = cursor.fetchone()
                if transaction_details:
                    self.clear_entry_form_fields_only()
                    self.current_edit_transaction_id = trans_id
                    self._update_action_button_states()
                    col_names = ['company', 'truck_plate', 'product', 'designation', 'sender', 'origin', 'destination', 'driver',
                                'gross_weight', 'tare_weight', 'net_weight', 'gross_date', 'gross_time', 'tare_date', 'tare_time',
                                'weight_type', 'ticket_no', 'status']
                    data_map = {col_names[i]: transaction_details[i] for i in range(len(col_names))}
                    self.ticket_no_label.config(text=str(data_map['ticket_no']) if data_map['ticket_no'] is not None else "N/A")
                
                    # Store original values for comparison
                    self.original_transaction_values = {}
                
                    for i, (display_name, internal_key) in enumerate(self.selected_entry_fields):
                        if internal_key in data_map and internal_key in self.entry_vars:
                            # Store the original value
                            self.original_transaction_values[internal_key] = data_map.get(internal_key, "")
                            # Set the StringVar
                            self.entry_vars[internal_key].set(data_map.get(internal_key, "") or "")

                            # Add binding to prevent change for company and truck_plate
                            if internal_key in ["company", "truck_plate"] and internal_key in self.entry_combos:
                                combo = self.entry_combos[internal_key]
                            
                                def prevent_change(event, key=internal_key):
                                    # Check if we're editing a transaction and the value was loaded from DB
                                    if self.current_edit_transaction_id and key in self.original_transaction_values:
                                        # Get the original value
                                        original_value = self.original_transaction_values[key]
                                        # Get the current value in the StringVar 
                                        current_value = self.entry_vars[key].get()
                                        # If the user tried to change it, revert it back
                                        if current_value != original_value:
                                            # Revert the StringVar to the original value
                                            self.entry_vars[key].set(original_value)
                            
                                # Bind the event handler
                                combo.bind('<<ComboboxSelected>>', prevent_change)

                    # Handle weight display and other logic...
                    gross_weight = transaction_details[8]
                    gross_date = transaction_details[11]
                    gross_time = transaction_details[12]
                    ticket_no = transaction_details[15]
                    self.ticket_no_label.config(text=str(ticket_no))
                    format_string = f"%.{self.decimal_places}f"
                    if gross_weight is not None:
                        self._update_weight_display_and_status(self.weight_value.get(), "Awaiting Weight Capture", "blue", clear_recalled=False)
                    self.recalled_record_no_var.set(ticket_no if ticket_no is not None else "N/A")

                else:
                    self.msg_box.showerror("Error", "Transaction not found.")
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to retrieve transaction details for editing: {e}")
        except Exception as e:
            self.msg_box.showerror("Error", "Failed to load transaction details for editing.")
        finally:
            pass

    def _edit_selected_transaction_from_entry_form(self):
        """
        Triggers the update process for a transaction that is currently loaded in the entry form.
        """
        try:
            if self.current_edit_transaction_id is None:
                self.msg_box.showerror("Error", "No transaction loaded for editing. Use 'New' to start a new entry or select a transaction from 'Completed Records' tab to edit.")
                return
            
            self.update_transaction_details()
        finally:
            self._update_action_button_states()


    def update_transaction_details(self):
        """
        Updates the details of a transaction in the database with the values from the entry form.
        """
        if not self.is_app_activated:
            self.msg_box.showerror("Activation Required", "Please activate the application to update transactions.")
            return
        if self.current_edit_transaction_id is None:
            self.msg_box.showerror("Error", "No transaction selected for update.")
            return

        update_data = {
            'company': None, 'truck_plate': None, 'product': None,
            'designation': None, 'sender': None, 'origin': None,
            'destination': None, 'driver': None
        }

        for _, internal_key in self.selected_entry_fields:
            if internal_key in self.entry_vars:
                update_data[internal_key] = self.entry_vars[internal_key].get().strip()
        
        original_values = {}
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT gross_weight, gross_date, gross_time, tare_weight, tare_date, tare_time, weight_type, status, net_weight, ticket_no
                    FROM transactions WHERE id = ?
                """, (self.current_edit_transaction_id,))
                orig_data = cursor.fetchone()
                if orig_data:
                    original_values = {
                        'gross_weight': orig_data[0], 'gross_date': orig_data[1], 'gross_time': orig_data[2],
                        'tare_weight': orig_data[3], 'tare_date': orig_data[4], 'tare_time': orig_data[5],
                        'weight_type': orig_data[6], 'status': orig_data[7], 'net_weight': orig_data[8],
                        'ticket_no': orig_data[9]
                    }
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", "Failed to retrieve original transaction data for update.")
            return

        gross_weight = original_values.get('gross_weight')
        gross_date = original_values.get('gross_date')
        gross_time = original_values.get('gross_time')
        tare_weight = original_values.get('tare_weight')
        tare_date = original_values.get('tare_date')
        tare_time = original_values.get('tare_time')
        net_weight = original_values.get('net_weight')
        weight_type = original_values.get('weight_type')
        status = original_values.get('status')

        # Calculate unit_price and total_price if computation is enabled
        unit_price = 0.0
        total_price = 0.0
        if self.price_computation_enabled_var.get() and net_weight is not None:
            try:
                weight_kg = float(net_weight)
                unit_price, total_price = self.calculate_price(weight_kg)
            except (ValueError, AttributeError):
                unit_price = 0.0
                total_price = 0.0

        # Check if we're in the right tab context before adding to master data
        current_tab_index = self.main_notebook.index(self.main_notebook.select())
        allowed_tabs = [
            self.main_notebook.index("Pending Records"),
            self.main_notebook.index("Completed Records"), 
            self.main_notebook.index("Reports")
        ]
        
        if current_tab_index in allowed_tabs:
            master_data_fields = ["company", "truck_plate", "product", "designation", "sender", "origin", "destination", "driver"]
            for field in master_data_fields:
                if field in self.entry_vars and self.entry_vars[field].get():
                    self.auto_add_to_master(field, self.entry_vars[field].get())

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE transactions SET
                        company = ?, truck_plate = ?, product = ?, designation = ?, sender = ?,
                        origin = ?, destination = ?, driver = ?,
                        gross_weight = ?, tare_weight = ?, net_weight = ?,
                        gross_date = ?, gross_time = ?, tare_date = ?, tare_time = ?,
                        weight_type = ?, status = ?, unit_price = ?, total_price = ?,
                        timestamp = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (
                    update_data['company'], update_data['truck_plate'], update_data['product'], update_data['designation'], update_data['sender'],
                    update_data['origin'], update_data['destination'], update_data['driver'],
                    gross_weight, tare_weight, net_weight,
                    gross_date, gross_time, tare_date, tare_time,
                    weight_type, status, unit_price, total_price,
                    self.current_edit_transaction_id
                ))
                conn.commit()

        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to update transaction: {e}")
        except Exception as e:
            pass
        finally:
            self._reset_entry_form_for_new_entry()
            self.load_transactions()

    def _open_master_data_for_field(self, field_key: str):
        """
        Switches to the settings tab and focuses the specified master data entry field.
        """
        self.main_notebook.select(self.main_notebook.index("Settings"))
        master_data_tab_index = -1
        for i, tab_name in enumerate(self.settings_notebook.tabs()):
            if self.settings_notebook.tab(tab_name, "text") == "Master Data":
                master_data_tab_index = i
                break
        
        if master_data_tab_index != -1:
            self.settings_notebook.select(master_data_tab_index)
            self.root.after(100, lambda: self._focus_master_data_entry(field_key))
        else:
            self.msg_box.showwarning("Navigation Error", "Could not find Master Data tab.")

    def _focus_master_data_entry(self, field_key: str):
        """
        Sets the focus and populates the text for a specific master data entry field.
        """
        if field_key in self.master_entries:
            self.master_entries[field_key]['entry_widget'].focus_set()
            if field_key in self.master_entries:
                self.master_entries[field_key]['entry_var'].set(self.entry_vars[field_key].get())


    def _on_truck_plate_dropdown_selected(self, event=None):
        """
        Triggers an autofill attempt when the truck plate field is updated.
        """
        # This method is now empty to disable autofill
        return

    def _autofill_from_last_transaction(self, truck_plate: str):
        """
        Automatically fills other fields in the entry form based on the last transaction
        for a given truck plate.
        """
        # This method is now empty to disable autofill
        return


    def _load_pending_for_tare(self, truck_plate_or_id: int | str):
        """
        Loads a pending transaction into the entry form for the second weighing (tare).
        """
        trans_id = None
        if isinstance(truck_plate_or_id, int):
            trans_id = truck_plate_or_id
        elif isinstance(truck_plate_or_id, str):
            trans_id = self._get_pending_transaction_id(truck_plate_or_id)

        # Corrected: Changed '===' to '==' for Python syntax
        if trans_id == None:
            self.msg_box.showinfo("No Pending", f"No pending gross weighing found for '{truck_plate_or_id}'.")
            self._reset_entry_form_for_new_entry()
            return

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT company, truck_plate, product, designation, sender, origin, destination, driver,
                           gross_weight, tare_weight, net_weight, gross_date, gross_time, tare_date, tare_time, ticket_no
                    FROM transactions WHERE id = ? AND status = 'Pending'
                """, (trans_id,))
                transaction_details = cursor.fetchone()

                if transaction_details:
                    self.clear_entry_form_fields_only()

                    self.current_edit_transaction_id = trans_id
                    self._update_action_button_states()
                    
                    self.weight_option_var.set("TWO_WAY")

                    field_names = ['company', 'truck_plate', 'product', 'designation', 'sender', 'origin', 'destination', 'driver']
                    for i, field_name in enumerate(field_names):
                        if field_name in self.entry_vars:
                            self.entry_vars[field_name].set(transaction_details[i] if transaction_details[i] else "")

                    gross_weight = transaction_details[8]
                    gross_date = transaction_details[11]
                    gross_time = transaction_details[12]
                    ticket_no = transaction_details[15]

                    self.ticket_no_label.config(text=str(ticket_no))

                    format_string = f"%.{self.decimal_places}f"

                    if gross_weight is not None:
                        self._update_weight_display_and_status(
                            self.weight_value.get(),
                            "Awaiting Weight Capture",
                            "blue",
                            clear_recalled=False
                        )

                        self.recalled_record_no_var.set(ticket_no if ticket_no is not None else "N/A")
                        self.recalled_weight_var.set(format_string % gross_weight)
                        self.recalled_weight_date_var.set(gross_date if gross_date else "N/A")
                        self.recalled_weight_time_var.set(gross_time if gross_time else "N/A")

                    else:
                        self._update_weight_display_and_status(self.weight_value.get(), "No Gross Weight Found", "red", clear_recalled=True)
                        self._reset_recalled_details()
                else:
                    self.msg_box.showerror("Error", "Failed to load pending transaction details by ID. It might no longer be pending.")
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to retrieve transaction details for tare weighing: {e}")
        except Exception as e:
            pass
        finally:
            pass

    def refresh_master_data(self):
        """
        Refreshes the dropdown lists for the entry form fields with the latest data from the master tables.
        """
        table_mapping = {
            "company": "companies", "truck_plate": "trucks", "product": "products",
            "designation": "designations", "sender": "senders", "origin": "origins",
            "destination": "destinations", "driver": "drivers"
        }

        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            for _, internal_key in self.selected_entry_fields:
                table = table_mapping.get(internal_key)
                if table and internal_key in self.entry_combos:
                    cursor.execute(f"SELECT name FROM {table} ORDER BY name COLLATE NOCASE ASC")
                    results = [row[0] for row in cursor.fetchall()]
                    self.entry_combos[internal_key]['values'] = results
                else:
                    pass

    def auto_add_to_master(self, category_field_name: str, value: str):
        """
        Automatically adds a new value to the corresponding master data table if it doesn't already exist.
        This is for convenience during data entry.
        """
        table_mapping = {
            "company": "companies", "truck_plate": "trucks", "product": "products",
            "designation": "designations", "sender": "senders", "origin": "origins",
            "destination": "destinations", "driver": "drivers"
        }
        table_name = table_mapping.get(category_field_name)

        if not table_name or not value:
            return

        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute(f"SELECT name FROM {table_name} WHERE name = ? COLLATE NOCASE", (value,))
            if cursor.fetchone() is None:
                try:
                    conn.execute(f"INSERT INTO {table_name} (name) VALUES (?)", (value,))
                    conn.commit()
                except sqlite3.IntegrityError:
                    # Catch the case where a record was added by another thread
                    pass
                except Exception as e:
                    # Catch any other unexpected database errors
                    pass

    def get_next_ticket_no(self) -> int:
        """
        Retrieves the next available ticket number from the database.
        Uses the configurable starting ticket number if no records exist.
        """
        # Get the starting ticket number from configuration
        STARTING_OFFSET = self.starting_ticket_number_var.get()

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()

                # --- Check if the transactions table is empty ---
                cursor.execute("SELECT COUNT(*) FROM transactions")
                count_result = cursor.fetchone()
                table_count = count_result[0] if count_result else 0

                # --- Determine the next ticket number ---
                if table_count == 0:
                    # Table is empty, start from the configured offset
                    return STARTING_OFFSET
                else:
                    # Table has records, find the max ticket number and increment
                    cursor.execute("SELECT MAX(ticket_no) FROM transactions")
                    result = cursor.fetchone()
                    max_ticket = result[0] if result else 0
                    # Return the next number after the maximum
                    return (max_ticket if max_ticket is not None else 0) + 1

        except sqlite3.Error as e:
            # Handle potential database errors gracefully
            print(f"Database error in get_next_ticket_no: {e}")
            # Fallback: Return a safe default (e.g., STARTING_OFFSET)
            # You might prefer to raise an exception or handle it differently based on your needs
            return STARTING_OFFSET # This ensures a number is always returned


    def _get_pending_transaction_id(self, truck_plate: str) -> int | None:
        """
        Retrieves the ID of a pending transaction by truck plate.
        """
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM transactions WHERE truck_plate = ? AND status = 'Pending'", (truck_plate,))
            result = cursor.fetchone()
            return result[0] if result else None

    def _get_pending_transaction_id_and_details(self, truck_plate: str) -> tuple | None:
        """
        Retrieves the ID and details of a pending transaction by truck plate.
        """
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, gross_weight, gross_date, gross_time, tare_weight, net_weight, ticket_no FROM transactions WHERE truck_plate = ? AND status = 'Pending'", (truck_plate,))
            result = cursor.fetchone()
            return result if result else None
    
    def _get_pending_transaction_id_and_details_by_id(self, trans_id: int) -> tuple | None:
        """
        Retrieves the ID and details of a pending transaction by ID.
        """
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, gross_weight, gross_date, gross_time, tare_weight, net_weight, ticket_no FROM transactions WHERE id = ? AND status = 'Pending'", (trans_id,))
            result = cursor.fetchone()
            return result if result else None
            
    def _on_pending_column_resize(self, event):
        """Handle column resize for pending tree and save widths to config."""
        try:
            column_widths = {}
            for col in self.pending_tree["columns"]:
                width = self.pending_tree.column(col, "width")
                column_widths[col] = width
            
            self.config["pending_column_widths"] = column_widths
            self.save_config()
        except Exception as e:
            logging.error(f"Error saving pending column widths: {e}")

    def _on_transaction_column_resize(self, event):
        """Handle column resize for transaction tree and save widths to config."""
        try:
            column_widths = {}
            for col in self.transaction_tree["columns"]:
                width = self.transaction_tree.column(col, "width")
                column_widths[col] = width
            
            self.config["transaction_column_widths"] = column_widths
            self.save_config()
        except Exception as e:
            logging.error(f"Error saving transaction column widths: {e}")

    def _on_report_column_resize(self, event):
        """Handle column resize for report tree and save widths to config."""
        try:
            column_widths = {}
            for col in self.report_tree["columns"]:
                width = self.report_tree.column(col, "width")
                column_widths[col] = width
            
            self.config["report_column_widths"] = column_widths
            self.save_config()
        except Exception as e:
            logging.error(f"Error saving report column widths: {e}")

    def _prompt_print_ticket_after_save(self, trans_id: int):
        """
        Prompt user to show print preview for ticket after successful save operation.
        """
        try:
            # Ask user if they want to see the print preview for the ticket
            result = self.msg_box.askyesno("Print Preview", "Do you want to see the print preview for this transaction?")
            if result:
                # Get transaction data for print preview
                data = self._get_transaction_data_for_print(trans_id)
                if data:
                    # Show print preview using existing functionality
                    self._print_ticket_from_data(data)
                else:
                    self.msg_box.showerror("Print Preview Error", "Could not retrieve transaction data for preview.")
        except Exception as e:
            self.msg_box.showerror("Print Preview Error", f"An error occurred while preparing print preview: {e}")

    def _print_ticket_from_data(self, data: dict):
        """
        Show print preview for ticket directly from transaction data using existing print functionality.
        """
        try:
            # Use the existing print preview functionality instead of direct printing
            self._open_print_preview(data, "Entry Form")
        except Exception as e:
            self.msg_box.showerror("Print Preview Error", f"Failed to open print preview: {e}")

    def _handle_save_button_click(self):
        """
        Central function for handling the "Save" button logic,
        validating inputs and routing to the correct saving method.
        """
        self.save_button.config(state=tk.DISABLED)
        try:
            company_value = self.entry_vars.get('company', tk.StringVar()).get().strip()
            truck_plate_value = self.entry_vars.get('truck_plate', tk.StringVar()).get().strip()

            if not company_value:
                self.msg_box.showerror("Missing Information", "Company is a mandatory field.")
                self._flash_weight_display("red")
                return
            
            if not truck_plate_value:
                self.msg_box.showerror("Missing Information", "Truck Plate is a mandatory field.")
                self._flash_weight_display("red")
                return

            if self.current_edit_transaction_id is not None:
                self.update_transaction_details_from_live_or_dialog()
            else:
                selected_option = self.weight_option_var.get()

                if selected_option == "ONE_WAY":
                    self.record_one_way_weight()
                elif selected_option == "TWO_WAY":
                    self.record_two_way_weight()
                else:
                    self.msg_box.showerror("Error", "Please select a Weight Option (ONE WAY or TWO WAY).")
        finally:
            # Always re-enable the button after the operation is complete
            self._update_action_button_states()

    def update_transaction_details_from_live_or_dialog(self):
        """
        Determines whether an update is a tare weighing or a simple field edit
        for a transaction loaded from a different tab.
        """
        if self.current_edit_transaction_id is None:
            self.msg_box.showerror("Error", "No transaction loaded for update.")
            return

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT status FROM transactions WHERE id = ?", (self.current_edit_transaction_id,))
                status = cursor.fetchone()[0]
        except Exception as e:
            self.msg_box.showerror("Database Error", "Could not verify transaction status.")
            self._reset_entry_form_for_new_entry()
            return

        if status == 'Pending':
            self.record_two_way_weight()
        else:
            self.update_transaction_details()

    def record_one_way_weight(self):
        """
        Records a new one-way weighing transaction.
        """
        if not self.is_app_activated:
            self.msg_box.showerror("Activation Required", "Please activate the application to save transactions.")
            self._flash_weight_display("red")
            return

        try:
            current_weight = float(self.weight_value.get())
        except ValueError:
            self.msg_box.showerror("Invalid Weight", "Captured weight is not a valid number. Cannot record One Way weighing.")
            self._flash_weight_display("red")
            return

        if current_weight <= 0:
            self.msg_box.showerror("Error", "Weight must be a positive value for One Way weighing.")
            self._flash_weight_display("red")
            return

        truck_plate = self.entry_vars.get('truck_plate', tk.StringVar()).get().strip()

        transaction_fields = {key: var_obj.get() for key, var_obj in self.entry_vars.items()}

        master_data_keys = ["company", "truck_plate", "product", "designation", "sender", "origin", "destination", "driver"]
        for field in master_data_keys:
            if transaction_fields.get(field):
                self.auto_add_to_master(field, transaction_fields[field])

        current_time_stamp = datetime.now()
        current_date = current_time_stamp.strftime("%m/%d/%Y")
        current_time = current_time_stamp.strftime("%I:%M:%S %p")

        # Calculate unit_price and total_price if computation is enabled
        unit_price = 0.0
        total_price = 0.0
        if self.price_computation_enabled_var.get():
            try:
                weight_kg = float(current_weight)
                unit_price, total_price = self.calculate_price(weight_kg)
            except (ValueError, AttributeError):
                unit_price = 0.0
                total_price = 0.0

        next_ticket = self.get_next_ticket_no()
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO transactions (company, truck_plate, product, designation, sender, origin, destination, driver,
                        gross_weight, tare_weight, net_weight, gross_date, gross_time,
                        tare_date, tare_time, weight_type, ticket_no, status, operator, unit_price, total_price, gross_total_price, tare_total_price)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    transaction_fields.get('company'), transaction_fields.get('truck_plate'),
                    transaction_fields.get('product'), transaction_fields.get('designation'), transaction_fields.get('sender'),
                    transaction_fields.get('origin'), transaction_fields.get('destination'), transaction_fields.get('driver'),
                    current_weight, 0.0, current_weight,
                    current_date, current_time,
                    current_date, current_time,
                    "ONE WAY WEIGHING", next_ticket, "Completed", self.logged_in_user,
                    unit_price, total_price, total_price, 0.0  # For one-way, gross_total_price = total_price, tare_total_price = 0
                ))
                trans_id = cursor.lastrowid
                conn.commit()
            self._flash_weight_display("green")
            
            # Prompt user to show print preview after successful save
            self._prompt_print_ticket_after_save(trans_id)
            
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to record One Way weight: {e}")
        finally:
            self._reset_entry_form_for_new_entry()
            self.load_transactions()

    def record_two_way_weight(self):
        """
        Records a new two-way weighing transaction. This can be a new 'gross' or a 'tare' for a pending gross.
        It also handles a swap if the tare weight is greater than the gross weight.
        """
        if not self.is_app_activated:
            self.msg_box.showerror("Activation Required", "Please activate the application to save transactions.")
            self._flash_weight_display("red")
            return

        try:
            current_weight = float(self.weight_value.get())
        except ValueError:
            self.msg_box.showerror("Invalid Weight", "Captured weight is not a valid number. Cannot record.")
            self._flash_weight_display("red")
            return

        if current_weight <= 0:
            self.msg_box.showerror("Error", "Weight must be a positive value.")
            self._flash_weight_display("red")
            return

        truck_plate = self.entry_vars.get('truck_plate', tk.StringVar()).get().strip()

        transaction_fields = {key: var_obj.get() for key, var_obj in self.entry_vars.items()}

        master_data_keys = ["company", "truck_plate", "product", "designation", "sender", "origin", "destination", "driver"]
        for field in master_data_keys:
            if transaction_fields.get(field):
                self.auto_add_to_master(field, transaction_fields[field])

        current_time_stamp = datetime.now()
        current_date = current_time_stamp.strftime("%m/%d/%Y")
        current_time = current_time_stamp.strftime("%I:%M:%S %p")

        if self.current_edit_transaction_id:
            # We are completing a pending transaction (Tare weighing)
            try:
                with sqlite3.connect(DB_FILE) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT gross_weight, gross_date, gross_time, tare_weight, tare_date, tare_time, ticket_no FROM transactions WHERE id = ?", (self.current_edit_transaction_id,))
                    existing_data = cursor.fetchone()

                    if existing_data:
                        gross_weight_from_db = existing_data[0]
                        original_gross_date = existing_data[1] # Corrected index for gross_date
                        original_gross_time = existing_data[2] # Corrected index for gross_time
                        ticket_no_from_db = existing_data[6] # Corrected index for ticket_no

                        tare_weight = current_weight
                        net_weight = gross_weight_from_db - tare_weight
                        status = "Completed"

                        # Calculate unit_price and total_price if computation is enabled
                        unit_price = 0.0
                        total_price = 0.0
                        gross_total_price = 0.0
                        tare_total_price = 0.0
                        if self.price_computation_enabled_var.get():
                            try:
                                # For two-way transactions, use current weight (tare) for total_price calculation
                                current_weight_for_price = float(tare_weight)
                                unit_price, total_price = self.calculate_price(current_weight_for_price)
                                
                                # Calculate separate prices for gross and tare weights
                                gross_unit_price, gross_total_price = self.calculate_price(float(gross_weight_from_db))
                                tare_unit_price, tare_total_price = self.calculate_price(float(tare_weight))
                            except (ValueError, AttributeError):
                                unit_price = 0.0
                                total_price = 0.0
                                gross_total_price = 0.0
                                tare_total_price = 0.0

                        # Handle the case where tare > gross, and swap them
                        if tare_weight > gross_weight_from_db:
                            temp_gross_weight = tare_weight
                            tare_weight = gross_weight_from_db
                            gross_weight_from_db = temp_gross_weight
                            
                            temp_gross_date_for_swap = original_gross_date
                            temp_gross_time_for_swap = original_gross_time

                            original_gross_date = current_date
                            original_gross_time = current_time
                            current_date = temp_gross_date_for_swap
                            current_time = temp_gross_time_for_swap

                            net_weight = gross_weight_from_db - tare_weight
                            weight_type = "TARE-GROSS"
                            
                            # Recalculate total_price for swapped weights
                            if self.price_computation_enabled_var.get():
                                try:
                                    # For two-way transactions, use current weight (tare) for total_price calculation
                                    current_weight_for_price = float(tare_weight)
                                    unit_price, total_price = self.calculate_price(current_weight_for_price)
                                    
                                    # Recalculate separate prices for swapped gross and tare weights
                                    gross_unit_price, gross_total_price = self.calculate_price(float(gross_weight_from_db))
                                    tare_unit_price, tare_total_price = self.calculate_price(float(tare_weight))
                                except (ValueError, AttributeError):
                                    unit_price = 0.0
                                    total_price = 0.0
                                    gross_total_price = 0.0
                                    tare_total_price = 0.0
                        else:
                            weight_type = "GROSS-TARE"

                        cursor.execute("""
                            UPDATE transactions SET
                                company = ?, truck_plate = ?, product = ?, designation = ?, sender = ?,
                                origin = ?, destination = ?, driver = ?,
                                gross_weight = ?, tare_weight = ?, net_weight = ?,
                                gross_date = ?, gross_time = ?, tare_date = ?, tare_time = ?,
                                weight_type = ?, status = ?, operator2 = ?, unit_price = ?, total_price = ?, gross_total_price = ?, tare_total_price = ?,
                                timestamp = CURRENT_TIMESTAMP
                            WHERE id = ?
                        """, (
                            transaction_fields.get('company'), transaction_fields.get('truck_plate'),
                            transaction_fields.get('product'), transaction_fields.get('designation'), transaction_fields.get('sender'),
                            transaction_fields.get('origin'), transaction_fields.get('destination'), transaction_fields.get('driver'),
                            gross_weight_from_db, tare_weight, net_weight,
                            original_gross_date, original_gross_time,
                            current_date, current_time,
                            weight_type, status, self.logged_in_user, unit_price, total_price, gross_total_price, tare_total_price,
                            self.current_edit_transaction_id
                        ))
                        conn.commit()
                        self._flash_weight_display("green")
                        
                        # Prompt user to show print preview after successful tare completion
                        self._prompt_print_ticket_after_save(self.current_edit_transaction_id)
                        
                    else:
                        self.msg_box.showerror("Error", "Pending transaction not found for tare weighing.")
                        self._flash_weight_display("red")
            except sqlite3.Error as e:
                self.msg_box.showerror("Database Error", f"Failed to record Tare weight: {e}")
                self._flash_weight_display("red")
            finally:
                self._reset_entry_form_for_new_entry()
                self.load_pending()
                self.load_transactions()

        else:
            # New Gross Weighing
            next_ticket = self.get_next_ticket_no()
            status = "Pending"
            weight_type = "GROSS-TARE"
            
            # Calculate gross_total_price for new gross transaction
            gross_total_price = 0.0
            if self.price_computation_enabled_var.get():
                try:
                    gross_unit_price, gross_total_price = self.calculate_price(float(current_weight))
                except (ValueError, AttributeError):
                    gross_total_price = 0.0
            
            try:
                with sqlite3.connect(DB_FILE) as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        INSERT INTO transactions (company, truck_plate, product, designation, sender, origin, destination, driver,
                            gross_weight, tare_weight, net_weight, gross_date, gross_time,
                            tare_date, tare_time, weight_type, ticket_no, status, operator, unit_price, total_price, gross_total_price, tare_total_price)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        transaction_fields.get('company'), transaction_fields.get('truck_plate'),
                        transaction_fields.get('product'), transaction_fields.get('designation'), transaction_fields.get('sender'),
                        transaction_fields.get('origin'), transaction_fields.get('destination'), transaction_fields.get('driver'),
                        current_weight, None, None,
                        current_date, current_time,
                        None, None,
                        weight_type, next_ticket, status, self.logged_in_user, 0.0, 0.0, gross_total_price, 0.0
                    ))
                    trans_id = cursor.lastrowid
                    conn.commit()
                self._flash_weight_display("green")
                
                # Prompt user to show print preview after successful gross save
                self._prompt_print_ticket_after_save(trans_id)
                
            except sqlite3.Error as e:
                self.msg_box.showerror("Database Error", f"Failed to record Gross weight: {e}")
            finally:
                self._reset_entry_form_for_new_entry()
                self.load_pending()
                self.load_transactions()
# Part 7: Pending Records Methods

    def build_pending_tab(self, tab: ttk.Frame):
        """
        Builds the GUI for the Pending Records tab.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        # Add Search & Filter section for Pending Records
        search_filter_frame = ttk.LabelFrame(tab, text="Search & Filter Pending Records", padding="10")
        search_filter_frame.pack(padx=10, pady=5, fill="x", expand=False)

        # Row 0: Search Query with real-time search
        ttk.Label(search_filter_frame, text="Search Query:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.pending_search_query_var = tk.StringVar()
        self.pending_search_entry = ttk.Entry(search_filter_frame, textvariable=self.pending_search_query_var, style="TEntry")
        self.pending_search_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        # Real-time search with debouncing
        self.pending_search_entry.bind("<KeyRelease>", self._debounced_pending_search)
        self.pending_search_entry.bind("<Return>", lambda event: self.load_pending(self.pending_search_query_var.get()))

        # Search status label
        self.pending_search_status_var = tk.StringVar(value="Ready")
        self.pending_search_status_label = ttk.Label(search_filter_frame, textvariable=self.pending_search_status_var, style="TLabel", foreground="gray")
        self.pending_search_status_label.grid(row=0, column=2, sticky="w", padx=5, pady=2)

        # Place Search and Clear Search buttons side-by-side (column 3)
        btn_search_frame = ttk.Frame(search_filter_frame)
        btn_search_frame.grid(row=0, column=3, padx=5, pady=2, sticky="w")

        ttk.Button(btn_search_frame, text="Search", command=lambda: self.load_pending(self.pending_search_query_var.get()), style="TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_search_frame, text="Clear Search", command=self.clear_pending_search, style="TButton").pack(side=tk.LEFT, padx=5)

        # Row 1: Date Filters with better layout
        date_frame = ttk.Frame(search_filter_frame)
        date_frame.grid(row=1, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        
        ttk.Label(date_frame, text="From Date:", style="TLabel").pack(side=tk.LEFT, padx=(0, 5))
        # Initialize with current month start date
        month_start, month_end = self._get_current_month_range()
        self.pending_from_date_var = tk.StringVar(value=month_start.strftime("%Y-%m-%d"))
        self.pending_from_date_entry = ttk.Entry(date_frame, textvariable=self.pending_from_date_var, style="TEntry", width=12)
        self.pending_from_date_entry.pack(side=tk.LEFT, padx=5)
        # Add date picker button
        ttk.Button(date_frame, text="📅", command=lambda: self._show_date_picker(self.pending_from_date_var), style="TButton", width=3).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(date_frame, text="To Date:", style="TLabel").pack(side=tk.LEFT, padx=(0, 5))
        # Initialize with current month end date
        self.pending_to_date_var = tk.StringVar(value=month_end.strftime("%Y-%m-%d"))
        self.pending_to_date_entry = ttk.Entry(date_frame, textvariable=self.pending_to_date_var, style="TEntry", width=12)
        self.pending_to_date_entry.pack(side=tk.LEFT, padx=5)
        # Add date picker button
        ttk.Button(date_frame, text="📅", command=lambda: self._show_date_picker(self.pending_to_date_var), style="TButton", width=3).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(date_frame, text="Apply Filters", command=self._apply_pending_filters, style="TButton").pack(side=tk.LEFT, padx=(10, 0))

        # Configure columns to allow expansion
        search_filter_frame.grid_columnconfigure(1, weight=1)

        # Extended columns to show more transaction details
        columns = ("ID", "Ticket No", "Company", "Truck Plate", "Product", "Driver", "Gross Weight", "Gross Date/Time")
        
        # Create main container for treeview and scrollbars
        tree_container = ttk.Frame(tab)
        tree_container.pack(fill="both", expand=True, padx=10, pady=(10, 0))
        
        # Add result count label above treeview
        self.pending_result_count_var = tk.StringVar(value="No records found")
        result_count_label = ttk.Label(tree_container, textvariable=self.pending_result_count_var, style="TLabel", font=("Arial", 9, "bold"))
        result_count_label.pack(anchor="w", pady=(0, 5))
        
        # Create treeview inside the container
        self.pending_tree = ttk.Treeview(tree_container, columns=columns, show='headings', style="Treeview")
        
        # Load saved column widths or use defaults
        pending_column_widths = self.config.get("pending_column_widths", {
            "ID": 50,
            "Ticket No": 80,
            "Company": 150,
            "Truck Plate": 120,
            "Product": 120,
            "Driver": 120,
            "Gross Weight": 100,
            "Gross Date/Time": 150
        })
        
        for col in columns:
            self.pending_tree.heading(col, text=col)
            width = pending_column_widths.get(col, 100)
            self.pending_tree.column(col, width=width, anchor=tk.CENTER)
        
        # Create vertical scrollbar
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.pending_tree.yview)
        self.pending_tree.configure(yscrollcommand=vsb.set)
        
        # Pack treeview and vertical scrollbar in container
        self.pending_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Bind column resize event to save widths
        self.pending_tree.bind("<ButtonRelease-1>", self._on_pending_column_resize)
        self.pending_tree.bind("<Double-Button-1>", lambda event: self.select_pending_for_tare_from_treeview())

        # Create horizontal scrollbar container above buttons
        hscroll_container = ttk.Frame(tab)
        hscroll_container.pack(fill="x", padx=10, pady=(0, 5))
        
        # Create horizontal scrollbar in its own container
        hsb = ttk.Scrollbar(hscroll_container, orient="horizontal", command=self.pending_tree.xview)
        self.pending_tree.configure(xscrollcommand=hsb.set)
        hsb.pack(fill="x")

        # Button frame at the bottom with proper padding
        button_frame = ttk.Frame(tab)
        button_frame.pack(pady=5, fill="x", padx=10)

        # Pack the buttons in the desired order:
        # 1. PRINT Button (first)
        self.print_to_file_pending_button = ttk.Button(button_frame, text="PRINT", command=self._debounced_print_pending_to_pdf, style="TButton")
        self.print_to_file_pending_button.pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        # 2. Recall for Second Weighing Button
        self.recall_for_second_weighing_button = ttk.Button(button_frame, text="Recall for Second Weighing", command=self._debounced_select_pending_for_tare, style="TButton")
        self.recall_for_second_weighing_button.pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        # 3. Delete Pending Button
        ttk.Button(button_frame, text="Delete Pending", command=self.delete_pending_transaction, style="TButton").pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        # 4. Refresh Pending List Button - HIDDEN
        # ttk.Button(button_frame, text="Refresh Pending List", command=self.load_pending, style="TButton").pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        self.load_pending()  # Initial load with date filters

    def _debounced_select_pending_for_tare(self):
        """
        Wrapper to prevent double-clicking on the button.
        """
        self.recall_for_second_weighing_button.config(state=tk.DISABLED)
        try:
            self.select_pending_for_tare_from_treeview()
        finally:
            self.root.after(500, lambda: self.recall_for_second_weighing_button.config(state=tk.NORMAL))

    def _debounced_print_selected_pending_ticket(self):
        """
        Wrapper to prevent double-clicking on the button.
        """
        self.print_selected_pending_button.config(state=tk.DISABLED)
        try:
            self.print_selected_ticket(self.pending_tree, "Pending", to_file=False)
        finally:
            self.root.after(500, lambda: self.print_selected_pending_button.config(state=tk.NORMAL))

    def _debounced_print_pending_to_pdf(self):
        """
        Wrapper to prevent double-clicking on the button. Now opens print preview instead of PDF.
        """
        self.print_to_file_pending_button.config(state=tk.DISABLED)
        try:
            self._open_preview_for_selection(self.pending_tree, "Pending")
        finally:
            self.root.after(500, lambda: self.print_to_file_pending_button.config(state=tk.NORMAL))


    def select_pending_for_tare_from_treeview(self):
        """
        Loads a selected pending transaction into the main entry form for a second weighing.
        """
        selected_item = self.pending_tree.selection()
        if not selected_item:
            self.msg_box.showerror("Error", "Please select a pending transaction to recall for second weighing.")
            return

        trans_id = self.pending_tree.item(selected_item[0])['values'][0]
        self._load_pending_for_tare(trans_id)
        self.main_notebook.select(0)


    def delete_pending_transaction(self):
        """
        Deletes a selected pending transaction from the database.
        Only allowed for admin and operator roles.
        """
        # Check user permissions - only admin and operator can delete
        if self.current_user_role not in ['admin', 'operator']:
            self.msg_box.showerror("Access Denied", "Only administrators and operators can delete pending transactions.")
            return
        
        selected = self.pending_tree.selection()
        if not selected:
            self.msg_box.showerror("Error", "No pending transaction selected to delete.")
            return

        trans_id = self.pending_tree.item(selected[0])['values'][0]
        
        # Use blocking dialog
        result = self.msg_box.askyesno("Confirm Delete", f"Are you sure you want to delete pending transaction ID {trans_id}?")
        if result:
            try:
                with sqlite3.connect(DB_FILE) as conn:
                    conn.execute("DELETE FROM transactions WHERE id = ?", (trans_id,))
                    conn.commit()
                self.msg_box.showinfo("Deleted", f"Pending transaction ID {trans_id} deleted.")
                self.load_pending()
            except sqlite3.Error as e:
                self.msg_box.showerror("Database Error", f"Failed to delete transaction: {e}")


    def load_pending(self, search_query=None, skip_date_filter=False):
        """
        Loads all pending transactions into the Treeview.
        Supports enhanced search and filtering functionality.
        """
        # Show searching status
        if hasattr(self, 'pending_search_status_var'):
            self.pending_search_status_var.set("Searching...")
            self.pending_search_status_label.config(foreground="blue")
        
        for i in self.pending_tree.get_children():
            self.pending_tree.delete(i)
        
        # Use provided search query or get from the search entry
        if search_query is None:
            search_query = self.pending_search_query_var.get().strip() if hasattr(self, 'pending_search_query_var') else ""
        
        from_date = self.pending_from_date_var.get() if hasattr(self, 'pending_from_date_var') else ""
        to_date = self.pending_to_date_var.get() if hasattr(self, 'pending_to_date_var') else ""
        
        # Skip date filter if requested
        if skip_date_filter:
            from_date = ""
            to_date = ""
            
        status_filter = self.pending_status_var.get() if hasattr(self, 'pending_status_var') else "All"
        
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                base_query = "SELECT id, ticket_no, company, truck_plate, product, driver, gross_weight, gross_date, gross_time, status FROM transactions"
                query_params = []
                where_clauses = []

                # Always filter for pending transactions only
                where_clauses.append("status = ?")
                query_params.append("Pending")

                # Apply additional status filter if specified (for future flexibility)
                if status_filter != "All" and status_filter != "Pending":
                    where_clauses[-1] = "status = ?"
                    query_params[-1] = status_filter

                if search_query:
                    # Enhanced search with better pattern matching
                    text_search_fields = ["company", "truck_plate", "product", "driver"]
                    text_search_pattern = f"%{search_query}%"
                    search_sub_clauses = [f"UPPER({field}) LIKE ?" for field in text_search_fields]
                    search_params = [text_search_pattern] * len(text_search_fields)
                    
                    # Also search by exact ticket number if numeric
                    if search_query.isdigit():
                        search_sub_clauses.append("ticket_no = ?")
                        search_params.append(int(search_query))
                    
                    where_clauses.append("(" + " OR ".join(search_sub_clauses) + ")")
                    query_params.extend(search_params)

                if from_date and to_date:
                    try:
                        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
                        to_dt = datetime.strptime(to_date, "%Y-%m-%d")
                        # Convert to MM/DD/YYYY format to match database format
                        from_date_converted = from_dt.strftime("%m/%d/%Y")
                        to_date_converted = to_dt.strftime("%m/%d/%Y")
                        where_clauses.append("(gross_date >= ? AND gross_date <= ?)")
                        query_params.extend([from_date_converted, to_date_converted])
                    except ValueError:
                        self.msg_box.showwarning("Date Format Error", "Invalid date format. Please use YYYY-MM-DD. Date filter ignored.")

                if where_clauses:
                    query = base_query + " WHERE " + " AND ".join(where_clauses)
                else:
                    query = base_query

                query += " ORDER BY timestamp DESC"
                
                cursor.execute(query, query_params)
                results = cursor.fetchall()
                
                # Update result count
                if hasattr(self, 'pending_result_count_var'):
                    if results:
                        self.pending_result_count_var.set(f"Found {len(results)} record{'s' if len(results) != 1 else ''}")
                    else:
                        self.pending_result_count_var.set("No records found")
                
                for row in results:
                    formatted_row = list(row)
                    format_string = f"%.{self.decimal_places}f"
                    formatted_row[6] = format_string % row[6] if row[6] is not None else "N/A"

                    gross_datetime = f"{row[7]} {row[8]}" if row[7] and row[8] else "N/A"

                    display_row = [
                        formatted_row[0],  # ID
                        formatted_row[1],  # Ticket No
                        formatted_row[2] or "N/A",  # Company
                        formatted_row[3] or "N/A",  # Truck Plate
                        formatted_row[4] or "N/A",  # Product
                        formatted_row[5] or "N/A",  # Driver
                        formatted_row[6],  # Gross Weight
                        gross_datetime  # Gross Date/Time
                    ]

                    self.pending_tree.insert("", "end", values=display_row)
                    
        except sqlite3.Error as e:
            # Handle silently, as this is a refresh function.
            pass
        finally:
            # Update search status
            if hasattr(self, 'pending_search_status_var'):
                self.pending_search_status_var.set("Ready")
                self.pending_search_status_label.config(foreground="gray")

    def clear_pending_search(self):
        """Clears the search fields for the Pending Records tab and reloads all pending transactions."""
        if hasattr(self, 'pending_search_query_var'):
            self.pending_search_query_var.set("")
        if hasattr(self, 'pending_from_date_var'):
            self.pending_from_date_var.set("")
        if hasattr(self, 'pending_to_date_var'):
            self.pending_to_date_var.set("")
        if hasattr(self, 'pending_status_var'):
            self.pending_status_var.set("All")
        self.load_pending()

    def _apply_pending_filters(self):
        """Apply date and status filters for pending records."""
        self.load_pending(self.pending_search_query_var.get())


    def build_completed_records_tab(self, tab: ttk.Frame):
        """Builds the GUI for the main Completed Records tab, including search and filters."""
        for widget in tab.winfo_children():
            widget.destroy()

        # FIX: Define search_filter_frame before it is used.
        search_filter_frame = ttk.LabelFrame(tab, text="Search & Filter Completed Records", padding="10")
        search_filter_frame.pack(padx=10, pady=5, fill="x", expand=False)

        # Row 0: Search Query with real-time search
        ttk.Label(search_filter_frame, text="Search Query:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.search_query_var.trace_add("write", lambda name, index, mode, var=self.search_query_var: var.set(var.get().upper()))
        self.search_entry = ttk.Entry(search_filter_frame, textvariable=self.search_query_var, style="TEntry")
        self.search_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        # Real-time search with debouncing
        self.search_entry.bind("<KeyRelease>", self._debounced_transaction_search)
        self.search_entry.bind("<Return>", lambda event: self.load_transactions(self.search_query_var.get()))

        # Search status label
        self.transaction_search_status_var = tk.StringVar(value="Ready")
        self.transaction_search_status_label = ttk.Label(search_filter_frame, textvariable=self.transaction_search_status_var, style="TLabel", foreground="gray")
        self.transaction_search_status_label.grid(row=0, column=2, sticky="w", padx=5, pady=2)

        # Place Search and Clear Search buttons side-by-side (column 3)
        btn_search_frame = ttk.Frame(search_filter_frame)
        btn_search_frame.grid(row=0, column=3, padx=5, pady=2, sticky="w")

        ttk.Button(btn_search_frame, text="Search", command=lambda: self.load_transactions(self.search_query_var.get()), style="TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_search_frame, text="Clear Search", command=self.clear_search, style="TButton").pack(side=tk.LEFT, padx=5)

        # Row 1: Status, Weight Type, and Date Filters
        filter_row1 = ttk.Frame(search_filter_frame)
        filter_row1.grid(row=1, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        
        ttk.Label(filter_row1, text="Status:", style="TLabel").pack(side=tk.LEFT, padx=(0, 5))
        self.transaction_status_var = tk.StringVar(value="All")
        status_combo = ttk.Combobox(filter_row1, textvariable=self.transaction_status_var, values=["All", "Pending", "Completed"], state="readonly", width=12)
        status_combo.pack(side=tk.LEFT, padx=5)
        status_combo.bind("<<ComboboxSelected>>", lambda event: self._apply_transaction_filters())
        
        ttk.Label(filter_row1, text="Weight Type:", style="TLabel").pack(side=tk.LEFT, padx=(20, 5))
        self.weight_type_var = tk.StringVar(value="All")
        weight_type_combo = ttk.Combobox(filter_row1, textvariable=self.weight_type_var, values=["All", "ONE WAY WEIGHING", "TWO WAY WEIGHING"], state="readonly", width=15)
        weight_type_combo.pack(side=tk.LEFT, padx=5)
        weight_type_combo.bind("<<ComboboxSelected>>", lambda event: self._apply_transaction_filters())
        
        # Date filters in the same row
        ttk.Label(filter_row1, text="From Date:", style="TLabel").pack(side=tk.LEFT, padx=(20, 5))
        self.from_date_entry = ttk.Entry(filter_row1, textvariable=self.from_date_var, style="TEntry", width=12)
        self.from_date_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_row1, text="📅", command=lambda: self._show_date_picker(self.from_date_var), style="TButton", width=3).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(filter_row1, text="To Date:", style="TLabel").pack(side=tk.LEFT, padx=(0, 5))
        self.to_date_entry = ttk.Entry(filter_row1, textvariable=self.to_date_var, style="TEntry", width=12)
        self.to_date_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_row1, text="📅", command=lambda: self._show_date_picker(self.to_date_var), style="TButton", width=3).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(filter_row1, text="Apply Filters", command=self._apply_transaction_filters, style="TButton").pack(side=tk.LEFT, padx=(10, 0))

        # Configure columns to allow expansion
        search_filter_frame.grid_columnconfigure(1, weight=1)

        # Extended columns to show more transaction details
        columns = ("ID", "Ticket No", "Company", "Truck Plate", "Product", "Driver", "Gross Weight", "Tare Weight", "Net Weight", "Gross Date/Time", "Tare Date/Time")
        
        # Create main container for treeview and scrollbars
        tree_container = ttk.Frame(tab)
        tree_container.pack(fill="both", expand=True, padx=10, pady=(10, 0))
        
        # Add result count label above treeview
        self.transaction_result_count_var = tk.StringVar(value="No records found")
        result_count_label = ttk.Label(tree_container, textvariable=self.transaction_result_count_var, style="TLabel", font=("Arial", 9, "bold"))
        result_count_label.pack(anchor="w", pady=(0, 5))
        
        # Create treeview inside the container
        self.transaction_tree = ttk.Treeview(tree_container, columns=columns, show='headings', style="Treeview")
        
        # Load saved column widths or use defaults
        transaction_column_widths = self.config.get("transaction_column_widths", {
            "ID": 40,
            "Ticket No": 80,
            "Company": 150,
            "Truck Plate": 120,
            "Product": 120,
            "Driver": 120,
            "Gross Weight": 100,
            "Tare Weight": 100,
            "Net Weight": 100,
            "Gross Date/Time": 140,
            "Tare Date/Time": 140
        })
        
        for col in columns:
            self.transaction_tree.heading(col, text=col)
            width = transaction_column_widths.get(col, 100)
            self.transaction_tree.column(col, width=width, anchor=tk.CENTER)
        
        # Create vertical scrollbar
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.transaction_tree.yview)
        self.transaction_tree.configure(yscrollcommand=vsb.set)
        
        # Pack treeview and vertical scrollbar in container
        self.transaction_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Bind column resize event to save widths
        self.transaction_tree.bind("<ButtonRelease-1>", self._on_transaction_column_resize)

        # Create horizontal scrollbar container above buttons
        hscroll_container = ttk.Frame(tab)
        hscroll_container.pack(fill="x", padx=10, pady=(0, 5))
        
        # Create horizontal scrollbar in its own container
        hsb = ttk.Scrollbar(hscroll_container, orient="horizontal", command=self.transaction_tree.xview)
        self.transaction_tree.configure(xscrollcommand=hsb.set)
        hsb.pack(fill="x")

        # Button frame at the bottom with proper padding
        button_frame = ttk.Frame(tab)
        button_frame.pack(pady=5, fill="x", padx=10)
        
        # Updated buttons with expand and fill parameters in desired order:
        # 1. PRINT Button (first)
        self.print_to_file_transaction_button = ttk.Button(button_frame, text="PRINT", command=self._debounced_print_transaction_to_pdf, style="TButton")
        self.print_to_file_transaction_button.pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        # 2. Refresh All Completed Records Button
        ttk.Button(button_frame, text="Refresh All Completed Records", command=lambda: self.load_transactions(), style="TButton").pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        # 3. Delete Selected Button (last)
        ttk.Button(button_frame, text="Delete Selected", command=self._delete_selected_transaction, style="TButton").pack(side=tk.LEFT, expand=True, fill="x", padx=2)

        self.load_transactions()  # Initial load with date filters

    def _debounced_print_selected_transaction_ticket(self):
        """
        Wrapper to prevent double-clicking on the button.
        """
        self.print_selected_transaction_button.config(state=tk.DISABLED)
        try:
            self.print_selected_ticket(self.transaction_tree, "Completed", to_file=False)
        finally:
            self.root.after(500, lambda: self.print_selected_transaction_button.config(state=tk.NORMAL))

    def _debounced_print_transaction_to_pdf(self):
        """
        Wrapper to prevent double-clicking on the button. Now opens print preview instead of PDF.
        """
        self.print_to_file_transaction_button.config(state=tk.DISABLED)
        try:
            self._open_preview_for_selection(self.transaction_tree, "Completed Records")
        finally:
            self.root.after(500, lambda: self.print_to_file_transaction_button.config(state=tk.NORMAL))

    def _apply_transaction_filters(self):
        """
        Applies search and date filters to the transactions list.
        """
        search_query = self.search_query_var.get()
        from_date = self.from_date_var.get()
        to_date = self.to_date_var.get()
        self.load_transactions(search_query=search_query, from_date=from_date, to_date=to_date)

    def clear_search(self):
        """Clears the search fields and reloads all completed records."""
        # Clear the search query for the Completed Records tab
        self.search_query_var.set("")
        # Clear the date filters for the Completed Records tab
        self.from_date_var.set("")
        self.to_date_var.set("")
        # Clear status and weight type filters
        if hasattr(self, 'transaction_status_var'):
            self.transaction_status_var.set("All")
        if hasattr(self, 'weight_type_var'):
            self.weight_type_var.set("All")
        # Reload all completed records
        self.load_transactions()
    
    def load_transactions(self, search_query: str = "", from_date: str = "", to_date: str = "", skip_date_filter=False):
        """
        Loads transactions from the database into the Treeview with improved filtering and search.
        Supports:
        - Text search across multiple fields (company, truck_plate, driver, operators, etc.)
        - Partial matching (e.g., "carlo" matches "carlo santos" or "carlo/argueles")
        - Case-insensitive search
        - Numeric search by ticket number
        - Status filtering
        - Weight type filtering
        - Date range filtering (YYYY-MM-DD format)
        """
        # Show searching status
        if hasattr(self, 'transaction_search_status_var'):
            self.transaction_search_status_var.set("Searching...")
            self.transaction_search_status_label.config(foreground="blue")
        
        for i in self.transaction_tree.get_children():
            self.transaction_tree.delete(i)

        # Get filter values
        if not search_query:
            search_query = self.search_query_var.get().strip() if hasattr(self, 'search_query_var') else ""
        if not from_date:
            from_date = self.from_date_var.get() if hasattr(self, 'from_date_var') else ""
        if not to_date:
            to_date = self.to_date_var.get() if hasattr(self, 'to_date_var') else ""
        
        # Skip date filter if requested
        if skip_date_filter:
            from_date = ""
            to_date = ""
            
        status_filter = self.transaction_status_var.get() if hasattr(self, 'transaction_status_var') else "All"
        weight_type_filter = self.weight_type_var.get() if hasattr(self, 'weight_type_var') else "All"

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                base_query = "SELECT id, ticket_no, company, truck_plate, product, driver, gross_weight, tare_weight, net_weight, gross_date, gross_time, tare_date, tare_time, status, weight_type, operator, operator2 FROM transactions"
                query_params = []
                where_clauses = []

                # Always filter for completed transactions only
                where_clauses.append("status = ?")
                query_params.append("Completed")

                # Apply status filter if specified (for future flexibility)
                if status_filter != "All" and status_filter != "Completed":
                    where_clauses[-1] = "status = ?"
                    query_params[-1] = status_filter
                
                # Apply weight type filter
                if weight_type_filter != "All":
                    where_clauses.append("weight_type = ?")
                    query_params.append(weight_type_filter)

                if search_query:
                    # Improved search with operator fields included
                    text_search_fields = ["company", "truck_plate", "product", "driver", "status", "designation", "sender", "origin", "destination", "operator", "operator2"]
                    text_search_pattern = f"%{search_query}%"

                    search_sub_clauses = [f"UPPER({field}) LIKE ?" for field in text_search_fields]
                    search_sub_params = [text_search_pattern] * len(text_search_fields)

                    if search_query.isdigit():
                        search_sub_clauses.append("ticket_no = ?")
                        search_sub_params.append(int(search_query))

                    # Also allow status matching if text matches a known status
                    if search_query.upper() in ["PENDING", "COMPLETED"]:
                        search_sub_clauses.append("UPPER(status) = ?")
                        search_sub_params.append(search_query.upper())

                    if search_sub_clauses:
                        where_clauses.append("(" + " OR ".join(search_sub_clauses) + ")")
                        query_params.extend(search_sub_params)

                if from_date and to_date:
                    try:
                        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
                        to_dt = datetime.strptime(to_date, "%Y-%m-%d")
                        # Convert to MM/DD/YYYY format to match database format
                        from_date_converted = from_dt.strftime("%m/%d/%Y")
                        to_date_converted = to_dt.strftime("%m/%d/%Y")
                        where_clauses.append("(gross_date >= ? AND gross_date <= ?)")
                        query_params.extend([from_date_converted, to_date_converted])
                    except ValueError:
                        self.msg_box.showwarning("Date Format Error", "Invalid date format. Please use YYYY-MM-DD. Date filter ignored.")

                if where_clauses:
                    query = base_query + " WHERE " + " AND ".join(where_clauses)
                else:
                    query = base_query

                query += " ORDER BY timestamp DESC"

                cursor.execute(query, query_params)
                results = cursor.fetchall()
                
                # Update result count
                if hasattr(self, 'transaction_result_count_var'):
                    if results:
                        self.transaction_result_count_var.set(f"Found {len(results)} record{'s' if len(results) != 1 else ''}")
                    else:
                        self.transaction_result_count_var.set("No records found")
                
                for row in results:
                    (
                        id, ticket_no, company, truck_plate, product, driver, gross_weight, tare_weight, net_weight,
                        gross_date, gross_time, tare_date, tare_time, status, weight_type, operator, operator2
                    ) = row

                    format_string = f"%.{self.decimal_places}f"

                    formatted_gross_weight = format_string % gross_weight if gross_weight is not None else "N/A"
                    formatted_tare_weight = format_string % tare_weight if tare_weight is not None else "N/A"
                    formatted_net_weight = format_string % net_weight if net_weight is not None else "N/A"

                    consolidated_gross_datetime = f"{gross_date} {gross_time}" if gross_date and gross_time else "N/A"

                    consolidated_tare_datetime = f"{tare_date} {tare_time}" if tare_date and tare_time else "N/A"

                    self.transaction_tree.insert("", "end", values=(
                        id, ticket_no, company or "N/A", truck_plate or "N/A", 
                        product or "N/A", driver or "N/A",
                        formatted_gross_weight, formatted_tare_weight, formatted_net_weight,
                        consolidated_gross_datetime, consolidated_tare_datetime
                    ))
        except sqlite3.Error as e:
            # Handle silently, as this is a refresh function.
            pass
        finally:
            # Update search status
            if hasattr(self, 'transaction_search_status_var'):
                self.transaction_search_status_var.set("Ready")
                self.transaction_search_status_label.config(foreground="gray")

    def _edit_selected_transaction(self):
        """
        Loads the selected transaction into the entry form for editing.
        """
        selected_item = self.transaction_tree.selection()
        if not selected_item:
            self.msg_box.showerror("Error", "No transaction selected for editing.")
            return

        trans_id = self.transaction_tree.item(selected_item[0])['values'][0]
        self._load_transaction_for_editing(trans_id)

    def _delete_selected_transaction(self):
        """
        Deletes a selected transaction from the database.
        Only allowed for admin and operator roles.
        """
        # Check user permissions - only admin and operator can delete
        if self.current_user_role not in ['admin', 'operator']:
            self.msg_box.showerror("Access Denied", "Only administrators and operators can delete transactions.")
            return
        
        selected_item = self.transaction_tree.selection()
        if not selected_item:
            self.msg_box.showerror("Error", "No transaction selected for deletion.")
            return

        trans_id = self.transaction_tree.item(selected_item[0])['values'][0]
        
        # Use blocking dialog
        result = self.msg_box.askyesno("Confirm Delete", f"Are you sure you want to delete transaction ID {trans_id}? This action is irreversible!")
        if result:
            try:
                with sqlite3.connect(DB_FILE) as conn:
                    conn.execute("DELETE FROM transactions WHERE id = ?", (trans_id,))
                    conn.commit()
                self.msg_box.showinfo("Deleted", f"Transaction ID {trans_id} deleted.")
                self.load_transactions()
                if self.current_edit_transaction_id == trans_id:
                    self._reset_entry_form_for_new_entry()
            except sqlite3.Error as e:
                self.msg_box.showerror("Database Error", f"Failed to delete transaction: {e}")

    def build_report_tab(self, tab: ttk.Frame):
        """Builds the GUI for the Reports tab, including search, filters, and export options."""
        for widget in tab.winfo_children():
            widget.destroy()

        # FIX: Define filter_frame before it is used.
        filter_frame = ttk.LabelFrame(tab, text="Report Filters", padding="10")
        filter_frame.pack(padx=10, pady=10, fill="x")

        # Row 0: Search Query with real-time search
        ttk.Label(filter_frame, text="Search Query:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        report_search_entry = ttk.Entry(filter_frame, textvariable=self.report_search_query_var)
        report_search_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        # Real-time search with debouncing
        report_search_entry.bind("<KeyRelease>", self._debounced_report_search)
        report_search_entry.bind("<Return>", lambda event: self._search_report_transactions())

        # Search status label
        self.report_search_status_var = tk.StringVar(value="Ready")
        self.report_search_status_label = ttk.Label(filter_frame, textvariable=self.report_search_status_var, style="TLabel", foreground="gray")
        self.report_search_status_label.grid(row=0, column=2, sticky="w", padx=5, pady=2)

        # Place Search and Clear Search buttons side-by-side (column 3)
        btn_search_frame = ttk.Frame(filter_frame)
        btn_search_frame.grid(row=0, column=3, padx=5, pady=2, sticky="w")
        ttk.Button(btn_search_frame, text="Search", command=self._search_report_transactions, style="TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_search_frame, text="Clear Search", command=self.clear_report_search, style="TButton").pack(side=tk.LEFT, padx=5)

        # Row 1: Status, Weight Type, and Date Filters
        filter_row1 = ttk.Frame(filter_frame)
        filter_row1.grid(row=1, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        
        ttk.Label(filter_row1, text="Status:", style="TLabel").pack(side=tk.LEFT, padx=(0, 5))
        self.report_status_var = tk.StringVar(value="All")
        status_combo = ttk.Combobox(filter_row1, textvariable=self.report_status_var, values=["All", "Pending", "Completed"], state="readonly", width=12)
        status_combo.pack(side=tk.LEFT, padx=5)
        status_combo.bind("<<ComboboxSelected>>", lambda event: self._search_report_transactions())
        
        ttk.Label(filter_row1, text="Weight Type:", style="TLabel").pack(side=tk.LEFT, padx=(20, 5))
        self.report_weight_type_var = tk.StringVar(value="All")
        weight_type_combo = ttk.Combobox(filter_row1, textvariable=self.report_weight_type_var, values=["All", "ONE WAY WEIGHING", "TWO WAY WEIGHING"], state="readonly", width=15)
        weight_type_combo.pack(side=tk.LEFT, padx=5)
        weight_type_combo.bind("<<ComboboxSelected>>", lambda event: self._search_report_transactions())
        
        # Date filters in the same row
        ttk.Label(filter_row1, text="From Date:", style="TLabel").pack(side=tk.LEFT, padx=(20, 5))
        report_from_date_entry = ttk.Entry(filter_row1, textvariable=self.report_from_date_var, width=12)
        report_from_date_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_row1, text="📅", command=lambda: self._show_date_picker(self.report_from_date_var), style="TButton", width=3).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(filter_row1, text="To Date:", style="TLabel").pack(side=tk.LEFT, padx=(0, 5))
        report_to_date_entry = ttk.Entry(filter_row1, textvariable=self.report_to_date_var, width=12)
        report_to_date_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_row1, text="📅", command=lambda: self._show_date_picker(self.report_to_date_var), style="TButton", width=3).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(filter_row1, text="Apply Filters", command=self._search_report_transactions, style="TButton").pack(side=tk.LEFT, padx=(10, 0))

        # Configure columns to allow expansion
        filter_frame.grid_columnconfigure(1, weight=1)

        # Extended columns to show more transaction details
        columns = ("ID", "Ticket No", "Company", "Truck Plate", "Product", "Driver", "Gross Weight", "Tare Weight", "Net Weight", "Gross Date/Time", "Tare Date/Time", "Status")
        
        # Create main container for treeview and scrollbars
        tree_container = ttk.Frame(tab)
        tree_container.pack(fill="both", expand=True, padx=10, pady=(10, 0))
        
        # Add result count label above treeview
        self.report_result_count_var = tk.StringVar(value="No records found")
        result_count_label = ttk.Label(tree_container, textvariable=self.report_result_count_var, style="TLabel", font=("Arial", 9, "bold"))
        result_count_label.pack(anchor="w", pady=(0, 5))
        
        # Create treeview inside the container
        self.report_tree = ttk.Treeview(tree_container, columns=columns, show='headings', selectmode="extended", style="Treeview")
        
        # Load saved column widths or use defaults
        report_column_widths = self.config.get("report_column_widths", {
            "ID": 50,
            "Ticket No": 80,
            "Company": 150,
            "Truck Plate": 120,
            "Product": 120,
            "Driver": 120,
            "Gross Weight": 100,
            "Tare Weight": 100,
            "Net Weight": 100,
            "Gross Date/Time": 160,
            "Tare Date/Time": 160,
            "Status": 80
        })
        
        for col in columns:
            self.report_tree.heading(col, text=col)
            width = report_column_widths.get(col, 100)
            self.report_tree.column(col, width=width, anchor=tk.CENTER)
        
        # Create vertical scrollbar
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.report_tree.yview)
        self.report_tree.configure(yscrollcommand=vsb.set)
        
        # Pack treeview and vertical scrollbar in container
        self.report_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Bind column resize event to save widths
        self.report_tree.bind("<ButtonRelease-1>", self._on_report_column_resize)

        # Create horizontal scrollbar container above buttons
        hscroll_container = ttk.Frame(tab)
        hscroll_container.pack(fill="x", padx=10, pady=(0, 5))
        
        # Create horizontal scrollbar in its own container
        hsb = ttk.Scrollbar(hscroll_container, orient="horizontal", command=self.report_tree.xview)
        self.report_tree.configure(xscrollcommand=hsb.set)
        hsb.pack(fill="x")

        # Button frame at the bottom with proper padding
        button_frame = ttk.Frame(tab)
        button_frame.pack(pady=5, fill="x", padx=10)

        # Pack each button with expand=True and fill="x" so they stretch
        ttk.Button(button_frame, text="Select All", command=self._select_all_report_transactions, style="TButton").pack(side=tk.LEFT, expand=True, fill="x", padx=2)
        ttk.Button(button_frame, text="EXPORT FILTERED TO CSV", command=lambda: self.export_to_csv(filtered=True), style="TButton").pack(side=tk.LEFT, expand=True, fill="x", padx=2)
        
        # Load completed transactions on initial tab display
        self._search_report_transactions()  # Initial load with date filters

    def _on_price_computation_toggled(self):
        """Handle price computation enabled/disabled toggle."""
        enabled = self.price_computation_enabled_var.get()
        self.config["price_computation_enabled"] = enabled
        self.save_config()
        
        # Update price displays immediately
        if hasattr(self, 'unit_price_var'):
            if enabled:
                self._update_price_display()
            else:
                self.unit_price_var.set("0.00")
                self.total_price_var.set("0")
        
        # Update transaction price display immediately
        if hasattr(self, 'transaction_total_price_var'):
            if enabled:
                self._update_transaction_price_display()
            else:
                self.transaction_total_price_var.set("0")
        
        # Update entry form total price field immediately
        if hasattr(self, 'total_price_entry_label') and hasattr(self, 'entry_vars') and 'total_price' in self.entry_vars:
            if enabled:
                # Trigger immediate price calculation with current weight
                self._update_transaction_price_display()
            else:
                self.entry_vars['total_price'].set("0")
        
        # Enable/disable one-way and two-way total price checkboxes
        if hasattr(self, 'one_way_total_price_cb'):
            state = "normal" if enabled else "disabled"
            self.one_way_total_price_cb.config(state=state)
            if not enabled:
                self.one_way_total_price_var.set(False)
                
        if hasattr(self, 'two_way_total_price_cb'):
            state = "normal" if enabled else "disabled"
            self.two_way_total_price_cb.config(state=state)
            if not enabled:
                self.two_way_total_price_var.set(False)

    def _save_price_settings(self):
        """Save price computation settings to config."""
        try:
            self.config["base_weight"] = self.base_weight_var.get()
            self.config["base_price"] = self.base_price_var.get()
            self.config["increment_weight"] = self.increment_weight_var.get()
            self.config["increment_price"] = self.increment_price_var.get()
            self.config["price_computation_enabled"] = self.price_computation_enabled_var.get()
            
            self.save_config()
            
            # Update price displays if enabled
            if self.price_computation_enabled_var.get():
                if hasattr(self, 'unit_price_var'):
                    self._update_price_display()
                if hasattr(self, 'transaction_total_price_var'):
                    self._update_transaction_price_display()
            
            self.msg_box.showinfo("Success", "Price computation settings saved successfully!")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save price settings: {e}")

    def build_ticket_output_tab(self, tab: ttk.Frame):
        """Builds the Ticket Format sub-tab (admin-only).
        Provides separate editors for One-way and Two-way ticket headers
        and page-size controls (per-type)."""
        # Only build once - don't destroy/rebuild on each call
        if tab.winfo_children():
            return

        container = ttk.Frame(tab, padding=10)
        container.pack(fill="both", expand=True)

        # One-way section
        one_frame = ttk.LabelFrame(container, text="One-way Ticket", padding=8)
        one_frame.pack(fill="x", pady=6)
        ttk.Label(one_frame, text="Header (Company Info):").grid(row=0, column=0, sticky="w")
        one_header = tk.Text(one_frame, height=6, wrap="word")
        one_header.grid(row=1, column=0, columnspan=4, sticky="we", pady=4)
        # Extract header from print template (remove transaction details)
        one_existing = self.one_way_print_template_var.get() if hasattr(self, 'one_way_print_template_var') else ""
        # Remove the transaction details section if present
        if "\n\n    TICKET NO : {ticket_no}" in one_existing:
            one_header_only = one_existing.split("\n\n    TICKET NO : {ticket_no}")[0]
        else:
            one_header_only = one_existing
        one_header.insert("1.0", one_header_only)
        self.ticket_header_one_way_text_widget = one_header

        ttk.Label(one_frame, text="Header Font Size:").grid(row=2, column=0, sticky="w", pady=4)
        one_font_size_var = tk.StringVar(value=str(self.config.get("ticket_header_font_size_one_way", "13")))
        self.ticket_header_font_size_one_var = one_font_size_var
        ttk.Combobox(one_frame, textvariable=one_font_size_var, values=["8", "9", "10", "11", "12", "13", "14", "16", "18", "20"], state="readonly", width=12).grid(row=2, column=1, sticky="w", padx=6)

        ttk.Label(one_frame, text="Page Size:").grid(row=2, column=2, sticky="w", pady=4)
        one_size_var = tk.StringVar(value=self.config.get("ticket_page_size_one_way", self.config.get("ticket_page_size", "A6")))
        self.ticket_page_size_one_var = one_size_var
        ttk.Combobox(one_frame, textvariable=one_size_var, values=["A4", "Letter", "A6", "Custom"], state="readonly", width=12).grid(row=2, column=3, sticky="w", padx=6)

        ttk.Label(one_frame, text="Width (mm):").grid(row=3, column=0, sticky="w")
        one_w = tk.StringVar(value=str(self.config.get("ticket_page_width_mm_one_way", self.config.get("ticket_page_width_mm", "105"))))
        self.ticket_page_width_one_var = one_w
        ttk.Entry(one_frame, textvariable=one_w, width=12).grid(row=3, column=1, sticky="w", padx=6)

        ttk.Label(one_frame, text="Height (mm):").grid(row=3, column=2, sticky="w")
        one_h = tk.StringVar(value=str(self.config.get("ticket_page_height_mm_one_way", self.config.get("ticket_page_height_mm", "148"))))
        self.ticket_page_height_one_var = one_h
        ttk.Entry(one_frame, textvariable=one_h, width=12).grid(row=3, column=3, sticky="w", padx=6)

        one_btns = ttk.Frame(one_frame)
        one_btns.grid(row=4, column=0, columnspan=4, pady=6, sticky="w")
        ttk.Button(one_btns, text="Preview One-way PDF", command=self._preview_ticket_pdf_one_way, style="TButton").pack(side=tk.LEFT, padx=4)

        # Two-way section
        two_frame = ttk.LabelFrame(container, text="Two-way Ticket", padding=8)
        two_frame.pack(fill="x", pady=6)
        ttk.Label(two_frame, text="Header (Company Info):").grid(row=0, column=0, sticky="w")
        two_header = tk.Text(two_frame, height=6, wrap="word")
        two_header.grid(row=1, column=0, columnspan=4, sticky="we", pady=4)
        # Extract header from print template (remove transaction details)
        two_existing = self.two_way_print_template_var.get() if hasattr(self, 'two_way_print_template_var') else ""
        # Remove the transaction details section if present
        if "\n\n    TICKET NO : {ticket_no}" in two_existing:
            two_header_only = two_existing.split("\n\n    TICKET NO : {ticket_no}")[0]
        else:
            two_header_only = two_existing
        two_header.insert("1.0", two_header_only)
        self.ticket_header_two_way_text_widget = two_header

        ttk.Label(two_frame, text="Header Font Size:").grid(row=2, column=0, sticky="w", pady=4)
        two_font_size_var = tk.StringVar(value=str(self.config.get("ticket_header_font_size_two_way", "13")))
        self.ticket_header_font_size_two_var = two_font_size_var
        ttk.Combobox(two_frame, textvariable=two_font_size_var, values=["8", "9", "10", "11", "12", "13", "14", "16", "18", "20"], state="readonly", width=12).grid(row=2, column=1, sticky="w", padx=6)

        ttk.Label(two_frame, text="Page Size:").grid(row=2, column=2, sticky="w", pady=4)
        two_size_var = tk.StringVar(value=self.config.get("ticket_page_size_two_way", self.config.get("ticket_page_size", "A6")))
        self.ticket_page_size_two_var = two_size_var
        ttk.Combobox(two_frame, textvariable=two_size_var, values=["A4", "Letter", "A6", "Custom"], state="readonly", width=12).grid(row=2, column=3, sticky="w", padx=6)

        ttk.Label(two_frame, text="Width (mm):").grid(row=3, column=0, sticky="w")
        two_w = tk.StringVar(value=str(self.config.get("ticket_page_width_mm_two_way", self.config.get("ticket_page_width_mm", "105"))))
        self.ticket_page_width_two_var = two_w
        ttk.Entry(two_frame, textvariable=two_w, width=12).grid(row=3, column=1, sticky="w", padx=6)

        ttk.Label(two_frame, text="Height (mm):").grid(row=3, column=2, sticky="w")
        two_h = tk.StringVar(value=str(self.config.get("ticket_page_height_mm_two_way", self.config.get("ticket_page_height_mm", "148"))))
        self.ticket_page_height_two_var = two_h
        ttk.Entry(two_frame, textvariable=two_h, width=12).grid(row=3, column=3, sticky="w", padx=6)

        two_btns = ttk.Frame(two_frame)
        two_btns.grid(row=4, column=0, columnspan=4, pady=6, sticky="w")
        ttk.Button(two_btns, text="Preview Two-way PDF", command=self._preview_ticket_pdf_two_way, style="TButton").pack(side=tk.LEFT, padx=4)

        # BOTTOM BUTTONS: Save for entire tab
        bottom_btns = ttk.Frame(container)
        bottom_btns.pack(fill="x", pady=(10, 0))
        ttk.Button(bottom_btns, text="Save Settings", command=self._save_all_ticket_settings).pack(side=tk.LEFT, fill="x", expand=True, padx=4)


    def _save_ticket_template(self):
        """Save the ticket header/template into configuration."""
        try:
            tpl = self.ticket_header_text_widget.get("1.0", "end").rstrip()
            self.config["ticket_header_template"] = tpl
            self.save_config()
            self.msg_box.showinfo("Saved", "Ticket header template saved.")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save template: {e}")

    def _save_ticket_page_size(self):
        """Save the selected page size and custom dimensions into configuration."""
        try:
            self.config["ticket_page_size"] = self.ticket_page_size_var.get()
            w = float(self.ticket_page_width_var.get())
            h = float(self.ticket_page_height_var.get())
            self.config["ticket_page_width_mm"] = w
            self.config["ticket_page_height_mm"] = h
            self.save_config()
            self.msg_box.showinfo("Saved", "Ticket page size saved.")
        except ValueError:
            self.msg_box.showerror("Invalid Value", "Width and Height must be numeric values (mm).")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save page size: {e}")

    def _preview_ticket_pdf(self):
        """Generate a simple preview PDF using the configured page size and header template."""
        if not PDF_PRINTING_ENABLED:
            self.msg_box.showerror("ReportLab Missing", "ReportLab library is not available. PDF preview is disabled.")
            return

        # Determine page size
        size_key = self.ticket_page_size_var.get() if hasattr(self, 'ticket_page_size_var') else self.config.get("ticket_page_size", "A6")
        try:
            if size_key == "Custom":
                w_mm = float(self.ticket_page_width_var.get())
                h_mm = float(self.ticket_page_height_var.get())
                mm_to_pt = 2.834645669
                page_size = (w_mm * mm_to_pt, h_mm * mm_to_pt)
            else:
                mapping = {"A4": A4, "Letter": letter, "A6": A6}
                page_size = mapping.get(size_key, A6)

            out_path = os.path.join(PROGRAM_DATA_DIR, "ticket_preview.pdf")
            c = canvas.Canvas(out_path, pagesize=page_size)
            header = self.ticket_header_text_widget.get("1.0", "end").rstrip()
            c.setFont("Helvetica", 10)
            margin = 20
            y = page_size[1] - margin
            for line in header.splitlines():
                c.drawString(margin, y, line)
                y -= 12
            c.showPage()
            c.save()
            self.msg_box.showinfo("Preview Saved", f"Preview saved to: {out_path}")
        except Exception as e:
            self.msg_box.showerror("Preview Error", f"Failed to generate preview PDF: {e}")
    
    def _save_ticket_format_one_way(self):
        """Save one-way header to database and page size to config."""
        try:
            header = self.ticket_header_one_way_text_widget.get("1.0", "end").rstrip()
            
            # Build complete template: header + transaction details
            one_way_details = (
                "\n\n"
                "    TICKET NO : {ticket_no}\n"
                "COMPANY       : {company}\n"
                "TRUCK PLATE   : {truck_plate}\n"
                "PRODUCT       : {product}\n"
                "\n\n"
                "GROSS WEIGHT  : {gross_weight} KG\n"
                "GROSS DATE    : {gross_date}\n"
                "GROSS TIME    : {gross_time}\n"
                "\n\n"
                "DATE PRINTED  : {date_printed}\n"
                "\n\n"
                "OPERATOR      : {logged_in_user}\n"
            )
            complete_template = header + one_way_details
            
            # Update the print template variable
            self.one_way_print_template_var.set(complete_template)
            # Save page size and font size to config
            self.config["ticket_page_size_one_way"] = self.ticket_page_size_one_var.get()
            self.config["ticket_page_width_mm_one_way"] = float(self.ticket_page_width_one_var.get())
            self.config["ticket_page_height_mm_one_way"] = float(self.ticket_page_height_one_var.get())
            self.config["ticket_header_font_size_one_way"] = int(self.ticket_header_font_size_one_var.get())
            self.save_config()
            # Save complete template to database permanently
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'ONE_WAY'", (complete_template,))
                conn.commit()
            self.msg_box.showinfo("Saved", "One-way ticket format saved successfully.")
        except ValueError:
            self.msg_box.showerror("Invalid Value", "Width and Height must be numeric values (mm).")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save one-way format: {e}")

    def _save_ticket_format_two_way(self):
        """Save two-way header to database and page size to config."""
        try:
            header = self.ticket_header_two_way_text_widget.get("1.0", "end").rstrip()
            
            # Build complete template: header + transaction details
            two_way_details = (
                "\n\n"
                "    TICKET NO : {ticket_no}\n"
                "COMPANY       : {company}\n"
                "TRUCK PLATE   : {truck_plate}\n"
                "PRODUCT       : {product}\n"
                "\n\n"
                "GROSS WEIGHT  : {gross_weight} KG\n"
                "GROSS DATE    : {gross_date}\n"
                "GROSS TIME    : {gross_time}\n"
                "\n\n"
                "TARE WEIGHT   : {tare_weight} KG\n"
                "TARE DATE     : {tare_date}\n"
                "TARE TIME     : {tare_time}\n"
                "\n\n"
                "NET WEIGHT    : {net_weight} KG\n"
                "\n\n"
                "DATE PRINTED  : {date_printed}\n"
                "\n\n"
                "OPERATOR      : {logged_in_user}\n"
            )
            complete_template = header + two_way_details
            
            # Update the print template variable
            self.two_way_print_template_var.set(complete_template)
            # Save page size and font size to config
            self.config["ticket_page_size_two_way"] = self.ticket_page_size_two_var.get()
            self.config["ticket_page_width_mm_two_way"] = float(self.ticket_page_width_two_var.get())
            self.config["ticket_page_height_mm_two_way"] = float(self.ticket_page_height_two_var.get())
            self.config["ticket_header_font_size_two_way"] = int(self.ticket_header_font_size_two_var.get())
            self.save_config()
            # Save complete template to database permanently
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'TWO_WAY'", (complete_template,))
                conn.commit()
            self.msg_box.showinfo("Saved", "Two-way ticket format saved successfully.")
        except ValueError:
            self.msg_box.showerror("Invalid Value", "Width and Height must be numeric values (mm).")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save two-way format: {e}")

    def _save_all_ticket_settings(self):
        """Save all ticket and print settings from the Print & Ticket Settings tab."""
        try:
            # Save one-way ticket format
            one_header = self.ticket_header_one_way_text_widget.get("1.0", "end").rstrip()
            
            # Build one-way details with optional fields
            one_way_details = (
                "\n\n"
                "    TICKET NO : {ticket_no}\n"
                "COMPANY       : {company}\n"
                "TRUCK PLATE   : {truck_plate}\n"
                "PRODUCT       : {product}\n"
            )
            
            # Add optional fields for one-way
            one_way_optional = self.config.get("one_way_optional_fields", [])
            if "Designation" in one_way_optional:
                one_way_details += "DESIGNATION   : {designation}\n"
            if "Sender" in one_way_optional:
                one_way_details += "SENDER        : {sender}\n"
            if "Origin" in one_way_optional:
                one_way_details += "ORIGIN        : {origin}\n"
            if "Destination" in one_way_optional:
                one_way_details += "DESTINATION   : {destination}\n"
            if "Driver" in one_way_optional:
                one_way_details += "DRIVER        : {driver}\n"
            if "Total Price" in one_way_optional:
                one_way_details += "TOTAL PRICE   : {total_price}\n"
            
            one_way_details += (
                "\n\n"
                "GROSS WEIGHT  : {gross_weight} KG\n"
                "GROSS DATE    : {gross_date}\n"
                "GROSS TIME    : {gross_time}\n"
                "\n\n"
                "DATE PRINTED  : {date_printed}\n"
                "\n\n"
                "OPERATOR      : {logged_in_user}\n"
            )
            one_complete_template = one_header + one_way_details
            self.one_way_print_template_var.set(one_complete_template)
            
            # Save two-way ticket format
            two_header = self.ticket_header_two_way_text_widget.get("1.0", "end").rstrip()
            
            # Build two-way details with optional fields
            two_way_details = (
                "\n\n"
                "    TICKET NO : {ticket_no}\n"
                "COMPANY       : {company}\n"
                "TRUCK PLATE   : {truck_plate}\n"
                "PRODUCT       : {product}\n"
            )
            
            # Add optional fields for two-way
            two_way_optional = self.config.get("two_way_optional_fields", [])
            if "Designation" in two_way_optional:
                two_way_details += "DESIGNATION   : {designation}\n"
            if "Sender" in two_way_optional:
                two_way_details += "SENDER        : {sender}\n"
            if "Origin" in two_way_optional:
                two_way_details += "ORIGIN        : {origin}\n"
            if "Destination" in two_way_optional:
                two_way_details += "DESTINATION   : {destination}\n"
            if "Driver" in two_way_optional:
                two_way_details += "DRIVER        : {driver}\n"
            if "Total Price" in two_way_optional:
                two_way_details += "TOTAL PRICE   : {total_price}\n"
            
            two_way_details += (
                "\n\n"
                "GROSS WEIGHT  : {gross_weight} KG\n"
                "GROSS DATE    : {gross_date}\n"
                "GROSS TIME    : {gross_time}\n"
                "\n\n"
                "TARE WEIGHT   : {tare_weight} KG\n"
                "TARE DATE     : {tare_date}\n"
                "TARE TIME     : {tare_time}\n"
                "\n\n"
                "NET WEIGHT    : {net_weight} KG\n"
                "\n\n"
                "DATE PRINTED  : {date_printed}\n"
                "\n\n"
                "OPERATOR      : {logged_in_user}\n"
            )
            two_complete_template = two_header + two_way_details
            self.two_way_print_template_var.set(two_complete_template)
            
            # Save all one-way settings to config
            self.config["ticket_page_size_one_way"] = self.ticket_page_size_one_var.get()
            self.config["ticket_page_width_mm_one_way"] = self.ticket_page_width_one_var.get()
            self.config["ticket_page_height_mm_one_way"] = self.ticket_page_height_one_var.get()
            self.config["ticket_header_font_size_one_way"] = self.ticket_header_font_size_one_var.get()
            
            # Save all two-way settings to config
            self.config["ticket_page_size_two_way"] = self.ticket_page_size_two_var.get()
            self.config["ticket_page_width_mm_two_way"] = self.ticket_page_width_two_var.get()
            self.config["ticket_page_height_mm_two_way"] = self.ticket_page_height_two_var.get()
            self.config["ticket_header_font_size_two_way"] = self.ticket_header_font_size_two_var.get()
            
            # Save printer settings from right panel
            self.config["selected_printer"] = self.selected_printer_var.get()
            self.config["print_copies"] = int(self.print_copies_var.get())
            self.config["pdf_page_size"] = self.pdf_page_size_var.get()
            self.config["pdf_orientation"] = self.pdf_orientation_var.get()
            self.config["starting_ticket_number"] = self.starting_ticket_number_var.get()
            
            # Save config to file
            self.save_config()
            
            # Save templates to database
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'ONE_WAY'", (one_complete_template,))
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'TWO_WAY'", (two_complete_template,))
                conn.commit()
            
            self.msg_box.showinfo("Settings Saved", "All Print & Ticket settings saved successfully and will persist on restart.")
        except ValueError:
            self.msg_box.showerror("Invalid Value", "Width, Height, and Font Size must be numeric values.")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save settings: {e}")

    def _reset_all_ticket_settings_to_default(self):
        """Reset all ticket and print settings to defaults."""
        if not self.msg_box.askyesno("Confirm Reset", "Reset all Print & Ticket settings to default values?", lambda res: None):
            return
        
        try:
            # Reset one-way settings
            default_one_way = (
                "Advantechnique\n"
                "advantechnique@gmail.com\n"
                "\n"
                "\n"
            )
            self.ticket_header_one_way_text_widget.config(state="normal")
            self.ticket_header_one_way_text_widget.delete("1.0", "end")
            self.ticket_header_one_way_text_widget.insert("1.0", default_one_way)
            self.ticket_header_one_way_text_widget.config(state="normal")
            
            self.ticket_header_font_size_one_var.set("13")
            self.ticket_page_size_one_var.set("Custom")
            self.ticket_page_width_one_var.set("105")
            self.ticket_page_height_one_var.set("148")
            
            # Reset two-way settings
            default_two_way = (
                "Advantechnique\n"
                "advantechnique@gmail.com\n"
                "\n"
                "\n"
            )
            self.ticket_header_two_way_text_widget.config(state="normal")
            self.ticket_header_two_way_text_widget.delete("1.0", "end")
            self.ticket_header_two_way_text_widget.insert("1.0", default_two_way)
            self.ticket_header_two_way_text_widget.config(state="normal")
            
            self.ticket_header_font_size_two_var.set("13")
            self.ticket_page_size_two_var.set("Custom")
            self.ticket_page_width_two_var.set("105")
            self.ticket_page_height_two_var.set("148")
            
            # Reset optional fields
            self.one_way_designation_var.set(False)
            self.one_way_sender_var.set(False)
            self.one_way_origin_var.set(False)
            self.one_way_destination_var.set(False)
            self.one_way_driver_var.set(False)
            self.one_way_total_price_var.set(False)
            
            self.two_way_designation_var.set(False)
            self.two_way_sender_var.set(False)
            self.two_way_origin_var.set(False)
            self.two_way_destination_var.set(False)
            self.two_way_driver_var.set(False)
            self.two_way_total_price_var.set(False)
            
            # Reset printer settings
            self.selected_printer_var.set("")
            self.print_copies_var.set(1)
            self.pdf_page_size_var.set("Letter")
            self.pdf_orientation_var.set("Portrait")
            self.starting_ticket_number_var.set(1000)
            
            # Now save everything
            self._save_all_ticket_settings()
            self.msg_box.showinfo("Reset Complete", "All settings reset to defaults and saved.")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to reset settings: {e}")

    def _update_one_way_optional_fields(self):
        """Update one-way optional fields when checkboxes are toggled."""
        selected_fields = []
        if self.one_way_designation_var.get():
            selected_fields.append("Designation")
        if self.one_way_sender_var.get():
            selected_fields.append("Sender")
        if self.one_way_origin_var.get():
            selected_fields.append("Origin")
        if self.one_way_destination_var.get():
            selected_fields.append("Destination")
        if self.one_way_driver_var.get():
            selected_fields.append("Driver")
        if self.one_way_total_price_var.get():
            selected_fields.append("Total Price")
            
        # Enable price computation if total price is selected
        if self.one_way_total_price_var.get() and not self.price_computation_enabled_var.get():
            self.price_computation_enabled_var.set(True)
            self._on_price_computation_toggled()
        
        self.config["one_way_optional_fields"] = selected_fields
        self.save_config()
        # Update the template with new optional fields
        self._update_ticket_templates()

    def _update_two_way_optional_fields(self):
        """Update two-way optional fields when checkboxes are toggled."""
        selected_fields = []
        if self.two_way_designation_var.get():
            selected_fields.append("Designation")
        if self.two_way_sender_var.get():
            selected_fields.append("Sender")
        if self.two_way_origin_var.get():
            selected_fields.append("Origin")
        if self.two_way_destination_var.get():
            selected_fields.append("Destination")
        if self.two_way_driver_var.get():
            selected_fields.append("Driver")
        if self.two_way_total_price_var.get():
            selected_fields.append("Total Price")
            
        # Enable price computation if total price is selected
        if self.two_way_total_price_var.get() and not self.price_computation_enabled_var.get():
            self.price_computation_enabled_var.set(True)
            self._on_price_computation_toggled()
        
        self.config["two_way_optional_fields"] = selected_fields
        self.save_config()
        # Update the template with new optional fields
        self._update_ticket_templates()
    
    def _update_ticket_templates(self):
        """Rebuild the ticket templates based on current optional field selections."""
        try:
            # Rebuild one-way template
            one_header = self.ticket_header_one_way_text_widget.get("1.0", "end").rstrip()
            
            # Build one-way details with optional fields
            one_way_details = (
                "\n\n"
                "    TICKET NO : {ticket_no}\n"
                "COMPANY       : {company}\n"
                "TRUCK PLATE   : {truck_plate}\n"
                "PRODUCT       : {product}\n"
            )
            
            # Add optional fields for one-way
            one_way_optional = self.config.get("one_way_optional_fields", [])
            if "Designation" in one_way_optional:
                one_way_details += "DESIGNATION   : {designation}\n"
            if "Sender" in one_way_optional:
                one_way_details += "SENDER        : {sender}\n"
            if "Origin" in one_way_optional:
                one_way_details += "ORIGIN        : {origin}\n"
            if "Destination" in one_way_optional:
                one_way_details += "DESTINATION   : {destination}\n"
            if "Driver" in one_way_optional:
                one_way_details += "DRIVER        : {driver}\n"
            if "Total Price" in one_way_optional:
                one_way_details += "TOTAL PRICE   : {total_price}\n"
            
            one_way_details += (
                "\n\n"
                "GROSS WEIGHT  : {gross_weight} KG\n"
                "GROSS DATE    : {gross_date}\n"
                "GROSS TIME    : {gross_time}\n"
                "\n\n"
                "DATE PRINTED  : {date_printed}\n"
                "\n\n"
                "OPERATOR      : {logged_in_user}\n"
            )
            one_complete_template = one_header + one_way_details
            self.one_way_print_template_var.set(one_complete_template)
            
            # Rebuild two-way template
            two_header = self.ticket_header_two_way_text_widget.get("1.0", "end").rstrip()
            
            # Build two-way details with optional fields
            two_way_details = (
                "\n\n"
                "    TICKET NO : {ticket_no}\n"
                "COMPANY       : {company}\n"
                "TRUCK PLATE   : {truck_plate}\n"
                "PRODUCT       : {product}\n"
            )
            
            # Add optional fields for two-way
            two_way_optional = self.config.get("two_way_optional_fields", [])
            if "Designation" in two_way_optional:
                two_way_details += "DESIGNATION   : {designation}\n"
            if "Sender" in two_way_optional:
                two_way_details += "SENDER        : {sender}\n"
            if "Origin" in two_way_optional:
                two_way_details += "ORIGIN        : {origin}\n"
            if "Destination" in two_way_optional:
                two_way_details += "DESTINATION   : {destination}\n"
            if "Driver" in two_way_optional:
                two_way_details += "DRIVER        : {driver}\n"
            if "Total Price" in two_way_optional:
                two_way_details += "TOTAL PRICE   : {total_price}\n"
            
            two_way_details += (
                "\n\n"
                "GROSS WEIGHT  : {gross_weight} KG\n"
                "GROSS DATE    : {gross_date}\n"
                "GROSS TIME    : {gross_time}\n"
                "\n\n"
                "TARE WEIGHT   : {tare_weight} KG\n"
                "TARE DATE     : {tare_date}\n"
                "TARE TIME     : {tare_time}\n"
                "\n\n"
                "NET WEIGHT    : {net_weight} KG\n"
                "\n\n"
                "DATE PRINTED  : {date_printed}\n"
                "\n\n"
                "OPERATOR      : {logged_in_user}\n"
            )
            two_complete_template = two_header + two_way_details
            self.two_way_print_template_var.set(two_complete_template)
            
            # Save to database
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'ONE_WAY'", (one_complete_template,))
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'TWO_WAY'", (two_complete_template,))
                conn.commit()
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to update templates: {e}")

    def _get_page_dimensions_for_preview(self, weight_type: str) -> tuple:
        """Get page dimensions in pixels for print preview based on configured size."""
        if weight_type == "ONE WAY WEIGHING":
            size_key = self.ticket_page_size_one_var.get()
            width_mm_var = self.ticket_page_width_one_var
            height_mm_var = self.ticket_page_height_one_var
        else:
            size_key = self.ticket_page_size_two_var.get()
            width_mm_var = self.ticket_page_width_two_var
            height_mm_var = self.ticket_page_height_two_var
        
        # Standard page sizes in mm (width x height)
        page_sizes_mm = {
            "A4": (210, 297),
            "Letter": (215.9, 279.4),
            "A6": (105, 148)
        }
        
        if size_key == "Custom":
            try:
                width_mm = float(width_mm_var.get())
                height_mm = float(height_mm_var.get())
            except (ValueError, tk.TclError):
                width_mm, height_mm = 105, 148  # Fallback to A6
        else:
            width_mm, height_mm = page_sizes_mm.get(size_key, (105, 148))
        
        # Convert mm to pixels (96 DPI standard)
        mm_to_px = 3.779527559  # 96 DPI
        width_px = int(width_mm * mm_to_px)
        height_px = int(height_mm * mm_to_px)
        
        return width_px, height_px

    def _format_wrapped_ticket_line(self, c, line, margin, max_width, detail_font_size):
        """
        Format a ticket line for PDF, ensuring proper alignment when text wraps.
        Returns list of (text, x_offset, y_offset) tuples for drawing.
        Wraps on word boundaries for clean output.
        """
        lines_to_draw = []
        
        # Check if line contains a colon (label: value format)
        if ':' not in line:
            lines_to_draw.append((line, margin, 0))
            return lines_to_draw
        
        # Split on colon
        colon_pos = line.find(':')
        label_part = line[:colon_pos + 1]  # Include colon
        value_part = line[colon_pos + 1:].strip()
        
        # Calculate the width of label + colon + space
        label_width = c.stringWidth(label_part + " ", "Courier", detail_font_size)
        value_start_x = margin + label_width
        
        # Check if entire line fits
        full_line_width = c.stringWidth(line, "Courier", detail_font_size)
        if full_line_width <= max_width:
            lines_to_draw.append((line, margin, 0))
            return lines_to_draw
        
        # Line is too long, need to wrap
        # First, draw label with colon
        lines_to_draw.append((label_part, margin, 0))
        
        # Now wrap the value part at word boundaries
        remaining_value = value_part
        y_offset = 0
        available_width = max_width - value_start_x
        
        while remaining_value:
            # Find how many words fit in available width
            words = remaining_value.split()
            if not words:
                break
            
            current_line = ""
            words_used = 0
            
            for i, word in enumerate(words):
                test_line = (current_line + (" " if current_line else "") + word).strip()
                test_width = c.stringWidth(test_line, "Courier", detail_font_size)
                
                if test_width <= available_width:
                    current_line = test_line
                    words_used = i + 1
                else:
                    break
            
            # If no words fit, force at least one word
            if words_used == 0 and words:
                current_line = words[0]
                words_used = 1
            
            if current_line:
                y_offset += (detail_font_size + 2)
                lines_to_draw.append((current_line, value_start_x, y_offset))
                remaining_value = ' '.join(words[words_used:])
            else:
                break
        
        return lines_to_draw

    def _show_print_preview_window(self, content, title):
        """Display a print preview window with the ticket content and print functionality."""
        preview_window = tk.Toplevel(self.root)
        preview_window.title(title)
        preview_window.geometry("800x600")
        
        # Create a frame for buttons
        button_frame = ttk.Frame(preview_window)
        button_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        # Print button
        def on_print():
            try:
                if sys.platform == "win32" and PRINTING_ENABLED:
                    try:
                        import win32print
                        import win32api
                        printer_name = win32print.GetDefaultPrinter()
                        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False)
                        temp_file.write(content)
                        temp_file.close()
                        win32api.ShellExecute(0, "print", temp_file.name, None, ".", 0)
                        self.msg_box.showinfo("Print", "Document sent to printer.")
                    except Exception as e:
                        self.msg_box.showerror("Print Error", f"Failed to print: {e}")
                else:
                    self.msg_box.showinfo("Print", "Print functionality is not available on this system.")
            except Exception as e:
                self.msg_box.showerror("Print Error", f"Error: {e}")
        
        ttk.Button(button_frame, text="Print", command=on_print).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Close", command=preview_window.destroy).pack(side=tk.LEFT, padx=5)
        
        text_frame = ttk.Frame(preview_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(text_frame, wrap=tk.NONE, yscrollcommand=scrollbar.set, font=("Courier", 9))
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        text_widget.insert(tk.END, content)
        text_widget.config(state=tk.DISABLED)

    def calculate_price(self, weight_kg: float) -> tuple[float, float]:
        """
        Calculate price based on weight using configured parameters.
        
        Args:
            weight_kg (float): Weight in kilograms
            
        Returns:
            tuple: (unit_price, total_price) where unit_price is price per kg and total_price is total cost
        """
        try:
            if not self.config.get("price_computation_enabled", True):
                return 0.0, 0.0
                
            base_weight = float(self.config.get("base_weight", DEFAULT_BASE_WEIGHT))
            base_price = float(self.config.get("base_price", DEFAULT_BASE_PRICE))
            increment_weight = float(self.config.get("increment_weight", DEFAULT_INCREMENT_WEIGHT))
            increment_price = float(self.config.get("increment_price", DEFAULT_INCREMENT_PRICE))
            
            if weight_kg <= 0:
                return 0.0, 0.0
                
            if weight_kg <= base_weight:
                total_price = base_price
                unit_price = total_price / weight_kg if weight_kg > 0 else 0.0
            else:
                # Calculate additional weight beyond base
                additional_weight = weight_kg - base_weight
                additional_units = additional_weight / increment_weight
                additional_price = additional_units * increment_price
                
                total_price = base_price + additional_price
                unit_price = total_price / weight_kg if weight_kg > 0 else 0.0
                
            return round(unit_price, 2), round(total_price, 0)
            
        except Exception as e:
            logging.error(f"Error calculating price: {e}")
            return 0.0, 0.0

    def _preview_ticket_pdf_one_way(self):
        """Show print preview for one-way ticket using its configured template and size."""
        try:
            # Create sample data for preview
            sample_data = {
                "company": "SAMPLE COMPANY",
                "ticket_no": "12345",
                "truck_plate": "ABC-123",
                "product": "GRAIN",
                "gross_weight": "25000.00",
                "gross_date": datetime.now().strftime("%m/%d/%Y"),
                "gross_time": datetime.now().strftime("%I:%M:%S %p"),
                "date_printed": datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
                "tare_weight": "0.00",
                "tare_date": "N/A",
                "tare_time": "N/A",
                "net_weight": "25000.00",
                "weight_type": "ONE WAY WEIGHING",
                "logged_in_user": self.logged_in_user if self.logged_in_user else "admin"
            }
            
            # Get the template
            template_str = self.one_way_print_template_var.get()
            
            # Format the template with sample data
            class SafeDict(dict):
                def __missing__(self, key):
                    return ''
            formatted_text = template_str.format_map(SafeDict(sample_data))
            
            # Show print preview window
            self._show_print_preview_window(formatted_text, "One-Way Ticket Preview")
        except Exception as e:
            self.msg_box.showerror("Preview Error", f"Failed to show preview: {e}")

    def _preview_ticket_pdf_two_way(self):
        """Show print preview for two-way ticket using its configured template and size."""
        try:
            # Create sample data for preview
            sample_data = {
                "company": "SAMPLE COMPANY",
                "ticket_no": "12345",
                "truck_plate": "ABC-123",
                "product": "GRAIN",
                "gross_weight": "30000.00",
                "gross_date": datetime.now().strftime("%m/%d/%Y"),
                "gross_time": datetime.now().strftime("%I:%M:%S %p"),
                "tare_weight": "5000.00",
                "tare_date": datetime.now().strftime("%m/%d/%Y"),
                "tare_time": datetime.now().strftime("%I:%M:%S %p"),
                "net_weight": "25000.00",
                "date_printed": datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
                "weight_type": "TWO WAY WEIGHING",
                "logged_in_user": self.logged_in_user if self.logged_in_user else "admin"
            }
            
            # Get the template
            template_str = self.two_way_print_template_var.get()
            
            # Format the template with sample data
            class SafeDict(dict):
                def __missing__(self, key):
                    return ''
            formatted_text = template_str.format_map(SafeDict(sample_data))
            
            # Show print preview window
            self._show_print_preview_window(formatted_text, "Two-Way Ticket Preview")
        except Exception as e:
            self.msg_box.showerror("Preview Error", f"Failed to show preview: {e}")
    
    def _search_report_transactions(self, skip_date_filter=False):
        """
        Executes an improved search query with advanced filtering options.
        Shows all transactions with enhanced filtering capabilities.
        Supports:
        - Text search across multiple fields (company, truck_plate, driver, etc.)
        - Partial matching (e.g., "carlo" matches "carlo santos")
        - Case-insensitive search
        - Numeric search by ticket number
        - Weight range filtering (optional: can add with ">" or "<" prefix)
        - Date range filtering
        - Status filtering
        - Weight type filtering
        """
        # Show searching status
        if hasattr(self, 'report_search_status_var'):
            self.report_search_status_var.set("Searching...")
            self.report_search_status_label.config(foreground="blue")
        
        for i in self.report_tree.get_children():
            self.report_tree.delete(i)
        
        search_query = self.report_search_query_var.get().strip()
        from_date = self.report_from_date_var.get()
        to_date = self.report_to_date_var.get()
        
        # Skip date filter if requested
        if skip_date_filter:
            from_date = ""
            to_date = ""
            
        status_filter = self.report_status_var.get() if hasattr(self, 'report_status_var') else "All"
        weight_type_filter = self.report_weight_type_var.get() if hasattr(self, 'report_weight_type_var') else "All"

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                base_query = "SELECT id, ticket_no, company, truck_plate, product, driver, gross_weight, tare_weight, net_weight, gross_date, gross_time, tare_date, tare_time, status, weight_type, operator, operator2 FROM transactions"
                query_params = []
                where_clauses = []

                # Apply status filter
                if status_filter != "All":
                    where_clauses.append("status = ?")
                    query_params.append(status_filter)
                
                # Apply weight type filter
                if weight_type_filter != "All":
                    where_clauses.append("weight_type = ?")
                    query_params.append(weight_type_filter)

                if search_query:
                    # Enhanced search with better pattern matching
                    text_search_fields = ["company", "truck_plate", "product", "driver", "designation", "sender", "origin", "destination", "operator", "operator2"]
                    
                    # Support for partial matching with wildcards
                    text_search_pattern = f"%{search_query}%"
                    search_sub_clauses = [f"UPPER({field}) LIKE ?" for field in text_search_fields]
                    search_params = [text_search_pattern] * len(text_search_fields)
                    
                    # Also search by exact ticket number if numeric
                    if search_query.isdigit():
                        search_sub_clauses.append("ticket_no = ?")
                        search_params.append(int(search_query))
                    
                    # Also search by status if status matches (Pending, Completed, etc.)
                    if search_query.upper() in ["PENDING", "COMPLETED"]:
                        search_sub_clauses.append("UPPER(status) = ?")
                        search_params.append(search_query.upper())
                    
                    where_clauses.append("(" + " OR ".join(search_sub_clauses) + ")")
                    query_params.extend(search_params)

                if from_date and to_date:
                    try:
                        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
                        to_dt = datetime.strptime(to_date, "%Y-%m-%d")
                        # Convert to MM/DD/YYYY format to match database format
                        from_date_converted = from_dt.strftime("%m/%d/%Y")
                        to_date_converted = to_dt.strftime("%m/%d/%Y")
                        where_clauses.append("(gross_date >= ? AND gross_date <= ?)")
                        query_params.extend([from_date_converted, to_date_converted])
                    except ValueError:
                        self.msg_box.showwarning("Date Format Error", "Invalid date format in report filter. Date filter ignored.")

                if where_clauses:
                    query = base_query + " WHERE " + " AND ".join(where_clauses)
                else:
                    query = base_query
                
                query += " ORDER BY timestamp DESC"
                
                cursor.execute(query, query_params)
                results = cursor.fetchall()
                
                # Update result count
                if hasattr(self, 'report_result_count_var'):
                    if results:
                        self.report_result_count_var.set(f"Found {len(results)} record{'s' if len(results) != 1 else ''}")
                    else:
                        self.report_result_count_var.set("No records found")

                self.all_report_transactions = []
                for row in results:
                    self.all_report_transactions.append(row)
                    (
                        id, ticket_no, company, truck_plate, product, driver, gross_weight, tare_weight, net_weight,
                        gross_date, gross_time, tare_date, tare_time, status, weight_type, operator, operator2
                    ) = row

                    format_string = f"%.{self.decimal_places}f"
                    formatted_gross_weight = format_string % gross_weight if gross_weight is not None else "N/A"
                    formatted_tare_weight = format_string % tare_weight if tare_weight is not None else "N/A"
                    formatted_net_weight = format_string % net_weight if net_weight is not None else "N/A"
                    consolidated_gross_datetime = f"{gross_date} {gross_time}" if gross_date and gross_time else "N/A"
                    consolidated_tare_datetime = f"{tare_date} {tare_time}" if tare_date and tare_time else "N/A"

                    self.report_tree.insert("", "end", values=(
                        id, ticket_no, company or "N/A", truck_plate or "N/A", 
                        product or "N/A", driver or "N/A",
                        formatted_gross_weight, formatted_tare_weight, formatted_net_weight,
                        consolidated_gross_datetime, consolidated_tare_datetime,
                        status or "N/A"
                    ))
        except sqlite3.Error as e:
            logging.error(f"Error loading report transactions: {e}")
            self.msg_box.showerror("Database Error", f"Failed to load transactions: {str(e)}")
        finally:
            # Update search status
            if hasattr(self, 'report_search_status_var'):
                self.report_search_status_var.set("Ready")
                self.report_search_status_label.config(foreground="gray")
            
    def clear_report_search(self):
        """Clears the search fields for the Reports tab and reloads all report transactions."""
        self.report_search_query_var.set("")
        self.report_from_date_var.set("")
        self.report_to_date_var.set("")
        if hasattr(self, 'report_status_var'):
            self.report_status_var.set("All")
        if hasattr(self, 'report_weight_type_var'):
            self.report_weight_type_var.set("All")
        self._search_report_transactions() # Call the search method to refresh with no filters        

    def _select_all_report_transactions(self):
        """
        Selects all items in the report treeview.
        """
        self.report_tree.selection_set(self.report_tree.get_children())
        
    def export_to_csv(self, filtered: bool = False):
        """
        Exports selected transactions to an Excel file with auto-fitted columns and rows.
        Only exports the transactions that are currently selected in the reports treeview.
        """
        # Try to import openpyxl for Excel export
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            excel_available = True
        except ImportError:
            excel_available = False
        
        # Choose file format based on availability
        if excel_available:
            filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
            default_ext = ".xlsx"
            title = "Export Selected Completed Records to Excel"
        else:
            filetypes = [("CSV files", "*.csv"), ("All files", "*.*")]
            default_ext = ".csv"
            title = "Export Selected Completed Records to CSV"
        
        filename = filedialog.asksaveasfilename(
            defaultextension=default_ext,
            filetypes=filetypes,
            title=title,
            parent=self.root
        )
        if not filename:
            return

        # Get selected items from the treeview
        selected_items = self.report_tree.selection()
        if not selected_items:
            self.msg_box.showwarning("No Selection", "Please select at least one transaction to export.")
            return

        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()

                cursor.execute("PRAGMA table_info(transactions)")
                all_col_names = [info[1] for info in cursor.fetchall()]
                # Exclude timestamp, unit_price, total_price, and photo_path from export
                cols_to_export = [name for name in all_col_names if name not in ['timestamp', 'unit_price', 'total_price', 'photo_path']]
                
                csv_headers = [name.replace('_', ' ').upper() for name in cols_to_export]

                # Get transaction IDs from selected treeview items
                selected_ids = []
                for item in selected_items:
                    values = self.report_tree.item(item, 'values')
                    if values:
                        selected_ids.append(values[0])  # ID is the first column

                if not selected_ids:
                    self.msg_box.showwarning("No Data", "No valid transactions found in selection.")
                    return

                # Build query to get selected transactions
                placeholders = ','.join(['?' for _ in selected_ids])
                query = f"SELECT {', '.join(cols_to_export)} FROM transactions WHERE id IN ({placeholders}) ORDER BY timestamp DESC"
                
                cursor.execute(query, selected_ids)

                # Get column indices for formatting
                gross_idx = cols_to_export.index('gross_weight') if 'gross_weight' in cols_to_export else -1
                tare_idx = cols_to_export.index('tare_weight') if 'tare_weight' in cols_to_export else -1
                net_idx = cols_to_export.index('net_weight') if 'net_weight' in cols_to_export else -1
                gross_total_price_idx = cols_to_export.index('gross_total_price') if 'gross_total_price' in cols_to_export else -1
                tare_total_price_idx = cols_to_export.index('tare_total_price') if 'tare_total_price' in cols_to_export else -1
                operator_idx = cols_to_export.index('operator') if 'operator' in cols_to_export else -1
                operator2_idx = cols_to_export.index('operator2') if 'operator2' in cols_to_export else -1
                weight_type_idx = cols_to_export.index('weight_type') if 'weight_type' in cols_to_export else -1

                # Collect all data rows
                data_rows = []
                for row in cursor.fetchall():
                    formatted_row = list(row)
                    format_string = f"%.{self.decimal_places}f"
                    
                    if gross_idx != -1 and formatted_row[gross_idx] is not None:
                        formatted_row[gross_idx] = format_string % formatted_row[gross_idx]
                    if tare_idx != -1 and formatted_row[tare_idx] is not None:
                        formatted_row[tare_idx] = format_string % formatted_row[tare_idx]
                    if net_idx != -1 and formatted_row[net_idx] is not None:
                        formatted_row[net_idx] = format_string % formatted_row[net_idx]
                    
                    # Handle gross_total_price
                    if gross_total_price_idx != -1:
                        existing_gross_total = formatted_row[gross_total_price_idx]
                        if existing_gross_total is not None and str(existing_gross_total).strip() not in ['0', '0.00', '']:
                            try:
                                gross_total_val = float(existing_gross_total)
                                if gross_total_val > 0:
                                    formatted_row[gross_total_price_idx] = "%.2f" % gross_total_val
                                else:
                                    formatted_row[gross_total_price_idx] = ""
                            except (ValueError, TypeError):
                                formatted_row[gross_total_price_idx] = ""
                        else:
                            formatted_row[gross_total_price_idx] = ""
                    
                    # Handle tare_total_price
                    if tare_total_price_idx != -1:
                        existing_tare_total = formatted_row[tare_total_price_idx]
                        if existing_tare_total is not None and str(existing_tare_total).strip() not in ['0', '0.00', '']:
                            try:
                                tare_total_val = float(existing_tare_total)
                                if tare_total_val > 0:
                                    formatted_row[tare_total_price_idx] = "%.2f" % tare_total_val
                                else:
                                    formatted_row[tare_total_price_idx] = ""
                            except (ValueError, TypeError):
                                formatted_row[tare_total_price_idx] = ""
                        else:
                            formatted_row[tare_total_price_idx] = ""
                    
                    # Handle operator field based on transaction type
                    if operator_idx != -1:
                        operator1 = formatted_row[operator_idx] or ""
                        if operator2_idx != -1:
                            operator2 = formatted_row[operator2_idx] or ""
                            # For two-way transactions, show operator1/operator2
                            if operator2 and operator1 != operator2:
                                formatted_row[operator_idx] = f"{operator1}/{operator2}"
                            elif operator2:
                                formatted_row[operator_idx] = f"{operator1}/{operator2}"
                    
                    data_rows.append(formatted_row)

                # Export based on file format
                if excel_available and filename.endswith('.xlsx'):
                    # Create Excel file with formatting
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Transactions"
                    
                    # Add headers with formatting
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    for col_idx, header in enumerate(csv_headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Add data with formatting
                    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    for row_idx, row_data in enumerate(data_rows, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
                            cell.alignment = data_alignment
                    
                    # Auto-fit column widths
                    for col_idx, column in enumerate(ws.columns, 1):
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(col_idx)
                        
                        # Check header length
                        header_cell = ws.cell(row=1, column=col_idx)
                        if header_cell.value:
                            max_length = max(max_length, len(str(header_cell.value)))
                        
                        # Check data cells
                        for cell in column:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                # For multi-line content, consider the longest line
                                if '\n' in str(cell.value):
                                    cell_length = max(len(line) for line in str(cell.value).split('\n'))
                                max_length = max(max_length, cell_length)
                        
                        # Set column width with padding
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Auto-fit row heights
                    for row in ws.iter_rows():
                        max_height = 0
                        for cell in row:
                            if cell.value:
                                # Calculate required height based on line breaks
                                lines = str(cell.value).split('\n')
                                max_height = max(max_height, len(lines))
                        
                        # Set row height (minimum 15, maximum 100)
                        if max_height > 1:
                            row_height = min(max_height * 15, 100)
                            ws.row_dimensions[row[0].row].height = row_height
                    
                    # Freeze header row
                    ws.freeze_panes = "A2"
                    
                    # Add borders to all cells
                    thin_border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.border = thin_border
                    
                    # Save Excel file
                    wb.save(filename)
                    
                else:
                    # Fallback to CSV export
                    import csv
                    with open(filename, "w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerow(csv_headers)
                        writer.writerows(data_rows)

            self.msg_box.showinfo("Exported", f"Transaction data successfully exported to:\n{filename}")
        except Exception as e:
            self.msg_box.showerror("Export Error", f"Failed to export data: {e}")
# Part 9: Printing and PDF Generation

    def _test_print(self, to_file: bool = False, file_type: str = "txt"):
        """
        Generates and prints a test ticket with sample data, useful for verifying
        the printer connection and the configured template.
        """
        # Get the current template based on the editor's mode
        current_template_mode = self.current_template_editor_mode_var.get()
        template_str = self.one_way_print_template_var.get() if current_template_mode == "ONE_WAY" else self.two_way_print_template_var.get()
        
        # Sample data for a completed two-way weighing
        date_printed_value = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
        sample_data = {
            "company": "SAMPLE CO.",
            "ticket_no": 12345,
            "truck_plate": "ABC 123",
            "product": "GRAIN",
            "designation": "BULK",
            "sender": "FARMER JOE",
            "origin": "FARM #1",
            "destination": "WAREHOUSE A",
            "driver": "JOHN DOE",
            "gross_weight": "25000.00",
            "gross_date": "2023-10-27",
            "gross_time": "10:30 AM",
            "tare_weight": "10000.00",
            "tare_date": "2023-10-27",
            "tare_time": "10:45 AM",
            "net_weight": "15000.00",
            "weight_type": "GROSS-TARE",
            "status": "Completed",
            "date_printed": datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
            "barcode": "12345",
            "logged_in_user": self.logged_in_user if self.logged_in_user else "admin"
        }
        
        line_spacing = self.print_line_spacing.get()
        include_barcode = self.print_include_barcode_var.get()
        print_encoding = self.print_encoding_var.get()

        if file_type == "pdf":
            self.print_pdf(sample_data, to_file=True)
            return

        try:
            content_to_print_bytes = self._prepare_print_content(
                template_str,
                sample_data,
                line_spacing,
                include_barcode,
                print_encoding,
                for_two_way_template=(current_template_mode == "TWO_WAY")
            )
            self._print_content(content_to_print_bytes, job_name="Test Ticket", to_file=to_file, file_type="txt")

        except Exception as e:
            self.msg_box.showerror("Printing Error", f"An error occurred while preparing the test print: {e}")

    def _get_transaction_data_for_print(self, trans_id: int) -> dict | None:
        """
        Retrieves a single transaction from the database and formats the data for printing.
        """
        try:
            with sqlite3.connect(DB_FILE) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM transactions WHERE id = ?", (trans_id,))
                row = cursor.fetchone()
                if row:
                    data = dict(row)
                    # Format weight values with the correct number of decimal places
                    format_string = f"%.{self.decimal_places}f"
                    for key in ['gross_weight', 'tare_weight', 'net_weight']:
                        data[key] = format_string % data[key] if data[key] is not None else "N/A"
                    data['ticket_no'] = str(data['ticket_no']) if data['ticket_no'] is not None else "N/A"
                    data['date_printed'] = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
                    data['barcode'] = str(data['ticket_no'])
                    
                    # Always calculate price based on most recent weight using timestamps
                    weight_for_price = 0.0
                    weight_type = data.get('weight_type', '')
                    
                    # Determine which weight is most recent based on timestamps
                    gross_datetime = None
                    tare_datetime = None
                    
                    if data.get('gross_date') and data.get('gross_time'):
                        try:
                            gross_datetime = datetime.strptime(f"{data['gross_date']} {data['gross_time']}", "%m/%d/%Y %I:%M:%S %p")
                        except (ValueError, TypeError):
                            pass
                    
                    if data.get('tare_date') and data.get('tare_time'):
                        try:
                            tare_datetime = datetime.strptime(f"{data['tare_date']} {data['tare_time']}", "%m/%d/%Y %I:%M:%S %p")
                        except (ValueError, TypeError):
                            pass
                    
                    # Use the most recent weight based on timestamps
                    if gross_datetime and tare_datetime:
                        if tare_datetime > gross_datetime:
                            # Tare is more recent
                            if data.get('tare_weight') and data['tare_weight'] != "N/A":
                                weight_for_price = float(data['tare_weight'])
                        else:
                            # Gross is more recent
                            if data.get('gross_weight') and data['gross_weight'] != "N/A":
                                weight_for_price = float(data['gross_weight'])
                    elif gross_datetime:
                        # Only gross has timestamp
                        if data.get('gross_weight') and data['gross_weight'] != "N/A":
                            weight_for_price = float(data['gross_weight'])
                    elif tare_datetime:
                        # Only tare has timestamp
                        if data.get('tare_weight') and data['tare_weight'] != "N/A":
                            weight_for_price = float(data['tare_weight'])
                    else:
                        # Fallback to weight type logic if no timestamps
                        if "ONE WAY" in weight_type:
                            if data.get('gross_weight') and data['gross_weight'] != "N/A":
                                weight_for_price = float(data['gross_weight'])
                        elif "GROSS-TARE" in weight_type:
                            if data.get('gross_weight') and data['gross_weight'] != "N/A":
                                weight_for_price = float(data['gross_weight'])
                        elif "TARE-GROSS" in weight_type:
                            if data.get('tare_weight') and data['tare_weight'] != "N/A":
                                weight_for_price = float(data['tare_weight'])
                        else:
                            if data.get('gross_weight') and data['gross_weight'] != "N/A":
                                weight_for_price = float(data['gross_weight'])
                    
                    # Calculate price based on most recent weight
                    unit_price, total_price = self.calculate_price(weight_for_price)
                    data['unit_price'] = f"{unit_price:.2f}"
                    data['total_price'] = f"{total_price:.2f}"
                    
                    # Handle operator information
                    operator1 = data.get('operator') if data.get('operator') else "N/A"
                    operator2 = data.get('operator2') if data.get('operator2') else None
                    
                    # For two-way transactions with both operators, format as operator1/operator2
                    if operator2 and operator2 != "N/A":
                        data['logged_in_user'] = f"{operator1}/{operator2}"
                    else:
                        data['logged_in_user'] = operator1
                    
                    return data
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to retrieve transaction data for printing: {e}")
            return None
        return None

    def _build_preview_text(self, template_str: str, data: dict, line_spacing: int) -> str:
        """Build a human-readable preview text from template and data without ESC/P controls."""
        processed_template = template_str.replace("\\n", "\n")
        # Create a safe dict that returns empty strings for missing keys
        class SafeDict(dict):
            def __missing__(self, key):
                return ''
        try:
            formatted = processed_template.format_map(SafeDict(data))
        except Exception as e:
            formatted = processed_template + f"\n\n[FORMAT ERROR: {e}]"
        # Replace barcode placeholder with a readable marker
        formatted = formatted.replace("{barcode}", f"[BARCODE: {data.get('ticket_no','N/A')}]")
        formatted = self._wrap_text_for_print(formatted, max_width=40)
        if line_spacing and int(line_spacing) > 0:
            formatted = formatted.replace("\n", "\n" * (1 + int(line_spacing)))
        return formatted

    def _open_print_preview(self, data: dict, source: str | None = None):
        """Open a modal print preview window with actions: Print, Save as PDF, Save as Text, Close."""
        weight_type = data.get('weight_type', '')
        line_spacing = self.print_line_spacing.get()
        include_barcode = self.print_include_barcode_var.get()
        print_encoding = self.print_encoding_var.get()
        
        # Always use print templates from one_way_print_template_var and two_way_print_template_var
        if weight_type == "ONE WAY WEIGHING":
            template_str = self.one_way_print_template_var.get()
        else:
            template_str = self.two_way_print_template_var.get()
        
        preview_text = self._build_preview_text(template_str, data, line_spacing)
        use_ticket_format = False

        # Get page dimensions based on configured size
        page_width_px, page_height_px = self._get_page_dimensions_for_preview(weight_type)
        
        # Add padding for window border and buttons
        padding = 50
        window_width = page_width_px + padding
        window_height = page_height_px + padding + 50  # Extra space for buttons

        win = tk.Toplevel(self.root)
        win.title("Print Preview")
        win.transient(self.root)
        win.grab_set()
        win.geometry(f"{window_width}x{window_height}")

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill="both", expand=True)

        # Get header font size from config
        header_font_size_key = "ticket_header_font_size_one_way" if weight_type == "ONE WAY WEIGHING" else "ticket_header_font_size_two_way"
        header_font_size = int(self.config.get(header_font_size_key, 13))
        detail_font_size = self.print_template_font_size_var.get()

        text = tk.Text(
            frame,
            wrap="none",  # Disable wrapping to prevent indentation issues
            font=(
                self.print_template_font_family_var.get(),
                detail_font_size,
                "bold" if self.print_template_font_bold_var.get() else "normal",
            ),
        )
        text.pack(fill="both", expand=True)
        
        # Configure tabs to prevent unwanted indentation
        text.config(tabs=("1c"))  # Set tab width to 1 character
        text.tag_configure("header", tabs=("1c"))
        text.tag_configure("detail", tabs=("1c"))
        
        # Configure text tags for header and detail styles
        text.tag_configure(
            "header",
            font=(
                self.print_template_font_family_var.get(),
                header_font_size,
                "bold" if self.print_template_font_bold_var.get() else "normal",
            )
        )
        text.tag_configure(
            "detail",
            font=(
                self.print_template_font_family_var.get(),
                detail_font_size,
                "bold" if self.print_template_font_bold_var.get() else "normal",
            )
        )
        
        # Insert text with proper styling - identify header lines
        lines = preview_text.split('\n')
        header_end_index = 0
        for i, line in enumerate(lines):
            if 'TICKET NO' in line.upper():
                header_end_index = i
                break
        
        for idx, line in enumerate(lines):
            if idx < header_end_index:
                text.insert(tk.END, line + '\n', 'header')
            else:
                text.insert(tk.END, line + '\n', 'detail')
        
        text.config(state="disabled")

        btns = ttk.Frame(frame)
        btns.pack(fill="x", pady=(8, 0))

        def do_print():
            try:
                # Get the number of copies from printer settings
                num_copies = self.print_copies_var.get()
                
                if use_ticket_format:
                    # Use the formatted content directly from preview
                    content_to_print_bytes = preview_text.encode(print_encoding)
                else:
                    content_to_print_bytes = self._prepare_print_content(
                        template_str,
                        data,
                        line_spacing,
                        include_barcode,
                        print_encoding,
                        for_two_way_template=(weight_type != "ONE WAY WEIGHING"),
                    )
                
                # Print the specified number of copies
                for copy_num in range(num_copies):
                    self._print_content(
                        content_to_print_bytes,
                        job_name=f"Ticket {data.get('ticket_no','')} (Copy {copy_num + 1}/{num_copies})",
                        to_file=False,
                        file_type="txt",
                        show_dialog=False  # Suppress individual dialogs for each copy
                    )
                
                # Show single success dialog for all copies
                self.msg_box.showinfo("Print Success", f"Ticket printed successfully! ({num_copies} copy/copies sent)")
            except Exception as e:
                self.msg_box.showerror("Print Error", f"Failed to print: {e}")

        def save_pdf():
            try:
                self.print_pdf(data, to_file=True)
            except Exception as e:
                self.msg_box.showerror("Save Error", f"Failed to save PDF: {e}")

        def save_txt():
            try:
                if use_ticket_format:
                    # Use the formatted content directly from preview
                    content_to_print_bytes = preview_text.encode(print_encoding)
                else:
                    content_to_print_bytes = self._prepare_print_content(
                        template_str,
                        data,
                        line_spacing,
                        include_barcode,
                        print_encoding,
                        for_two_way_template=(weight_type != "ONE WAY WEIGHING"),
                    )
                self._print_content(
                    content_to_print_bytes,
                    job_name=f"Ticket {data.get('ticket_no','')}",
                    to_file=True,
                    file_type="txt",
                )
            except Exception as e:
                self.msg_box.showerror("Save Error", f"Failed to save text: {e}")

        ttk.Button(btns, text="Print", command=do_print).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="Save as PDF", command=save_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="Save as Text", command=save_txt).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="Close", command=win.destroy).pack(side=tk.RIGHT)

    def _open_preview_for_selection(self, tree, source: str):
        """Fetch selected transaction and open preview. Single selection only for preview."""
        selected_items = tree.selection()
        if not selected_items:
            self.msg_box.showwarning("No Selection", "Please select a transaction to preview.")
            return
        if len(selected_items) > 1:
            self.msg_box.showwarning("Selection Error", "Please select only one transaction to preview.")
            return
        item = selected_items[0]
        trans_id = tree.item(item)['values'][0]
        data = self._get_transaction_data_for_print(trans_id)
        if not data:
            self.msg_box.showwarning("No Data", "Could not retrieve transaction data for preview.")
            return
        self._open_print_preview(data, source)

    def print_selected_ticket(self, tree, source: str, to_file: bool = True, file_type: str = "pdf"):
        """
        Prints or exports a selected ticket(s) to a PDF file.
        If 'to_file' is True, it generates a temporary PDF file,
        opens it in the system's default viewer, and schedules deletion.
        If 'to_file' is False, it uses the original logic (e.g., direct printer).
        """
        # --- Original logic for non-to_file cases ---
        if not to_file:
            return # Return early if direct printing is intended

        # --- Modified Logic for 'to_file' = True (New Behavior) ---

        # --- Step 1: Get selected items ---
        selected_items = tree.selection()
        if not selected_items:
            self.msg_box.showwarning("No Selection", "Please select a transaction to export.")
            return

        # --- Step 2: Prepare Data ---
        # Handle multiple selections for the report tab
        if source == "Report":
            transactions_to_print = []
            for item in selected_items:
                trans_id = tree.item(item)['values'][0] # Assuming ID is first column
                data = self._get_transaction_data_for_print(trans_id)
                if data:
                    transactions_to_print.append(data)
            if not transactions_to_print:
                self.msg_box.showwarning("No Data", "No valid transaction data could be retrieved for printing.")
                return
            # --- Use Batch PDF Generation for Report ---
            # For simplicity, use the batch method directly for report selections
            # This might need adjustment if you want different handling for report vs. single entries
            self.print_pdf_batch(transactions_to_print)
            return # Early return after batch PDF creation if needed

        # --- Handle Single Item Selection (Pending/Completed Records) ---
        # Assuming only one item is selected for pending/completed records tabs
        # (This logic assumes single selection for those tabs)
        if len(selected_items) != 1:
            self.msg_box.showwarning("Selection Error", f"For {source} tab, please select exactly one item.")
            return

        item = selected_items[0]
        trans_id = tree.item(item)['values'][0] # Assuming ID is first column

        # --- Step 3: Get Data for Print ---
        data = self._get_transaction_data_for_print(trans_id)
        if not data:
            self.msg_box.showwarning("No Data", "Could not retrieve transaction data for printing.")
            return

        # --- Step 4: Generate Temporary PDF File ---
        try:
            # Create a temporary file with .pdf extension
            # Using 'delete=False' initially so we can pass the path to the PDF generator
            temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            temp_file.close() # Close the file handle immediately after creation
            temp_pdf_path = temp_file.name # Get the path

            # --- Step 5: Generate PDF Content ---
            # --- Adapted from _generate_pdf_ticket ---
            # Get the appropriate template based on weight type
            template_str = self.one_way_print_template_var.get() if data.get('weight_type', '') == "ONE WAY WEIGHING" else self.two_way_print_template_var.get()

            # --- Create document template ---
            doc_template = SimpleDocTemplate(
                temp_pdf_path, 
                pagesize=self._get_pdf_page_size(),
                leftMargin=0.1 * inch,  # Minimal left margin
                rightMargin=0.1 * inch, # Reasonable right margin
                topMargin=0.1 * inch,   # Reasonable top margin
                bottomMargin=0.1 * inch # Reasonable bottom margin
            )

            # --- Styles ---
            styles = getSampleStyleSheet()
            
            # - Create custom style with proper paragraph spacing -
            pdf_style = ParagraphStyle(name='CustomStyle',
                parent=styles['Normal'],
                fontSize=self.pdf_print_template_font_size_var.get(),
                fontName=self._get_pdf_font_name(),  # You'll need to implement this
                bold=self.pdf_print_template_font_bold_var.get(),
                leading=self.pdf_print_template_font_size_var.get(),  # <-- KEY CHANGE: Set leading equal to font size
                spaceAfter=5,  # <-- KEEP: This adds space after each paragraph
            )

            # Create header style with the configured header font size
            weight_type = data.get('weight_type', '')
            header_font_size_key = "ticket_header_font_size_one_way" if weight_type == "ONE WAY WEIGHING" else "ticket_header_font_size_two_way"
            header_font_size = int(self.config.get(header_font_size_key, 13))
            
            pdf_header_style = ParagraphStyle(name='HeaderStyle',
                parent=styles['Normal'],
                fontSize=header_font_size,
                fontName=self._get_pdf_font_name(),
                bold=self.pdf_print_template_font_bold_var.get(),
                leading=header_font_size,
                spaceAfter=5,
            )

            # --- Story (Content) ---
            story = []

            # --- Register Font if needed ---
            self._register_pdf_font(self.pdf_print_template_font_family_var.get()) # Uncomment if needed

            # --- Apply Template Content ---
            # This is the key part - using the correct template for the data type
            # Create safe dict for template formatting
            class SafeDict(dict):
                def __missing__(self, key):
                    return ''
            
            formatted_text = template_str.format_map(SafeDict(data))
            lines = formatted_text.split('\n')
            
            # Find where the header ends and transaction details begin
            # The header ends before the first line containing "TICKET NO" (case-insensitive)
            header_end_index = 0
            for i, line in enumerate(lines):
                if 'TICKET NO' in line.upper():
                    header_end_index = i
                    break
            
            # --- Process lines, handling barcodes ---
            line_spacing_inch = self.print_line_spacing.get() * 0.05 * inch
            
            # Add spacing before the first content
            story.append(Spacer(1, 0.01 * inch))
            
            for idx, line in enumerate(lines):
                if '{barcode}' in line:
                    ticket_no = data.get('ticket_no', 'N/A')
                    if str(ticket_no).isdigit():
                        # Add barcode
                        barcode128 = code128.Code128(str(ticket_no))
                        drawing = Drawing(200, 50)
                        drawing.add(barcode128)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1 * inch))
                    else:
                        # If no valid ticket number, just add an empty spacer
                        story.append(Spacer(1, 0.1 * inch))
                else:
                    # Use header style for header lines, regular style for transaction details
                    style = pdf_header_style if idx < header_end_index else pdf_style
                    # Add regular paragraph with proper spacing
                    story.append(Paragraph(line, style))
                    story.append(Spacer(1, line_spacing_inch))
            
            # - ADD LOGO -
            # - ADD LOGO -
            logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
            if os.path.exists(logo_path):
                # Optional: Set logo dimensions (adjust as needed)
                logo_width = 2.80 * inch
                logo_height = 0.40 * inch
                # Create a Table with one cell to center the logo
                logo_table = Table([[Image(logo_path, width=logo_width, height=logo_height)]], colWidths=[logo_width])
                logo_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                               ('ALIGN', (0, 0), (-1, -1), 'CENTER')]))

                # Insert the logo table BEFORE the first piece of content (the company name)
                # This will place it centered above the first paragraph.
                story.insert(0, logo_table)  # <-- This line stays

                # Remove the following line (or comment it out):
                # story.insert(0, Spacer(0, 0.1 * inch))  # This was causing the extra space

            # - END ADD LOGO -
            
            # --- Build Document ---
            try:
                doc_template.build(story)
                
                # --- Step 6: Open PDF in System Viewer ---
                # --- Step 7: Schedule Deletion ---
                # --- Schedule Deletion in Background Thread ---
                def delete_temp_file(path):
                    """Function to delete the temporary file after delay."""
                    time.sleep(30) # Wait 30 seconds
                    try:
                        os.unlink(path) # Remove the file
                        print(f"Temporary file {path} deleted.")
                    except OSError as e:
                        print(f"Error deleting temporary file {path}: {e}")

                # Start the deletion thread
                deletion_thread = threading.Thread(target=delete_temp_file, args=(temp_pdf_path,), daemon=True)
                deletion_thread.start()

                # --- Open the PDF ---
                # --- Platform-specific opening ---
                try:
                    if sys.platform == "win32":
                        os.startfile(temp_pdf_path) # Windows
                    elif sys.platform == "darwin":
                        subprocess.run(["open", temp_pdf_path]) # macOS
                    else:
                        subprocess.run(["xdg-open", temp_pdf_path]) # Linux
                except Exception as e:
                    self.msg_box.showerror("Opening Error", f"Failed to open PDF viewer: {e}")
                    # Even if opening fails, the file is still created and scheduled for deletion
                    # Optionally, inform user about manual opening or deletion

                self.msg_box.showinfo("PDF Generated", f"PDF created temporarily at:\n{temp_pdf_path}\nADJUST TO HIGHER DPI SETTINGS OR RESOLUTION TO PRINT CLEARLY.")

            except Exception as e:
                self.msg_box.showerror("PDF Error", f"An error occurred while generating the PDF: {e}")
                # Clean up the temporary file if PDF generation failed
                try:
                    os.unlink(temp_pdf_path)
                except OSError:
                    pass # Already deleted or doesn't exist

        except Exception as e:
            self.msg_box.showerror("Error", f"An unexpected error occurred: {e}")
    
    def print_pdf(self, data: dict, to_file: bool = True):
        """
        Creates a PDF from a single transaction using coordinate positioning.
        """
        if not PDF_PRINTING_ENABLED:
            self.msg_box.showerror("PDF Error", "PDF generation is disabled because ReportLab is not installed. Please run 'pip install ReportLab' from your command line.")
            return
        
        if to_file:
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Save Transaction to PDF",
                parent=self.root
            )
            if not filename:
                return
            pdf_path = filename
        else:
            # Create a temporary file for printing
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            pdf_path = temp_file.name
            temp_file.close()
        
        try:
            # Use coordinate-based PDF generation
            self._generate_coordinate_pdf(data, pdf_path)
            
            if not to_file:
                # If printing directly, send to printer
                self._print_pdf_file(pdf_path)
                os.remove(pdf_path)  # Clean up temp file
                
        except Exception as e:
            self.msg_box.showerror("PDF Error", f"Failed to generate PDF: {e}")
            if not to_file and os.path.exists(pdf_path):
                os.remove(pdf_path)


    def print_pdf_batch(self, data_list: list[dict]):
        """
        Generates a multi-page PDF from a list of transactions.
        """
        if not PDF_PRINTING_ENABLED:
            self.msg_box.showerror("PDF Error", "PDF generation is disabled because ReportLab is not installed.\n\nPlease run 'pip install ReportLab' from your command line.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Save Multiple Completed Records to PDF",
            parent=self.root
        )
        if not filename:
            return

        doc_template = SimpleDocTemplate(
            filename, 
            pagesize=self._get_pdf_page_size(),
            leftMargin=0.1 * inch,  # Minimal left margin
            rightMargin=0.1 * inch, # Reasonable right margin
            topMargin=0.1 * inch,   # Reasonable top margin
            bottomMargin=0.1 * inch # Reasonable bottom margin
        )
        styles = getSampleStyleSheet()
        story = []
        
        self._register_pdf_font(self.pdf_print_template_font_family_var.get())
        pdf_style = ParagraphStyle(
            'CustomStyle',
            parent=styles['Normal'],
            fontName=self._get_pdf_font_name(),
            fontSize=self.pdf_print_template_font_size_var.get(),
            alignment=TA_LEFT
        )
        if self.pdf_print_template_font_bold_var.get():
            pdf_style.fontName = self._get_pdf_font_name() + "-Bold"
            
        line_spacing_inch = self.print_line_spacing.get() * 0.05 * inch

        for i, data in enumerate(data_list):
            template_str = self.one_way_print_template_var.get() if data.get('weight_type', '') == "ONE WAY WEIGHING" else self.two_way_print_template_var.get()

            # Create safe dict for template formatting
            class SafeDict(dict):
                def __missing__(self, key):
                    return ''
            
            formatted_text = template_str.format_map(SafeDict(data))
            lines = formatted_text.split('\n')
            
            for line in lines:
                if '{barcode}' in line:
                    ticket_no = data.get('ticket_no', 'N/A')
                    if str(ticket_no).isdigit():
                        barcode128 = code128.Code128(ticket_no)
                        drawing = Drawing(200, 50)
                        drawing.add(barcode128)
                        story.append(drawing)
                        story.append(Spacer(1, 0.1 * inch))
                else:
                    story.append(Paragraph(line, pdf_style))
                    story.append(Spacer(1, line_spacing_inch))
            
            # Add a page break between tickets, but not after the last one
            if i < len(data_list) - 1:
                story.append(PageBreak())

        try:
            doc_template.build(story)
            self.msg_box.showinfo("PDF Created", f"Batch PDF created successfully at:\n{filename}")
        except Exception as e:
            self.msg_box.showerror("PDF Error", f"An error occurred while generating the batch PDF: {e}")


    def _get_pdf_page_size(self):
        """Get the appropriate page size based on settings."""
        page_size = A6  # Default
        
        if self.pdf_page_size_var.get() == "A6":
            page_size = (3.15 * inch, 5.5* inch)
        elif self.pdf_page_size_var.get() == "A6":
            page_size = (3.15 * inch, 5.5* inch)
        
        # Handle orientation
        if self.pdf_orientation_var.get() == "landscape":
            page_size = (page_size[1], page_size[0])
        
        return page_size

    def _register_pdf_font(self, font_family: str):
        """
        Registers a TTF font for ReportLab, if it's available.
        This allows for more font choices beyond the standard 14.
        """
        # This is a placeholder for a more robust font registration system.
        # Currently, it only handles the standard fonts and a custom MonospaceFont
        # which is registered below as a fallback.
        if font_family in ["Helvetica", "Times-Roman", "Courier"]:
            pass # These are built-in
        else:
            try:
                # Assuming 'MonospaceFont.ttf' is in a 'fonts' directory
                # or some other known path. We'll add this check.
                font_path = os.path.join(os.path.dirname(__file__), "assets", "MonospaceFont.ttf")
                if os.path.exists(font_path) and not self.registered_pdf_font:
                    registerFont(TTFont('MonospaceFont', font_path))
                    addMapping('MonospaceFont', 0, 0, 'MonospaceFont')
                    addMapping('MonospaceFont', 0, 1, 'MonospaceFont-Bold') # Use bold variant if available
                    self.registered_pdf_font = True
            except TTFError:
                pass # Font file not found or invalid
    
    def _get_pdf_font_name(self):
        """
        Returns the font name to use for PDF generation based on selected font.
        """
        font_family = self.pdf_print_template_font_family_var.get()
        if font_family in ["Helvetica", "Times-Roman", "Courier"]:
            if self.pdf_print_template_font_bold_var.get():
                return f"{font_family}-Bold"
            return font_family
        else:
            # Use the generic registered font if not a standard one
            if self.pdf_print_template_font_bold_var.get():
                return "MonospaceFont-Bold"
            return "MonospaceFont"

    def _print_pdf_file(self, file_path):
        """
        Prints a PDF file to the selected printer.
        This uses the default system PDF viewer to send the print job.
        This is a platform-dependent solution.
        """
        if sys.platform == "win32":
            try:
                import win32print  # type: ignore
                import win32api  # type: ignore
                # Use the selected printer name or the default
                printer_name = self.selected_printer_var.get() or win32print.GetDefaultPrinter()
                
                # Check if the printer is online before attempting to print
                h_printer = win32print.OpenPrinter(printer_name, {'DesiredAccess': win32print.PRINTER_ALL_ACCESS})
                printer_status = win32print.GetPrinter(h_printer, 2)['Status']
                win32print.ClosePrinter(h_printer)
                
                if printer_status != win32print.PRINTER_STATUS_PAUSED and \
                   printer_status != win32print.PRINTER_STATUS_OFFLINE and \
                   printer_status != win32print.PRINTER_STATUS_ERROR:
                       
                    import subprocess
                    for i in range(self.print_copies_var.get()):
                        win32api.ShellExecute(
                            0,
                            "printto",
                            file_path,
                            printer_name,
                            ".",
                            0
                        )
                else:
                    self.msg_box.showerror("Printing Error", f"Printer '{printer_name}' is not ready or is offline. Please check the printer status.")

            except Exception as e:
                self.msg_box.showerror("Printing Error", f"Failed to print PDF: {e}\n\nEnsure a default PDF reader is installed and configured to handle 'printto' actions.")
        else:
            self.msg_box.showwarning("Printing Not Supported", "Direct PDF printing is not yet supported on this operating system. Please print the saved file manually.")

    def _wrap_text_for_print(self, text: str, max_width: int = 40) -> str:
        """
        Wraps text to fit within max_width characters.
        For lines with colons (label: value), value starts on the same line after colon.
        Continuation lines are indented to align with the start of the value.
        """
        lines = text.split('\n')
        wrapped_lines = []
        
        for line in lines:
            if len(line) <= max_width:
                wrapped_lines.append(line)
            else:
                # Check if line contains a colon (label: value format)
                if ':' in line:
                    colon_pos = line.find(':')
                    label_part = line[:colon_pos + 1]  # Include colon
                    value_part = line[colon_pos + 1:].strip()
                    
                    # Calculate indent for wrapped lines - align with where value starts
                    # Value starts at colon_pos + 2 (": ")
                    value_indent = ' ' * (colon_pos + 2)  # +2 for ": "
                    indent_width = len(value_indent)
                    
                    # First line: label + colon + space + as much value as fits
                    first_line_prefix = label_part + ' '
                    available_for_value = max_width - len(first_line_prefix)
                    
                    if available_for_value <= 0:
                        # Label is too long by itself, put value on next line
                        wrapped_lines.append(label_part)
                        remaining = value_part
                        available_width = max_width - indent_width
                    else:
                        # Fit value starting on same line
                        remaining = value_part
                        current_line = ""
                        words = remaining.split()
                        words_on_first = 0
                        
                        # Pack as many words as fit on first line
                        for i, word in enumerate(words):
                            test = (current_line + (" " if current_line else "") + word).strip()
                            if len(first_line_prefix) + len(test) <= max_width:
                                current_line = test
                                words_on_first = i + 1
                            else:
                                break
                        
                        wrapped_lines.append(first_line_prefix + current_line)
                        remaining = ' '.join(words[words_on_first:])
                        available_width = max_width - indent_width
                    
                    # Wrap remaining value with proper indentation
                    while remaining:
                        if len(remaining) <= available_width:
                            wrapped_lines.append(value_indent + remaining)
                            remaining = ''
                        else:
                            # Find last space within available width
                            chunk = remaining[:available_width]
                            last_space = chunk.rfind(' ')
                            
                            if last_space > 0:
                                # Wrap at word boundary
                                wrapped_lines.append(value_indent + chunk[:last_space])
                                remaining = remaining[last_space:].lstrip()
                            else:
                                # No space found, break at max width
                                wrapped_lines.append(value_indent + chunk)
                                remaining = remaining[available_width:]
                else:
                    # Line without colon - wrap normally without indentation
                    wrapped_lines.append(line[:max_width])
                    line = line[max_width:]
                    while len(line) > max_width:
                        wrapped_lines.append(line[:max_width])  # No indentation for continuation
                        line = line[max_width:]
                    if line:
                        wrapped_lines.append(line)  # No indentation for last line
        
        return '\n'.join(wrapped_lines)

    def _prepare_print_content(self, template_str: str, data_map: dict,
                                line_spacing: int, include_barcode: bool,
                                print_encoding: str, for_two_way_template: bool = False) -> bytes:
        """
        Prepares the print content from a template string and data dictionary.
        It handles placeholders, line spacing, and optional barcode generation.
        This version is specifically for ESC/P compatible printers.
        """
        # ESC/P commands for an Epson dot matrix printer
        init_printer = b'\x1B\x40'
        # This command is for a printer with a continuous paper feed.
        form_feed = b'\x0C'
        align_left = b'\x1B\x61\x00'
        align_center = b'\x1B\x61\x01'
        
        # Prepare content string with newlines and spacing
        processed_template = template_str.replace("\\n", "\n")
        if line_spacing > 0:
            processed_template = processed_template.replace("\n", "\n" * (1 + line_spacing))
        
        try:
            formatted_content_str = processed_template.format(**data_map)
        except KeyError as e:
            formatted_content_str = f"ERROR: Missing placeholder in template: {e}. Check your template.\n\n" + processed_template
        except Exception as e:
            formatted_content_str = f"ERROR: An error occurred formatting the template: {e}.\n\n" + processed_template

        formatted_content_str = self._wrap_text_for_print(formatted_content_str, max_width=40)
        final_content_bytes = formatted_content_str.encode(print_encoding, errors='replace')
        
        # Barcode generation and insertion
        if include_barcode:
            ticket_no_str = str(data_map.get('ticket_no', 'N/A'))
            if ticket_no_str.isdigit():
                # Correct ESC/P commands for barcode printing on Epson LX-310
                # Set barcode height: GS h n
                set_height = b'\x1D\x68\x64'
                # Select Code 128: GS k m n
                select_code128 = b'\x1D\x6B\x49'
                barcode_data_bytes = ticket_no_str.encode('ascii')
                barcode_len_byte = len(barcode_data_bytes).to_bytes(1, 'little')
                
                barcode_command_sequence = (
                    align_center + 
                    set_height + 
                    select_code128 + 
                    barcode_len_byte + 
                    barcode_data_bytes +
                    b'\x00' + # Null terminator
                    b'\r\n\r\n' # Add line feeds after barcode
                )
                
                # Find placeholder position and replace
                placeholder_pos = final_content_bytes.find(b'{barcode}')
                if placeholder_pos != -1:
                    final_content_bytes = (
                        final_content_bytes[:placeholder_pos] +
                        barcode_command_sequence +
                        final_content_bytes[placeholder_pos + len(b'{barcode}'):]
                    )
                else:
                    final_content_bytes = final_content_bytes.replace(b'{barcode}', b'[BARCODE DATA HERE]\n')
            else:
                final_content_bytes = final_content_bytes.replace(b'{barcode}', b'')

        # Concatenate all parts
        return init_printer + align_left + final_content_bytes + form_feed

    def _print_content(self, content_to_print: bytes, job_name: str = "Truck Scale Ticket", 
                   print_encoding: str = "utf-8", to_file: bool = False, file_type: str = "txt", show_dialog: bool = True):
        """
        Modified to support coordinate-based PDF printing with proper spacing.
        """
        if to_file and file_type == "txt":
            filename = filedialog.asksaveasfilename(
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                title="Save Print Content to File",
                parent=self.root
            )
            if not filename:
                return
            if not filename.lower().endswith(".txt"):
                filename += ".txt"
            try:
                with open(filename, 'wb') as f:
                    f.write(content_to_print)
                self.msg_box.showinfo("Success", "Content saved to file successfully!")
            except Exception as e:
                self.msg_box.showerror("Save Error", f"Failed to save file: {e}")
            return
        
        elif to_file and file_type == "pdf":
            # Note: PDF generation from content_to_print (bytes) is not directly supported.
            # Use the print_pdf() method with data dictionary instead.
            self.msg_box.showwarning("PDF Warning", "PDF generation from raw content is not supported.\nUse the print_pdf() method with transaction data instead.")
            return
        
        # For physical printing (existing code)
        if sys.platform == "win32":
            try:
                import win32print  # type: ignore
                import win32api  # type: ignore
                import win32con  # type: ignore
                
                printer_name = self.selected_printer_var.get()
                if not printer_name:
                    try:
                        printer_name = win32print.GetDefaultPrinter()
                    except Exception:
                        self.msg_box.showerror("Printer Error", "No printer selected and no default printer is set.")
                        return
                
                # Print to the selected printer
                hPrinter = win32print.OpenPrinter(printer_name)
                try:
                    hJob = win32print.StartDocPrinter(hPrinter, 1, ("Truck Scale Ticket", None, "RAW"))
                    try:
                        win32print.StartPagePrinter(hPrinter)
                        win32print.WritePrinter(hPrinter, content_to_print)
                        win32print.EndPagePrinter(hPrinter)
                    finally:
                        win32print.EndDocPrinter(hPrinter)
                finally:
                    win32print.ClosePrinter(hPrinter)
                    
                if show_dialog:
                    self.msg_box.showinfo("Print Success", "Ticket printed successfully!")
            except Exception as e:
                self.msg_box.showerror("Print Error", f"Failed to print: {e}")
        else:
            # For non-Windows systems, attempt to use lpr
            try:
                import subprocess
                import tempfile
                import os
                
                # Write content to temporary file
                with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.txt') as tmp:
                    tmp.write(content_to_print)
                    tmp_path = tmp.name
                
                # Try to print with lpr
                subprocess.run(['lpr', '-P', self.selected_printer_var.get(), tmp_path], 
                              check=True, capture_output=True)
                os.unlink(tmp_path)  # Clean up temp file
                
                if show_dialog:
                    self.msg_box.showinfo("Print Success", "Ticket sent to printer successfully!")
            except Exception as e:
                self.msg_box.showerror("Print Error", f"Failed to send to printer: {e}")
    
    def _generate_coordinate_pdf(self, data, filename):
        """Generate a PDF with proper template formatting including all details."""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter, A6
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Get the appropriate template based on weight type
            weight_type = data.get('weight_type', '')
            template_str = self.one_way_print_template_var.get() if weight_type == "ONE WAY WEIGHING" else self.two_way_print_template_var.get()
            
            # Format the template with data
            class SafeDict(dict):
                def __missing__(self, key):
                    return ''
            formatted_text = template_str.format_map(SafeDict(data))
            
            # Determine page size
            page_size = A6  # Default to A6
            if self.pdf_page_size_var.get() == "Letter":
                page_size = letter
                
            # Create canvas
            c = canvas.Canvas(filename, pagesize=page_size)
            
            # Set base font
            font_name = self._get_pdf_font_name()
            font_size = self.pdf_print_template_font_size_var.get()
            header_font_size_key = "ticket_header_font_size_one_way" if weight_type == "ONE WAY WEIGHING" else "ticket_header_font_size_two_way"
            header_font_size = int(self.config.get(header_font_size_key, 13))
            
            # Register custom fonts if needed
            try:
                c.setFont(font_name, font_size)
            except:
                c.setFont("Helvetica", font_size)
            
            # Parse lines and determine which are header lines
            lines = formatted_text.split('\n')
            y_position = page_size[1] - 20  # Start from top with margin
            line_height = font_size + 2  # Line spacing
            
            # Find where the header ends (before TICKET NO line)
            header_end_index = 0
            for i, line in enumerate(lines):
                if 'TICKET NO' in line.upper():
                    header_end_index = i
                    break
            
            # Draw lines with proper font size for header vs details
            for idx, line in enumerate(lines):
                if line.strip():  # Skip empty lines for positioning, but draw them for spacing
                    # Determine if this is a header line or detail line
                    if idx < header_end_index:
                        # Header line - use header font size
                        c.setFont(font_name, header_font_size)
                        current_font_size = header_font_size
                        current_line_height = header_font_size + 2
                    else:
                        # Detail line - use standard font size
                        c.setFont(font_name, font_size)
                        current_font_size = font_size
                        current_line_height = line_height
                    
                    c.drawString(10, y_position, line)
                    y_position -= current_line_height
                else:
                    # Empty line - add spacing
                    y_position -= line_height
            
            # Draw operator signature line
            c.setFont(font_name, font_size)
            y_position -= 10
            c.drawString(10, y_position, "___________________")
            
            # Save the PDF
            c.save()
            
        except Exception as e:
            self.msg_box.showerror("PDF Error", f"Failed to create PDF: {e}")
        
    def _get_pdf_font_name(self):
        """Get the appropriate font name for PDF generation."""
        font_family = self.pdf_print_template_font_family_var.get()
        if font_family in ['Helvetica', 'Times-Roman', 'Courier']:
            return font_family
        else:
            # Return default font if unknown
            return 'Helvetica'
    
    def _print_to_pdf(self, transaction_data):
        """Print transaction data to PDF with enhanced formatting."""
        try:
            # Get the current template from the editor
            template_str = self.template_text_widget.get("1.0", tk.END)
            
            # Prepare data for formatting
            formatted_data = self._prepare_transaction_data_for_printing(transaction_data)
            
            # Create a temporary PDF file
            temp_pdf_path = tempfile.mktemp(suffix=".pdf")
            
            # Determine page size and orientation
            page_size = self._get_pdf_page_size()
            if self.pdf_orientation_var.get() == "landscape":
                page_size = (page_size[1], page_size[0])
                
            # Create PDF document
            doc = SimpleDocTemplate(
                temp_pdf_path,
                pagesize=page_size,
                leftMargin=self.margin_left_var.get() * inch,
                rightMargin=self.margin_right_var.get() * inch,
                topMargin=self.margin_top_var.get() * inch,
                bottomMargin=self.margin_bottom_var.get() * inch
            )
            
            story = []
            
            # Define styles with proper leading (line spacing) - THIS IS THE KEY IMPROVEMENT
            styles = getSampleStyleSheet()
            normal_style = styles["Normal"]
            normal_style.leading = 14  # This controls line spacing within paragraphs
            normal_style.spaceAfter = 6  # This adds space after each paragraph
            
            # Format the template text with the transaction data
            formatted_content = template_str.format(**formatted_data)
            
            # Process content lines
            lines = formatted_content.split('\n')
            line_spacing_inch = self.print_line_spacing.get() * 0.05 * inch  # Use user-defined spacing
            
            for line in lines:
                if line.strip():
                    # Apply different styles based on content
                    if ':' in line:
                        parts = line.split(':', 1)
                        key = parts[0].strip()
                        value = parts[1].strip()
                        
                        # Apply bold styling to headers
                        if key in ["TICKET NO", "COMPANY", "TRUCK PLATE", "PRODUCT", "GROSS WEIGHT", 
                                  "GROSS DATE", "GROSS TIME", "TARE WEIGHT", "TARE DATE", "TARE TIME", 
                                  "NET WEIGHT", "WEIGHT TYPE", "STATUS", "DATE PRINTED"]:
                            story.append(Paragraph(f"<b>{key}</b>: {value}", normal_style))
                        else:
                            story.append(Paragraph(line, normal_style))
                    else:
                        story.append(Paragraph(line, normal_style))
                else:
                    # For empty lines, add a spacer
                    story.append(Spacer(1, 0.2 * inch))  # Use smaller spacer for empty lines
                
                # Add consistent spacing after each line (including empty lines)
                story.append(Spacer(1, line_spacing_inch))
                
            # Build the PDF
            doc.build(story)
            
            # Print the PDF
            self._print_pdf_file(temp_pdf_path)
            
            # Clean up
            if os.path.exists(temp_pdf_path):
                os.unlink(temp_pdf_path)
                
        except Exception as e:
            self.msg_box.showerror("Print Error", f"Failed to generate or print PDF: {e}")
    
    def _prepare_transaction_data_for_printing(self, transaction_data):
        """Prepare transaction data for template formatting."""
        # Convert numeric values to strings with proper formatting
        formatted_data = {}
        for key, value in transaction_data.items():
            if isinstance(value, (int, float)):
                if key in ['gross_weight', 'tare_weight', 'net_weight']:
                    formatted_data[key] = f"{value:.2f}"
                else:
                    formatted_data[key] = str(value)
            else:
                formatted_data[key] = str(value) if value is not None else ""
        
        # Handle total_price based on the specific transaction's data
        existing_total_price = transaction_data.get('total_price')
        if existing_total_price is not None and str(existing_total_price).strip() not in ['0', '0.00', '']:
            # Transaction has a valid total price, format it properly
            try:
                total_price_val = float(existing_total_price)
                formatted_data['total_price'] = f"{total_price_val:.0f}"
            except (ValueError, TypeError):
                formatted_data['total_price'] = "0"
        else:
            # Transaction has no valid total price, remove it from formatted data
            # This prevents {total_price} from appearing in printouts
            if 'total_price' in formatted_data:
                del formatted_data['total_price']
        
        # Add additional fields
        formatted_data['date_printed'] = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
        return formatted_data

    def _print_pdf_file(self, pdf_path):
        """Print a PDF file using the selected printer."""
        try:
            printer_name = self.selected_printer_var.get()
            copies = self.print_copies_var.get()
            
            if sys.platform == "win32":
                # Windows-specific printing
                import win32print  # type: ignore
                import win32api  # type: ignore
                
                # Get default printer if none selected
                if not printer_name:
                    printer_name = win32print.GetDefaultPrinter()
                
                # Print the file
                win32api.ShellExecute(
                    0, "print", pdf_path, f'/d:"{printer_name}"', ".", 0
                )
            else:
                # Cross-platform solution using system print command
                if sys.platform.startswith('linux'):
                    subprocess.run(['lp', '-n', str(copies), pdf_path])
                elif sys.platform == 'darwin':  # macOS
                    subprocess.run(['lpr', '-n', str(copies), pdf_path])
                else:
                    # Fallback for other platforms
                    self.msg_box.showwarning("Print Warning", "Direct PDF printing not supported on this platform.")
                    
        except Exception as e:
            self.msg_box.showerror("Print Error", f"Failed to print PDF: {e}")

# Part 10: General Settings and Regex Handling

    def build_general_settings_sub_tab(self, tab: ttk.Frame):
        """
        Builds the General Settings tab for COM ports and hardware.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        baud_rates = [300, 600, 900, 1200, 2400, 4800, 9600, 19200, 38400, 57600, 115200]

        main_settings_frame = ttk.Frame(tab)
        main_settings_frame.pack(fill="x", expand=True, padx=5, pady=5)

        main_settings_frame.grid_columnconfigure(0, weight=1)
        main_settings_frame.grid_columnconfigure(1, weight=1)
        main_settings_frame.grid_columnconfigure(2, weight=1)

        emulator_frame = ttk.LabelFrame(main_settings_frame, text="Weight Emulator", padding="10")
        emulator_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5, ipady=5)

        ttk.Label(emulator_frame, text="Emulator COM Port:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self._emulator_port_entry = ttk.Combobox(emulator_frame, values=self.get_serial_ports(), textvariable=self._emulator_port_entry_var, style="TCombobox")
        self._emulator_port_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)

        ttk.Label(emulator_frame, text="Emulator Baud Rate:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self._emulator_baud_entry = ttk.Combobox(emulator_frame, values=baud_rates, textvariable=self._emulator_baud_entry_var, style="TCombobox")
        self._emulator_baud_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)

        ttk.Label(emulator_frame, text="Weight to Send:", style="TLabel").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(emulator_frame, textvariable=self._emulator_weight_to_send_var, style="TEntry").grid(row=2, column=1, sticky="ew", padx=5, pady=2)

        ttk.Label(emulator_frame, text="Send Interval (sec):", style="TLabel").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(emulator_frame, textvariable=self._emulator_send_interval_var, style="TEntry").grid(row=3, column=1, sticky="ew", padx=5, pady=2)
        
        self._emulator_auto_connect_var = tk.BooleanVar(value=self.config.get("emulator_auto_connect", False))
        ttk.Checkbutton(emulator_frame, text="Auto Connect", variable=self._emulator_auto_connect_var, style="TCheckbutton").grid(row=4, column=0, columnspan=2, sticky="w", padx=5, pady=2)
        
        emulator_button_frame = ttk.Frame(emulator_frame)
        emulator_button_frame.grid(row=5, column=0, columnspan=2, pady=5)
        self._emulator_connect_button = ttk.Button(emulator_button_frame, text="Connect", command=self._start_emulator_connection_from_settings, style="TButton")
        self._emulator_connect_button.pack(side=tk.LEFT, padx=2)
        self._emulator_disconnect_button = ttk.Button(emulator_button_frame, text="Disconnect", command=self._emulator_disconnect_serial, state=tk.DISABLED, style="TButton")
        self._emulator_disconnect_button.pack(side=tk.LEFT, padx=2)
        
        emulator_send_button_frame = ttk.Frame(emulator_frame)
        emulator_send_button_frame.grid(row=6, column=0, columnspan=2, pady=2)
        self._emulator_start_send_button = ttk.Button(emulator_send_button_frame, text="Start Sending", command=self._emulator_start_continuous_sending, state=tk.DISABLED, style="TButton")
        self._emulator_start_send_button.pack(side=tk.LEFT, padx=2)
        self._emulator_stop_send_button = ttk.Button(emulator_send_button_frame, text="Stop", command=self._emulator_stop_continuous_sending, state=tk.DISABLED, style="TButton")
        self._emulator_stop_send_button.pack(side=tk.LEFT, padx=2)

        ttk.Label(emulator_frame, textvariable=self._emulator_status_var, font=("Helvetica", 9, "italic")).grid(row=7, column=0, columnspan=2, sticky="w", padx=5, pady=(5,0))
        emulator_frame.grid_columnconfigure(1, weight=1)
        self._update_emulator_button_states()

        big_display_frame = ttk.LabelFrame(main_settings_frame, text="Big Display Output", padding="10")
        big_display_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5, ipady=5)

        ttk.Label(big_display_frame, text="Display COM Port:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.big_display_port_entry = ttk.Combobox(big_display_frame, values=self.get_serial_ports(), textvariable=self.big_display_port_var, style="TCombobox")
        self.big_display_port_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)

        ttk.Label(big_display_frame, text="Display Baud Rate:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.big_display_baud_entry = ttk.Combobox(big_display_frame, values=baud_rates, textvariable=self.big_display_baud_var, style="TCombobox")
        self.big_display_baud_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
        
        self.big_display_auto_connect_checkbox = ttk.Checkbutton(big_display_frame, text="Auto Connect", variable=self.big_display_auto_connect_var, style="TCheckbutton")
        self.big_display_auto_connect_checkbox.grid(row=3, column=0, columnspan=2, sticky="w", padx=5, pady=2)
        
        big_display_button_frame = ttk.Frame(big_display_frame)
        big_display_button_frame.grid(row=4, column=0, columnspan=2, pady=5)
        self.big_display_connect_button = ttk.Button(big_display_button_frame, text="Connect", command=self._connect_big_display, style="TButton")
        self.big_display_connect_button.pack(side=tk.LEFT, padx=2)
        self.big_display_disconnect_button = ttk.Button(big_display_button_frame, text="Disconnect", command=self._disconnect_big_display, state=tk.DISABLED, style="TButton")
        self.big_display_disconnect_button.pack(side=tk.LEFT, padx=2)

        ttk.Label(big_display_frame, textvariable=self.big_display_status_var, font=("Helvetica", 9, "italic")).grid(row=5, column=0, columnspan=2, sticky="w", padx=5, pady=(5,0))
        big_display_frame.grid_columnconfigure(1, weight=1)
        self._update_big_display_button_states()

        serial_frame = ttk.LabelFrame(main_settings_frame, text="Main Scale Connection", padding="10")
        serial_frame.grid(row=0, column=2, sticky="nsew", padx=5, pady=5, ipady=5)
        
        ttk.Label(serial_frame, text="COM Port:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.port_entry = ttk.Combobox(serial_frame, values=self.get_serial_ports(), textvariable=self.port_entry_var, style="TCombobox")
        self.port_entry.grid(row=0, column=1, columnspan=2, sticky="ew", padx=5, pady=2)

        ttk.Label(serial_frame, text="Baud Rate:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.baud_entry = ttk.Combobox(serial_frame, values=baud_rates, textvariable=self.baud_entry_var, style="TCombobox")
        self.baud_entry.grid(row=1, column=1, columnspan=2, sticky="ew", padx=5, pady=2)
        
        ttk.Label(serial_frame, text="Data Format Regex:", style="TLabel").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.data_format_regex_combobox = ttk.Combobox(serial_frame, values=self.predefined_regexes, textvariable=self.data_format_regex_var, style="TCombobox")
        self.data_format_regex_combobox.grid(row=2, column=1, sticky="ew", padx=5, pady=2)
        self.data_format_regex_combobox.bind("<<ComboboxSelected>>", self._on_regex_selection)
        ttk.Button(serial_frame, text="Auto-Detect", command=self._manual_detect_regex_action, style="TButton").grid(row=2, column=2, padx=(5, 0))

        
        ttk.Label(serial_frame, text="Display Decimals:", style="TLabel").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        self.decimal_places_var = tk.StringVar(value=str(self.config.get("decimal_places", 0)))
        self.decimal_places_spinbox = ttk.Spinbox(serial_frame, from_=0, to=3, textvariable=self.decimal_places_var, wrap=True, style="TSpinbox", width=5)
        self.decimal_places_spinbox.grid(row=3, column=1, columnspan=2, sticky="w", padx=5, pady=2)

        ttk.Label(serial_frame, text="Read Interval (ms):", style="TLabel").grid(row=4, column=0, sticky="w", padx=5, pady=2)
        self.read_loop_interval_spinbox = ttk.Spinbox(serial_frame, from_=10, to=1000, textvariable=self.read_loop_interval_ms, wrap=True, style="TSpinbox", width=5)
        self.read_loop_interval_spinbox.grid(row=4, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        
        self.auto_connect_var = tk.BooleanVar(value=self.config.get("auto_connect", True))
        ttk.Checkbutton(serial_frame, text="Auto Connect on Startup", variable=self.auto_connect_var, style="TCheckbutton").grid(row=5, column=0, columnspan=3, sticky="w", padx=5, pady=2)
        
        serial_button_frame = ttk.Frame(serial_frame)
        serial_button_frame.grid(row=6, column=0, columnspan=3, pady=5)
        self.connect_button = ttk.Button(serial_button_frame, text="Start Connection", command=self._start_main_serial_connection_from_settings, style="TButton")
        self.connect_button.pack(side=tk.LEFT, padx=2)
        self.stop_button = ttk.Button(serial_button_frame, text="Stop", command=self.stop_serial_connection, style="TButton")
        self.stop_button.pack(side=tk.LEFT, padx=2)

        if self.serial_running:
            self.connect_button.config(state=tk.DISABLED)
            self.stop_button.config(state=tk.NORMAL)
        else:
            self.connect_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
        serial_frame.grid_columnconfigure(1, weight=1)

        # Camera Settings Frame - Enhanced with better styling
        camera_frame = ttk.LabelFrame(main_settings_frame, text="📷 Camera Settings", padding="10")
        camera_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=5, pady=8, ipady=5)
        
        # Configure grid columns for side-by-side layout
        camera_frame.grid_columnconfigure(0, weight=1)
        camera_frame.grid_columnconfigure(1, weight=1)

        # Left Column - Device Selection & Options
        left_column = ttk.Frame(camera_frame)
        left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        # Camera Device Selection - Compact
        device_frame = ttk.LabelFrame(left_column, text="📹 Device Selection", padding="5")
        device_frame.pack(fill=tk.X, pady=(0, 8))
        device_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Label(device_frame, text="Camera:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.camera_device_combobox = ttk.Combobox(device_frame, textvariable=self.camera_device_var, style="TCombobox")
        self.camera_device_combobox.grid(row=0, column=1, sticky="ew", padx=(2, 2), pady=2)
        
        # Bind selection change to auto-save
        self.camera_device_combobox.bind('<<ComboboxSelected>>', lambda e: self.save_config())
        
        refresh_btn = ttk.Button(device_frame, text="🔄 Refresh Cameras", command=self._refresh_camera_list, style="TButton")
        refresh_btn.grid(row=0, column=2, padx=5, pady=2)
        
        # Options - Compact
        options_frame = ttk.LabelFrame(left_column, text="⚙️ Options", padding="8")
        options_frame.pack(fill=tk.X, pady=(0, 8))
        
        self.camera_auto_connect_var = tk.BooleanVar(value=self.config.get("camera_auto_connect", False))
        auto_connect_check = ttk.Checkbutton(options_frame, text="🚀 Auto Connect", 
                                            variable=self.camera_auto_connect_var, 
                                            command=lambda: self.save_config(),
                                            style="TCheckbutton")
        auto_connect_check.pack(anchor="w", padx=5, pady=2)
        
        self.camera_use_image_check = ttk.Checkbutton(options_frame, text="🖼️ Use Image", 
                                                      variable=self.camera_use_image_var, 
                                                      command=self._on_camera_image_toggle,
                                                      style="TCheckbutton")
        self.camera_use_image_check.pack(anchor="w", padx=5, pady=2)
        
        self.camera_mirror_check = ttk.Checkbutton(options_frame, text="🪞 Mirror Mode", 
                                                  variable=self.camera_mirror_mode_var,
                                                  command=self._on_camera_mirror_toggle,
                                                  style="TCheckbutton")
        self.camera_mirror_check.pack(anchor="w", padx=5, pady=2)

        # Right Column - Image Upload & Controls
        right_column = ttk.Frame(camera_frame)
        right_column.grid(row=0, column=1, sticky="nsew", padx=(5, 0), rowspan=2)
        
        # Image Upload Frame - Compact
        self.camera_image_upload_frame = ttk.LabelFrame(right_column, text="📁 Image Upload", padding="8")
        self.camera_image_upload_frame.pack(fill=tk.X, pady=(0, 8))
        self.camera_image_upload_frame.grid_columnconfigure(0, weight=1)
        
        ttk.Label(self.camera_image_upload_frame, text="File:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        self.camera_image_path_var = tk.StringVar(value=self.camera_uploaded_image_path)
        self.camera_image_path_entry = ttk.Entry(self.camera_image_upload_frame, textvariable=self.camera_image_path_var, style="TEntry")
        self.camera_image_path_entry.grid(row=1, column=0, sticky="ew", padx=5, pady=2)
        
        browse_btn = ttk.Button(self.camera_image_upload_frame, text="📂 Browse", command=self._browse_camera_image, style="TButton")
        browse_btn.grid(row=2, column=0, sticky="ew", padx=5, pady=3)
        
        # Camera Control Frame - Compact
        control_frame = ttk.LabelFrame(right_column, text="🎮 Control", padding="8")
        control_frame.pack(fill=tk.X, pady=(0, 8))
        
        self.camera_connect_button = ttk.Button(control_frame, text="📷 Connect", 
                                              command=self._toggle_camera_connection, 
                                              style="TButton")
        self.camera_connect_button.pack(fill=tk.X, padx=5, pady=3)
        
        # Bottom row - Save Settings button and Camera Status side by side
        bottom_row_frame = ttk.Frame(camera_frame)
        bottom_row_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(5, 0))
        bottom_row_frame.grid_columnconfigure(0, weight=1)
        bottom_row_frame.grid_columnconfigure(1, weight=1)
        
        # Save Settings Button - Left side
        save_settings_btn = ttk.Button(bottom_row_frame, text="💾 Save Settings", 
                                      command=self.save_config, 
                                      style="TButton")
        save_settings_btn.grid(row=0, column=0, sticky="ew", padx=(5, 2), pady=5)
        
        # Camera Status - Right side
        camera_status_frame = ttk.LabelFrame(bottom_row_frame, text="📊 Camera Status", padding="5")
        camera_status_frame.grid(row=0, column=1, sticky="ew", padx=(2, 5), pady=5)
        camera_status_frame.grid_columnconfigure(0, weight=1)
        
        ttk.Label(camera_status_frame, textvariable=self.camera_status_var, 
                 foreground="blue", 
                 style="TLabel").pack(anchor="w", padx=5, pady=2)
        
        # Show/hide upload frame based on toggle state
        self._on_camera_image_toggle()
        
        # Configure grid weights for the upload frame
        self.camera_image_upload_frame.grid_columnconfigure(0, weight=1)

        # ttk.Label(serial_frame, text="Predefined regex patterns for data format:", font=("Helvetica", 9, "bold")).grid(row=7, column=0, columnspan=3, sticky="w", padx=5, pady=(10, 2))
        # ttk.Label(serial_frame, text="    (\\d+\\.\\d+)", font=("Courier New", 9)).grid(row=8, column=0, columnspan=3, sticky="w", padx=5)
        # ttk.Label(serial_frame, text="    W=(\\d+\\.\\d+)", font=("Courier New", 9)).grid(row=9, column=0, columnspan=3, sticky="w", padx=5)
        # ttk.Label(serial_frame, text="    ST,GS,(\\d+\\.\\d+)", font=("Courier New", 9)).grid(row=10, column=0, columnspan=3, sticky="w", padx=5)
        # ttk.Label(serial_frame, text="    ([\\+\\-]?\\d+\\.\\d+)", font=("Courier New", 9)).grid(row=11, column=0, columnspan=3, sticky="w", padx=5)
        # ttk.Label(serial_frame, text="    (\\d+)\\s+kg", font=("Courier New", 9)).grid(row=12, column=0, columnspan=3, sticky="w", padx=5)
        # ttk.Label(serial_frame, text="    (\\d+)", font=("Courier New", 9)).grid(row=13, column=0, columnspan=3, sticky="w", padx=5)

    def _on_regex_selection(self, event):
        """
        Handles the selection of a predefined regex or the 'Custom' option.
        """
        selected_regex = self.data_format_regex_var.get()
        if selected_regex == "Custom":
            self.data_format_regex_combobox.config(state=tk.DISABLED)
            try:
                custom_value = simpledialog.askstring("Custom Regex", "Enter your custom regular expression:", parent=self.root)
                if custom_value:
                    try:
                        re.compile(custom_value)
                        self.data_format_regex_var.set(custom_value)
                    except re.error as e:
                        self.msg_box.showerror("Invalid Regex", f"The provided regex is invalid: {e}")
                        self.data_format_regex_var.set(self.config.get("data_format_regex", r"([\+\-]?\d+\.\d+)"))
                else:
                    self.data_format_regex_var.set(self.config.get("data_format_regex", r"([\+\-]?\d+\.\d+)"))
            finally:
                self.data_format_regex_combobox.config(state=tk.NORMAL)
        else:
            # Update the regex variable with the selected pattern
            self.data_format_regex_var.set(selected_regex)



    def _update_print_preview(self):
        """Update the preview with sample data using coordinate positioning."""
        try:
            # Sample data for preview
            date_printed_value = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
            sample_data = {
                "company": "SAMPLE CO.",
                "ticket_no": "12345",
                "truck_plate": "ABC 123",
                "product": "GRAIN",
                "gross_weight": "25000.00",
                "gross_date": "2023-10-27",
                "gross_time": "10:30 AM",
                "date_printed": date_printed_value,
                "tare_weight": "15000.00",
                "tare_date": "2023-10-27",
                "tare_time": "10:30 AM",
                "net_weight": "10000.00",
                "weight_type": "ONE WAY WEIGHING",
                "status": "COMPLETED",
                "logged_in_user": self.logged_in_user if self.logged_in_user else "admin"
            }
            
            # Generate preview using coordinate-based approach
            preview_content = self._generate_coordinate_based_label(sample_data)
            # Update preview text widget
            self.print_preview_text.config(state="normal")
            self.print_preview_text.delete("1.0", tk.END)
            self.print_preview_text.insert("1.0", preview_content)
            
        except Exception as e:
            self.print_preview_text.config(state="normal")
            self.print_preview_text.delete("1.0", tk.END)
            self.print_preview_text.insert("1.0", f"Preview Error: {str(e)}")
            self.print_preview_text.config(state="disabled")

    def build_print_settings_tab(self, tab: ttk.Frame):
        """Builds the GUI for the Print Settings tab with advanced features."""
        for widget in tab.winfo_children():
            widget.destroy()

        # --- Printer Settings ---
        printer_frame = ttk.LabelFrame(tab, text="Printer Settings", padding="10")
        printer_frame.pack(fill="x", padx=10, pady=(5, 0))

        ttk.Label(printer_frame, text="Printer:").grid(row=0, column=0, sticky="w", padx=2)
        self.selected_printer_var = tk.StringVar()
        printer_combo = ttk.Combobox(printer_frame, textvariable=self.selected_printer_var, state="readonly")
        printer_combo.grid(row=0, column=1, sticky="ew", padx=2, pady=2)
        printer_combo.bind("<<ComboboxSelected>>", lambda e: self._on_printer_selected())
        ttk.Button(printer_frame, text="Refresh", command=self._populate_printer_list).grid(row=0, column=2, padx=(5, 0), pady=2)

        ttk.Label(printer_frame, text="Number of Copies:").grid(row=1, column=0, sticky="w", padx=2)
        self.print_copies_var = tk.IntVar(value=1)
        copies_spinbox = ttk.Spinbox(printer_frame, from_=1, to=99, textvariable=self.print_copies_var, width=5)
        copies_spinbox.grid(row=1, column=1, sticky="w", padx=2)

        # --- Printout Format Settings ---
        format_frame = ttk.LabelFrame(tab, text="Printout Format Settings", padding="10")
        format_frame.pack(fill="x", padx=10, pady=(5, 0))

        # Add labels and text boxes for each printout detail
        ttk.Label(format_frame, text="Company Name:").grid(row=0, column=0, sticky="w", padx=2)
        self.company_name_var = tk.StringVar()
        company_name_entry = ttk.Entry(format_frame, textvariable=self.company_name_var)
        company_name_entry.grid(row=0, column=1, sticky="ew", padx=2, pady=2)

        ttk.Label(format_frame, text="Truck Plate:").grid(row=1, column=0, sticky="w", padx=2)
        self.truck_plate_var = tk.StringVar()
        truck_plate_entry = ttk.Entry(format_frame, textvariable=self.truck_plate_var)
        truck_plate_entry.grid(row=1, column=1, sticky="ew", padx=2, pady=2)

        ttk.Label(format_frame, text="Product:").grid(row=2, column=0, sticky="w", padx=2)
        self.product_var = tk.StringVar()
        product_entry = ttk.Entry(format_frame, textvariable=self.product_var)
        product_entry.grid(row=2, column=1, sticky="ew", padx=2, pady=2)

        ttk.Label(format_frame, text="Gross Weight:").grid(row=3, column=0, sticky="w", padx=2)
        self.gross_weight_var = tk.StringVar()
        gross_weight_entry = ttk.Entry(format_frame, textvariable=self.gross_weight_var)
        gross_weight_entry.grid(row=3, column=1, sticky="ew", padx=2, pady=2)

        ttk.Label(format_frame, text="Date Format:").grid(row=4, column=0, sticky="w", padx=2)
        self.date_format_var = tk.StringVar()
        date_format_entry = ttk.Entry(format_frame, textvariable=self.date_format_var)
        date_format_entry.grid(row=4, column=1, sticky="ew", padx=2, pady=2)

        ttk.Label(format_frame, text="Time Format:").grid(row=5, column=0, sticky="w", padx=2)
        self.time_format_var = tk.StringVar()
        time_format_entry = ttk.Entry(format_frame, textvariable=self.time_format_var)
        time_format_entry.grid(row=5, column=1, sticky="ew", padx=2, pady=2)

        # --- Save/Load Buttons ---
        save_button = ttk.Button(format_frame, text="Save", command=self._save_printout_format)
        save_button.grid(row=6, column=0, sticky="w", padx=2, pady=5)

        load_button = ttk.Button(format_frame, text="Load", command=self._load_printout_format)
        load_button.grid(row=6, column=1, sticky="e", padx=2, pady=5)

        # --- PDF Settings ---
        pdf_frame = ttk.LabelFrame(tab, text="PDF Settings", padding="10")
        pdf_frame.pack(fill="x", padx=10, pady=(5, 0))

        ttk.Label(pdf_frame, text="Page Size:").grid(row=0, column=0, sticky="w", padx=2)
        self.pdf_page_size_var = tk.StringVar(value="A6")
        page_size_combo = ttk.Combobox(pdf_frame, textvariable=self.pdf_page_size_var, values=["A6"], state="readonly")
        page_size_combo.grid(row=0, column=1, sticky="ew", padx=2, pady=2)

        ttk.Label(pdf_frame, text="Orientation:").grid(row=1, column=0, sticky="w", padx=2)
        self.pdf_orientation_var = tk.StringVar(value="portrait")
        orientation_combo = ttk.Combobox(pdf_frame, textvariable=self.pdf_orientation_var, values=["portrait", "landscape"], state="readonly")
        orientation_combo.grid(row=1, column=1, sticky="ew", padx=2, pady=2)

        # --- Template Selection & Editor ---
        template_frame = ttk.LabelFrame(tab, text="Print Templates", padding="10")
        template_frame.pack(fill="x", padx=10, pady=(5, 0))

        # Radio buttons for template selection
        self.current_template_editor_mode_var = tk.StringVar(value="Two Way Template")
        ttk.Radiobutton(template_frame, text="One Way Template", variable=self.current_template_editor_mode_var, value="One Way Template").grid(row=0, column=0, sticky="w", padx=2, pady=2)
        ttk.Radiobutton(template_frame, text="Two Way Template", variable=self.current_template_editor_mode_var, value="Two Way Template").grid(row=0, column=1, sticky="w", padx=2, pady=2)

        # Placeholder List and Insert Button
        placeholder_frame = ttk.Frame(template_frame)
        placeholder_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=2, pady=2)
        placeholder_frame.columnconfigure(0, weight=1)

        ttk.Label(placeholder_frame, text="Available Placeholders:").grid(row=0, column=0, sticky="w", padx=2)
        self.placeholder_listbox = tk.Listbox(placeholder_frame, height=8)
        self.placeholder_listbox.grid(row=1, column=0, sticky="ew", padx=2, pady=2)
        self.placeholder_listbox.bind("<Double-Button-1>", self._on_placeholder_double_click)
        self._populate_placeholder_list()

        ttk.Button(placeholder_frame, text="Insert Selected", command=self._insert_placeholder).grid(row=1, column=1, padx=(5, 0), pady=2)

        # Enhanced Text Editor with Syntax Highlighting
        editor_frame = ttk.Frame(template_frame)
        editor_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=2, pady=2)
        editor_frame.columnconfigure(0, weight=1)
        editor_frame.rowconfigure(0, weight=1)

        # Create a custom Text widget for syntax highlighting
        self.template_text_widget = tk.Text(editor_frame, wrap="word", font=("Courier New", 10))
        self.template_text_widget.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        self.template_text_widget.tag_configure("placeholder", foreground="blue")
        self.template_text_widget.tag_configure("keyword", foreground="green")
        self.template_text_widget.tag_configure("default", foreground="black")

        # Add scrollbars
        scrollbar_y = ttk.Scrollbar(editor_frame, orient="vertical", command=self.template_text_widget.yview)
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        self.template_text_widget.configure(yscrollcommand=scrollbar_y.set)

        # Bind events for real-time updates
        self.template_text_widget.bind("<KeyRelease>", self._on_template_text_change)
        self.template_text_widget.bind("<ButtonRelease-1>", self._highlight_syntax)

        # Initialize the editor with default templates
        self._load_default_templates()

        # --- Preview Panel ---
        preview_frame = ttk.LabelFrame(tab, text="Preview", padding="10")
        preview_frame.pack(fill="both", expand=True, padx=10, pady=(5, 0))

        # Create a frame for the preview content
        preview_content_frame = ttk.Frame(preview_frame)
        preview_content_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # Use a Canvas to hold the generated PDF image or display error message
        self.preview_canvas = tk.Canvas(preview_content_frame, bg="white", relief="sunken", bd=2)
        self.preview_canvas.pack(fill="both", expand=True)

        # --- Controls below the preview ---
        control_frame = ttk.Frame(preview_frame)
        control_frame.pack(fill="x", padx=2, pady=(2, 0))

        ttk.Button(control_frame, text="Generate Preview", command=self._generate_pdf_preview).pack(side="left", padx=(0, 5))
        ttk.Button(control_frame, text="Reset Template", command=self._reset_template_to_default).pack(side="right")

        # Configure grid weights for proper resizing
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)  # Make the preview area expandable

    def _setup_advanced_print_controls(self, tab):
        """Setup advanced print controls including margins and header/footer."""
        # --- Advanced Print Settings ---
        advanced_frame = ttk.LabelFrame(tab, text="Advanced Print Settings", padding="10")
        advanced_frame.pack(fill="x", padx=10, pady=(5, 0))

        # Margins
        margin_frame = ttk.Frame(advanced_frame)
        margin_frame.pack(fill="x", pady=5)

        ttk.Label(margin_frame, text="Margins (inches):").grid(row=0, column=0, sticky="w", padx=2)
        ttk.Label(margin_frame, text="Top:").grid(row=1, column=0, sticky="w", padx=2)
        self.margin_top_var = tk.DoubleVar(value=0.5)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.1, textvariable=self.margin_top_var, width=5).grid(row=1, column=1, sticky="w", padx=2)

        ttk.Label(margin_frame, text="Bottom:").grid(row=2, column=0, sticky="w", padx=2)
        self.margin_bottom_var = tk.DoubleVar(value=0.5)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.1, textvariable=self.margin_bottom_var, width=5).grid(row=2, column=1, sticky="w", padx=2)

        ttk.Label(margin_frame, text="Left:").grid(row=1, column=2, sticky="w", padx=2)
        self.margin_left_var = tk.DoubleVar(value=0.5)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.1, textvariable=self.margin_left_var, width=5).grid(row=1, column=3, sticky="w", padx=2)

        ttk.Label(margin_frame, text="Right:").grid(row=2, column=2, sticky="w", padx=2)
        self.margin_right_var = tk.DoubleVar(value=0.5)
        ttk.Spinbox(margin_frame, from_=0, to=2, increment=0.1, textvariable=self.margin_right_var, width=5).grid(row=2, column=3, sticky="w", padx=2)

        # Header/Footer
        header_footer_frame = ttk.Frame(advanced_frame)
        header_footer_frame.pack(fill="x", pady=5)

        ttk.Label(header_footer_frame, text="Header Text:").grid(row=0, column=0, sticky="w", padx=2)
        self.header_text_var = tk.StringVar(value="")
        ttk.Entry(header_footer_frame, textvariable=self.header_text_var, width=50).grid(row=1, column=0, columnspan=2, sticky="ew", padx=2, pady=2)

        ttk.Label(header_footer_frame, text="Footer Text:").grid(row=2, column=0, sticky="w", padx=2)
        self.footer_text_var = tk.StringVar(value="")
        ttk.Entry(header_footer_frame, textvariable=self.footer_text_var, width=50).grid(row=3, column=0, columnspan=2, sticky="ew", padx=2, pady=2)

        # Font Management
        font_frame = ttk.LabelFrame(advanced_frame, text="Font Management", padding="10")
        font_frame.pack(fill="x", pady=5)

        ttk.Label(font_frame, text="Custom Font File:").grid(row=0, column=0, sticky="w", padx=2)
        self.custom_font_path_var = tk.StringVar()
        ttk.Entry(font_frame, textvariable=self.custom_font_path_var, width=40).grid(row=1, column=0, sticky="ew", padx=2, pady=2)
        ttk.Button(font_frame, text="Browse...", command=self._browse_font_file).grid(row=1, column=1, padx=2)

        # Line Spacing
        ttk.Label(advanced_frame, text="Line Spacing (lines):").pack(anchor="w", padx=2, pady=(5, 0))
        self.print_line_spacing = tk.DoubleVar(value=1.0)
        ttk.Spinbox(advanced_frame, from_=0.5, to=3.0, increment=0.1, textvariable=self.print_line_spacing, width=5).pack(anchor="w", padx=2)

        # Template Management
        template_management_frame = ttk.Frame(advanced_frame)
        template_management_frame.pack(fill="x", pady=5)

        ttk.Button(template_management_frame, text="Save Current Template", command=self._save_current_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_management_frame, text="Load Template", command=self._load_template_dialog).pack(side=tk.LEFT, padx=2)
        ttk.Button(template_management_frame, text="Export Template", command=self._export_template).pack(side=tk.LEFT, padx=2)

    def _browse_font_file(self):
        """Open a file dialog to select a custom font file."""
        file_path = filedialog.askopenfilename(
            title="Select Font File",
            filetypes=[("Font files", "*.ttf *.otf"), ("All files", "*.*")]
        )
        if file_path:
            self.custom_font_path_var.set(file_path)
            self._register_custom_font(file_path)

    def _register_custom_font(self, font_path):
        """Register a custom font with the system."""
        try:
            # This would integrate with reportlab's font system
            # For now, we'll just store the path
            self.custom_font_path = font_path
            self.msg_box.showinfo("Font Loaded", f"Custom font loaded: {os.path.basename(font_path)}")
        except Exception as e:
            self.msg_box.showerror("Font Error", f"Failed to load font: {e}")

    def _load_template_dialog(self):
        """Open a dialog to load a saved template."""
        # This would implement a template selection dialog
        # For now, we'll just show a message
        self.msg_box.showinfo("Template Loading", "Template loading feature would be implemented here.")

    def _export_template(self):
        """Export the current template to a file."""
        template_content = self.template_text_widget.get("1.0", tk.END)
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    f.write(template_content)
                self.msg_box.showinfo("Export Complete", "Template exported successfully.")
            except Exception as e:
                self.msg_box.showerror("Export Error", f"Failed to export template: {e}")

    
    def _populate_printer_list(self):
        """Populate the printer dropdown list and default selection."""
        try:
            if sys.platform == "win32" and PRINTING_ENABLED:
                printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
                printer_names = [printer[2] for printer in printers]
                # Try to use Windows default if available
                try:
                    default_name = win32print.GetDefaultPrinter()
                except Exception:
                    default_name = None
            else:
                printer_names = ["Generic Printer"]
                default_name = printer_names[0]

            # Store list
            self.available_printers = printer_names

            # Update combobox if present
            if hasattr(self, 'printer_combobox') and isinstance(self.printer_combobox, ttk.Combobox):
                self.printer_combobox['values'] = printer_names

            current = self.selected_printer_var.get()
            if current and current in printer_names:
                # keep current selection
                pass
            else:
                # prefer default if available, otherwise first in list
                if default_name and default_name in printer_names:
                    self.selected_printer_var.set(default_name)
                elif printer_names:
                    self.selected_printer_var.set(printer_names[0])
                else:
                    self.selected_printer_var.set("")
        except Exception as e:
            print(f"Error populating printer list: {e}")

    def _on_printer_selected(self):
        """Handle printer selection change."""
        # This could be used to validate the selected printer or update settings.
        pass

    def _populate_placeholder_list(self):
        """Populate the list of available placeholders."""
        placeholders = [
            "ticket_no", "company", "truck_plate", "product",
            "designation", "sender", "origin", "destination", "driver",
            "gross_weight", "gross_date", "gross_time", "tare_weight",
            "tare_date", "tare_time", "net_weight", "weight_type", "status",
            "date_printed", "barcode"
        ]
        self.placeholder_listbox.delete(0, tk.END)
        for p in placeholders:
            self.placeholder_listbox.insert(tk.END, f"{{{p}}}")

    def _on_placeholder_double_click(self, event=None):
        """Handle double-click on a placeholder in the list."""
        selection = self.placeholder_listbox.curselection()
        if selection:
            item = self.placeholder_listbox.get(selection[0])
            # Insert the placeholder at the cursor position
            cursor_pos = self.template_text_widget.index(tk.INSERT)
            self.template_text_widget.insert(cursor_pos, item)

    def _insert_placeholder(self):
        """Insert the currently selected placeholder into the editor."""
        selection = self.placeholder_listbox.curselection()
        if selection:
            item = self.placeholder_listbox.get(selection[0])
            # Insert the placeholder at the cursor position
            cursor_pos = self.template_text_widget.index(tk.INSERT)
            self.template_text_widget.insert(cursor_pos, item)

    def _load_default_templates(self):
        """Load the default templates based on the selected mode."""
        if self.current_template_editor_mode_var.get() == "One Way Template":
            self.template_text_widget.delete("1.0", tk.END)
            self.template_text_widget.insert("1.0", self._get_one_way_template())
        else:
            self.template_text_widget.delete("1.0", tk.END)
            self.template_text_widget.insert("1.0", self._get_two_way_template())

    def _get_one_way_template(self):
        """Return the default one-way template string."""
        return """Advantechnique
advantechnique@gmail.com

TICKET NO : {ticket_no}

COMPANY     : {company}
TRUCK PLATE : {truck_plate}
PRODUCT     : {product}

GROSS WEIGHT : {gross_weight} KG
GROSS DATE   : {gross_date}
GROSS TIME   : {gross_time}
DATE PRINTED : {date_printed}
"""

    def _get_two_way_template(self):
        """Return the default two-way template string."""
        return """Advantechnique
advantechnique@gmail.com

TICKET NO : {ticket_no}

COMPANY         : {company}
TRUCK PLATE     : {truck_plate}
PRODUCT         : {product}
DESIGNATION     : {designation}
SENDER          : {sender}
ORIGIN          : {origin}
DESTINATION     : {destination}
DRIVER          : {driver}

GROSS WEIGHT    : {gross_weight} KG
GROSS DATE      : {gross_date}
GROSS TIME      : {gross_time}

TARE WEIGHT     : {tare_weight} KG
TARE DATE       : {tare_date}
TARE TIME       : {tare_time}

NET WEIGHT      : {net_weight} KG
WEIGHT TYPE     : {weight_type}
STATUS          : {status}
DATE PRINTED    : {date_printed}
{barcode}
"""

    def _highlight_syntax(self, event=None):
        """Apply syntax highlighting to the template text."""
        # Remove existing tags
        self.template_text_widget.tag_remove("placeholder", "1.0", tk.END)
        self.template_text_widget.tag_remove("keyword", "1.0", tk.END)
        
        # Find and highlight placeholders
        start = "1.0"
        while True:
            pos = self.template_text_widget.search(r"\{[^}]*\}", start, tk.END, regexp=True)
            if not pos:
                break
            end = f"{pos}+{len(pos)-1}c"
            self.template_text_widget.tag_add("placeholder", pos, end)
            start = end

        # Find and highlight keywords (like "TICKET NO", "COMPANY", etc.)
        keywords = ["TICKET NO", "COMPANY", "TRUCK PLATE", "PRODUCT", "GROSS WEIGHT", "GROSS DATE", "GROSS TIME", "TARE WEIGHT", "TARE DATE", "TARE TIME", "NET WEIGHT", "WEIGHT TYPE", "STATUS", "DATE PRINTED"]
        for keyword in keywords:
            start = "1.0"
            while True:
                pos = self.template_text_widget.search(keyword, start, tk.END)
                if not pos:
                    break
                end = f"{pos}+{len(keyword)}c"
                self.template_text_widget.tag_add("keyword", pos, end)
                start = end

        # Apply default styling to everything else
        self.template_text_widget.tag_add("default", "1.0", tk.END)

    def _on_template_text_change(self, event=None):
        """Handle changes to the template text (debounced)."""
        # Debounce logic to prevent excessive calls
        self._debounce_template_updates()

    def _debounce_template_updates(self):
        """Debounce function to delay the actual update until user stops typing."""
        # This method would call _update_all_previews after a short delay.
        # Implementation depends on your debouncing strategy.
        pass

    def _generate_pdf_preview(self):
        """Generate a PDF preview using reportlab and display it in the canvas."""
        # Check if required widgets exist before trying to use them
        if not hasattr(self, 'template_text_widget') or not self.template_text_widget:
            return
        if not hasattr(self, 'preview_canvas') or not self.preview_canvas:
            return
            
        try:
            # Get the current template text
            template_str = self.template_text_widget.get("1.0", tk.END)
            
            # Create sample data for the preview
            sample_data = {
                "ticket_no": "12345",
                "company": "ABC Corp",
                "truck_plate": "XYZ123",
                "product": "Gravel",
                "designation": "Construction",
                "sender": "Supplier Inc.",
                "origin": "Manila",
                "destination": "Quezon City",
                "driver": "John Doe",
                "gross_weight": "15000.00",
                "gross_date": "2023-10-27",
                "gross_time": "10:30 AM",
                "tare_weight": "5000.00",
                "tare_date": "2023-10-27",
                "tare_time": "10:45 AM",
                "net_weight": "10000.00",
                "weight_type": "TWO WAY WEIGHING",
                "status": "Completed",
                "date_printed": datetime.now().strftime('%m/%d/%Y %I:%M:%S %p'),
                "barcode": "12345",
                "logged_in_user": self.logged_in_user if self.logged_in_user else "admin"
            }
            
            # Create a temporary file path for the PDF
            temp_pdf_path = tempfile.mktemp(suffix=".pdf")
            
            # Determine page size and orientation
            page_size = letter
            if self.pdf_page_size_var.get() == "A6":
                page_size = (4.13 * inch, 5.5 * inch)
            elif self.pdf_page_size_var.get() == "A6":
                page_size = (4.13 * inch, 5.5 * inch)
                
            if self.pdf_orientation_var.get() == "landscape":
                page_size = (page_size[1], page_size[0])
                
            # Create a SimpleDocTemplate for the PDF
            doc = SimpleDocTemplate(temp_pdf_path, pagesize=page_size, 
                                   rightMargin=0.1 * inch, leftMargin=0.1 * inch, 
                                   topMargin=0.1 * inch, bottomMargin=0.1 * inch)
            story = []
            
            # Define styles
            styles = getSampleStyleSheet()
            normal_style = styles["Normal"]
            bold_style = styles["Heading3"]
            
            # Format the template text with the sample data
            formatted_content = template_str.format(**sample_data)
            
            # - Process lines, handling barcodes -
            line_spacing_inch = self.print_line_spacing.get() * 0.05 * inch
            lines = formatted_content.split('\n')

            for line in lines:
                if line.strip():
                    # Split the line into key-value pairs if it contains a colon
                    if ':' in line:
                        parts = line.split(':', 1)
                        key = parts[0].strip()
                        value = parts[1].strip()

                        # Apply different styles based on key
                        # Use the standard 'bold' style from the reportlab stylesheet
                        story.append(Paragraph(f"<b>{key}</b>: {value}", bold_style))
                    else:
                        # For lines without a colon, use the normal style
                        story.append(Paragraph(line, normal_style))
                else:
                    # Add space for empty lines
                    story.append(Spacer(1, 12))
                    
            # Build the PDF
            doc.build(story)
            
            # Display success message in preview canvas
            self.preview_canvas.delete("all")
            self.preview_canvas.create_text(10, 10, anchor="nw", 
                                           text="Preview Generated Successfully!", 
                                           fill="green", font=("Arial", 10))
            
        except Exception as e:
            # Handle any errors during PDF generation
            self.preview_canvas.delete("all")
            self.preview_canvas.create_text(10, 10, anchor="nw", 
                                           text=f"Preview Error: {str(e)}", 
                                           fill="red", font=("Arial", 10))
        finally:
            # Clean up the temporary file if it exists
            if 'temp_pdf_path' in locals() and os.path.exists(temp_pdf_path):
                try:
                    os.unlink(temp_pdf_path)
                except:
                    pass

    def _display_pdf_image(self, pdf_path):
        """Display the PDF as an image in the preview canvas."""
        # This is a simplified example. In practice, you'd need to convert the PDF to an image.
        # Libraries like `pdf2image` can be used for this purpose.
        # Here, we'll just clear the canvas and show a placeholder message.
        self.preview_canvas.delete("all")
        self.preview_canvas.create_text(10, 10, anchor="nw", text="PDF Preview Image Would Appear Here", fill="gray")

    def _reset_template_to_default(self):
        """Reset the template to its default state."""
        self.template_text_widget.delete("1.0", tk.END)
        self.template_text_widget.insert("1.0", self._get_default_template())

    def _get_default_template(self):
        """Get the default template based on the current mode."""
        if self.current_template_editor_mode_var.get() == "One Way Template":
            return self._get_one_way_template()
        else:
            return self._get_two_way_template()


    def _update_all_previews(self):
        """Update all previews (e.g., PDF preview) based on current settings."""
        # Only update previews if the preview widgets exist (they're in the Print Templates tab)
        if hasattr(self, 'print_preview_text') and self.print_preview_text:
            try:
                self._generate_pdf_preview()
            except Exception as e:
                logging.debug(f"Could not update print preview: {e}")
        if hasattr(self, 'preview_canvas') and self.preview_canvas:
            try:
                self._generate_pdf_preview()
            except Exception as e:
                logging.debug(f"Could not update PDF canvas preview: {e}")
        
    def _on_template_text_change(self, event=None):
        """Handle changes to the template text with debouncing."""
        # Cancel any pending updates
        if hasattr(self, '_template_update_timer'):
            self.root.after_cancel(self._template_update_timer)
        
        # Schedule a new update after 500ms of inactivity
        self._template_update_timer = self.root.after(500, self._debounced_template_update)

    def _debounced_template_update(self):
        """Actually update the template content."""
        self._highlight_syntax()
        self._update_all_previews()

    def build_activation_tab(self, tab: ttk.Frame):
        """
        Builds the GUI for the application activation, admin password change, and user management.
        Only accessible by admin users.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        # Create main frame with expanded layout for user management
        main_frame = ttk.Frame(tab)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Configure grid columns for distribution (now only 2 columns since admin frame is hidden)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)

        # ===== Application Activation Section =====
        activation_frame = ttk.LabelFrame(main_frame, text="Application Activation", padding="10")
        activation_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5, ipady=5)
        activation_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Label(activation_frame, text="Activation Status:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(activation_frame, textvariable=self.activation_status_var, font=("Helvetica", 10, "italic"), foreground="blue").grid(row=0, column=1, sticky="ew", padx=5, pady=2)

        ttk.Label(activation_frame, text="Enter Activation Code:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.activation_code_entry = ttk.Entry(activation_frame, textvariable=self.activation_code_entry_var, show="*", style="TEntry")
        self.activation_code_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)

        activation_button_frame = ttk.Frame(activation_frame)
        activation_button_frame.grid(row=2, column=0, columnspan=2, pady=5)
        ttk.Button(activation_button_frame, text="Activate", command=self._attempt_activation, style="TButton").pack(side=tk.LEFT, padx=2)

        # ===== Change Admin Password Section =====
        # HIDDEN - This section is now hidden from the GUI
        # admin_frame = ttk.LabelFrame(main_frame, text="Change Admin Password", padding="10")
        # admin_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5, ipady=5)
        # admin_frame.grid_columnconfigure(1, weight=1)
        # 
        # ttk.Label(admin_frame, text="Current Password:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        # self.current_admin_password_entry = ttk.Entry(admin_frame, show="*", style="TEntry")
        # self.current_admin_password_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        # 
        # ttk.Label(admin_frame, text="New Password:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        # self.new_admin_password_entry = ttk.Entry(admin_frame, textvariable=self.new_admin_password_var, show="*", style="TEntry")
        # self.new_admin_password_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
        # 
        # ttk.Label(admin_frame, text="Confirm New Password:", style="TLabel").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        # self.new_admin_password_confirm_entry = ttk.Entry(admin_frame, textvariable=self.new_admin_password_confirm_var, show="*", style="TEntry")
        # self.new_admin_password_confirm_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=2)
        # 
        # admin_button_frame = ttk.Frame(admin_frame)
        # admin_button_frame.grid(row=3, column=0, columnspan=2, pady=5)
        # ttk.Button(admin_button_frame, text="Update Password", command=self._update_admin_password, style="TButton").pack(side=tk.LEFT, padx=2)

        # ===== Create New User Section =====
        user_frame = ttk.LabelFrame(main_frame, text="Create New User", padding="10")
        user_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5, ipady=5)
        user_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Label(user_frame, text="Username:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.new_user_username_entry = ttk.Entry(user_frame, style="TEntry")
        self.new_user_username_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        
        ttk.Label(user_frame, text="Password:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.new_user_password_entry = ttk.Entry(user_frame, show="*", style="TEntry")
        self.new_user_password_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=2)
        
        ttk.Label(user_frame, text="Confirm Password:", style="TLabel").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.new_user_password_confirm_entry = ttk.Entry(user_frame, show="*", style="TEntry")
        self.new_user_password_confirm_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=2)
        
        ttk.Label(user_frame, text="User Role:", style="TLabel").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        self.new_user_role_var = tk.StringVar(value="user")
        role_dropdown = ttk.Combobox(user_frame, textvariable=self.new_user_role_var, values=["user", "operator", "admin"], state="readonly", style="TCombobox")
        role_dropdown.grid(row=3, column=1, sticky="ew", padx=5, pady=2)
        
        user_button_frame = ttk.Frame(user_frame)
        user_button_frame.grid(row=4, column=0, columnspan=2, pady=5)
        ttk.Button(user_button_frame, text="Save", command=self._create_new_user, style="TButton").pack(side=tk.LEFT, padx=2)

        # ===== User Management Section (New Row) =====
        user_mgmt_frame = ttk.LabelFrame(main_frame, text="Manage Existing Users", padding="10")
        user_mgmt_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5, ipady=5)
        user_mgmt_frame.grid_columnconfigure(0, weight=1)
        user_mgmt_frame.grid_columnconfigure(1, weight=1)

        # User List (Left side)
        list_subframe = ttk.Frame(user_mgmt_frame)
        list_subframe.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        list_subframe.grid_columnconfigure(0, weight=1)

        ttk.Label(list_subframe, text="Existing Users:", style="TLabel").pack(anchor="w", pady=(0, 5))

        # Create treeview for user list with integrated scrollbar
        columns = ("Username", "Role")
        self.user_tree = ttk.Treeview(list_subframe, columns=columns, show="headings", height=6)
        self.user_tree.heading("Username", text="Username")
        self.user_tree.heading("Role", text="Role")
        self.user_tree.column("Username", width=150)
        self.user_tree.column("Role", width=100)
        self.user_tree.bind("<<TreeviewSelect>>", self._on_user_selected)

        # Create and configure scrollbar inside the treeview
        scrollbar = ttk.Scrollbar(self.user_tree, orient="vertical", command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview with scrollbar integrated
        self.user_tree.pack(fill="both", expand=True, pady=(0, 5))
        scrollbar.pack(side="right", fill="y")

        # Refresh button
        ttk.Button(list_subframe, text="Refresh List", command=self._load_user_list, style="TButton").pack(pady=5)

        # Edit User (Right side)
        edit_subframe = ttk.LabelFrame(user_mgmt_frame, text="Edit Selected User", padding="10")
        edit_subframe.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        edit_subframe.grid_columnconfigure(1, weight=1)

        # Username field
        ttk.Label(edit_subframe, text="Username:", style="TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.edit_username_var = tk.StringVar()
        self.edit_username_entry = ttk.Entry(edit_subframe, textvariable=self.edit_username_var, style="TEntry")
        self.edit_username_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=2)

        # Current password display
        ttk.Label(edit_subframe, text="Current Password:", style="TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.current_password_label = ttk.Label(edit_subframe, text="••••••••", font=("Helvetica", 10, "italic"))
        self.current_password_label.grid(row=1, column=1, sticky="w", padx=5, pady=2)

        # New password field
        ttk.Label(edit_subframe, text="New Password:", style="TLabel").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.edit_password_var = tk.StringVar()
        self.edit_password_entry = ttk.Entry(edit_subframe, textvariable=self.edit_password_var, show="*", style="TEntry")
        self.edit_password_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=2)

        # Confirm new password field
        ttk.Label(edit_subframe, text="Confirm Password:", style="TLabel").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        self.edit_password_confirm_var = tk.StringVar()
        self.edit_password_confirm_entry = ttk.Entry(edit_subframe, textvariable=self.edit_password_confirm_var, show="*", style="TEntry")
        self.edit_password_confirm_entry.grid(row=3, column=1, sticky="ew", padx=5, pady=2)

        # Role field
        ttk.Label(edit_subframe, text="User Role:", style="TLabel").grid(row=4, column=0, sticky="w", padx=5, pady=2)
        self.edit_role_var = tk.StringVar()
        self.edit_role_dropdown = ttk.Combobox(edit_subframe, textvariable=self.edit_role_var, 
                                             values=["user", "operator", "admin"], state="readonly", style="TCombobox")
        self.edit_role_dropdown.grid(row=4, column=1, sticky="ew", padx=5, pady=2)

        # Buttons
        button_frame = ttk.Frame(edit_subframe)
        button_frame.grid(row=5, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Update User", command=self._update_user, style="TButton").pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="Delete User", command=self._delete_user, style="TButton").pack(side=tk.LEFT, padx=2)

        # Initially disable edit fields
        self._set_edit_fields_state("disabled")

        # Load user list
        self._load_user_list()

        self._update_activation_status_label_on_gui()

    def _update_admin_password(self):
        """
        Changes the admin password in the database and updates the in-memory variable.
        Now requires the current password for confirmation.
        """
        current_password = self.current_admin_password_entry.get()
        new_password = self.new_admin_password_var.get()
        confirm_password = self.new_admin_password_confirm_var.get()
        
        # Validate current password
        current_password_hash = hashlib.sha256(current_password.encode()).hexdigest()
        if current_password_hash != self.admin_password:
            self.msg_box.showerror("Authentication Failed", "Incorrect current password.")
            self.current_admin_password_entry.delete(0, tk.END)
            return

        if new_password != confirm_password:
            self.msg_box.showerror("Password Mismatch", "The new password and confirmation do not match.")
            return

        if not new_password:
            self.msg_box.showerror("Invalid Password", "Password cannot be empty.")
            return

        new_password_hash = hashlib.sha256(new_password.encode()).hexdigest()
        
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET password_hash = ? WHERE username = ?", (new_password_hash, ADMIN_USERNAME))
                conn.commit()
                self.admin_password = new_password_hash
            
            self.msg_box.showinfo("Success", "Admin password updated successfully.")
            self.current_admin_password_entry.delete(0, tk.END)
            self.new_admin_password_var.set("")
            self.new_admin_password_confirm_var.set("")

        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to update admin password in database: {e}")
    
    def _create_new_user(self):
        """
        Creates a new user account in the database.
        """
        username = self.new_user_username_entry.get().strip()
        password = self.new_user_password_entry.get()
        confirm_password = self.new_user_password_confirm_entry.get()
        role = self.new_user_role_var.get()
        
        # Validation
        if not username:
            self.msg_box.showerror("Invalid Input", "Username cannot be empty.")
            return
        
        if len(username) < 3:
            self.msg_box.showerror("Invalid Input", "Username must be at least 3 characters long.")
            return
        
        if not password:
            self.msg_box.showerror("Invalid Input", "Password cannot be empty.")
            return
        
        if len(password) < 6:
            self.msg_box.showerror("Invalid Input", "Password must be at least 6 characters long.")
            return
        
        if password != confirm_password:
            self.msg_box.showerror("Password Mismatch", "Passwords do not match.")
            return
        
        if not role:
            self.msg_box.showerror("Invalid Input", "Please select a user role.")
            return
        
        # Hash the password
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                
                # Check if user already exists
                cursor.execute("SELECT username FROM users WHERE username = ?", (username,))
                if cursor.fetchone():
                    self.msg_box.showerror("User Exists", f"Username '{username}' already exists. Please choose a different username.")
                    return
                
                # Insert new user
                cursor.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                              (username, password_hash, role))
                conn.commit()
            
            self.msg_box.showinfo("Success", f"User '{username}' created successfully with role '{role}'.\nThey can now log in with their password.")
            
            # Clear the form
            self.new_user_username_entry.delete(0, tk.END)
            self.new_user_password_entry.delete(0, tk.END)
            self.new_user_password_confirm_entry.delete(0, tk.END)
            self.new_user_role_var.set("user")
            
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to create user: {e}")
        
    def _update_activation_status_label_on_gui(self):
        """
        Updates the activation status label in the GUI.
        """
        today = datetime.now().date()
        expiry_date_str = self.config.get("expiry_date")

        if self.is_app_activated:
            expiry_str = self.config.get("expiry_date", "N/A")
            self.activation_status_var.set(f"Application Activated. Expires: {expiry_str}")
        else:
            if expiry_date_str:
                expiry_date = datetime.strptime(expiry_date_str, "%m/%d/%Y").date()
                if today > expiry_date:
                    self.activation_status_var.set(f"Trial Expired on {expiry_date_str}. Please activate.")
                else:
                    self.activation_status_var.set(f"Trial Active. Expires: {expiry_date_str}")
            else:
                self.activation_status_var.set("Status Unknown. Please try activating.")

    def _load_user_list(self):
        """Load the list of users from the database."""
        try:
            # Clear existing items
            for item in self.user_tree.get_children():
                self.user_tree.delete(item)
            
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT username, role FROM users ORDER BY username")
                for row in cursor.fetchall():
                    self.user_tree.insert("", "end", values=(row[0], row[1]))
                    
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to load user list: {e}")

    def _on_user_selected(self, event):
        """Handle user selection in the treeview."""
        selection = self.user_tree.selection()
        if selection:
            item = self.user_tree.item(selection[0])
            username, role = item['values']
            
            # Load user data into edit fields
            self.edit_username_var.set(username)
            self.edit_role_var.set(role)
            self.edit_password_var.set("")
            self.edit_password_confirm_var.set("")
            
            # Enable edit fields
            self._set_edit_fields_state("normal")
            
            # Disable username editing for the currently logged-in admin user
            if username == ADMIN_USERNAME and self.logged_in_user == ADMIN_USERNAME:
                self.edit_username_entry.config(state="disabled")
            else:
                self.edit_username_entry.config(state="normal")

    def _set_edit_fields_state(self, state):
        """Set the state of all edit fields."""
        self.edit_username_entry.config(state=state)
        self.edit_password_entry.config(state=state)
        self.edit_password_confirm_entry.config(state=state)
        self.edit_role_dropdown.config(state=state)

    def _update_user(self):
        """Update the selected user's information."""
        selection = self.user_tree.selection()
        if not selection:
            self.msg_box.showerror("No Selection", "Please select a user to update.")
            return
        
        old_username = self.user_tree.item(selection[0])['values'][0]
        new_username = self.edit_username_var.get().strip()
        new_password = self.edit_password_var.get().strip()
        confirm_password = self.edit_password_confirm_var.get().strip()
        new_role = self.edit_role_var.get()
        
        # Validation
        if not new_username:
            self.msg_box.showerror("Invalid Input", "Username cannot be empty.")
            return
        
        if new_password and new_password != confirm_password:
            self.msg_box.showerror("Password Mismatch", "New password and confirmation do not match.")
            return
        
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                
                # Check if new username already exists (if username is being changed)
                if new_username != old_username:
                    cursor.execute("SELECT username FROM users WHERE username = ?", (new_username,))
                    if cursor.fetchone():
                        self.msg_box.showerror("Username Exists", f"Username '{new_username}' already exists.")
                        return
                
                # Update user information
                if new_password:
                    # Update username, password, and role
                    password_hash = hashlib.sha256(new_password.encode()).hexdigest()
                    cursor.execute("UPDATE users SET username = ?, password_hash = ?, role = ? WHERE username = ?",
                                 (new_username, password_hash, new_role, old_username))
                else:
                    # Update username and role only (keep existing password)
                    cursor.execute("UPDATE users SET username = ?, role = ? WHERE username = ?",
                                 (new_username, new_role, old_username))
                
                conn.commit()
                
                # Update current session if user changed their own username
                if old_username == self.logged_in_user:
                    self.logged_in_user = new_username
                
                self.msg_box.showinfo("Success", f"User '{old_username}' updated successfully.")
                self._load_user_list()
                
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to update user: {e}")

    def _delete_user(self):
        """Delete the selected user."""
        selection = self.user_tree.selection()
        if not selection:
            self.msg_box.showerror("No Selection", "Please select a user to delete.")
            return
        
        username = self.user_tree.item(selection[0])['values'][0]
        
        # Prevent deletion of the currently logged-in user
        if username == self.logged_in_user:
            self.msg_box.showerror("Cannot Delete", "You cannot delete your own account while logged in.")
            return
        
        # Prevent deletion of the default admin user
        if username == ADMIN_USERNAME:
            self.msg_box.showerror("Cannot Delete", "Cannot delete the default administrator account.")
            return
        
        # Confirm deletion
        confirmed = self.msg_box.askyesno("Confirm Delete", 
                                        f"Are you sure you want to delete user '{username}'? This action cannot be undone.")
        if confirmed:
            try:
                with sqlite3.connect(DB_FILE) as conn:
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM users WHERE username = ?", (username,))
                    conn.commit()
                    
                    self.msg_box.showinfo("Success", f"User '{username}' deleted successfully.")
                    self._load_user_list()
                    self._set_edit_fields_state("disabled")
                    
            except sqlite3.Error as e:
                self.msg_box.showerror("Database Error", f"Failed to delete user: {e}")

    def _save_print_settings(self):
        """
        Saves all print-related settings to the config file.
        """
        try:
            # Save printer settings
            self.config["selected_printer"] = self.selected_printer_var.get()
            self.config["print_copies"] = int(self.print_copies_var.get())
            
            # Save PDF settings
            self.config["pdf_page_size"] = self.pdf_page_size_var.get()
            self.config["pdf_orientation"] = self.pdf_orientation_var.get()
            
            # Save all settings to file
            self.save_config()
            
            self.msg_box.showinfo("Settings Saved", "Printer and PDF settings saved successfully.")
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to save print settings: {e}")

    def _on_template_text_change(self, event=None):
        """
        Event handler for changes in the print template text editor.
        Updates the corresponding variable and triggers a debounced preview and save.
        """
        # Only process if the template text widget exists
        if not hasattr(self, 'print_template_text') or not self.print_template_text:
            return
            
        current_text_content = self.print_template_text.get("1.0", tk.END).strip()
        
        # Save to the correct variable based on the selected radio button
        if self.current_template_editor_mode_var.get() == "ONE_WAY":
            self.one_way_print_template_var.set(current_text_content)
        else: # "TWO_WAY"
            self.two_way_print_template_var.set(current_text_content)

        self.print_template_text.edit_modified(False)
        self._debounce_template_updates()

    def _debounce_template_updates(self):
        """
        Debounces and schedules a function to update the previews and save the templates.
        """
        self.root.after_cancel(getattr(self, '_template_update_id', ''))
        self._template_update_id = self.root.after(1000, self._update_and_save_templates)

    def _update_and_save_templates(self):
        """
        Updates the previews and saves the current templates to the database.
        """
        self._update_all_previews()
        self._save_templates_to_db()

    def _save_templates_to_db(self):
        """
        Saves the print templates from the in-memory variables to the database.
        """
        try:
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'ONE_WAY'",
                             (self.one_way_print_template_var.get(),))
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'TWO_WAY'",
                             (self.two_way_print_template_var.get(),))
                conn.commit()
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Save Error", f"Failed to save print templates to database: {e}")

    def _update_all_previews(self):
        """Updates both the raw text and PDF previews."""
        self._update_print_preview()
        
    def _update_template_text_and_previews(self):
        """
        Updates the template text editor and refreshes all previews.
        This is called when a user changes the radio button.
        Only updates if the widgets exist (they're in the Print Templates tab).
        """
        # Check if the template text widget exists
        if not hasattr(self, 'print_template_text') or not self.print_template_text:
            return
            
        current_mode = self.current_template_editor_mode_var.get()
        
        template_content = ""
        if current_mode == "ONE_WAY":
            template_content = self.one_way_print_template_var.get()
        else: # "TWO_WAY"
            template_content = self.two_way_print_template_var.get()

        self.print_template_text.config(state=tk.NORMAL) # Allow editing
        self.print_template_text.delete("1.0", tk.END)
        self.print_template_text.insert("1.0", template_content)
        self.print_template_text.edit_modified(False)
        self.print_template_text.config(state=tk.DISABLED) # Revert to disabled to prevent unintended edits
        
        # Force a redraw to ensure the text widget updates immediately
        self.root.update_idletasks() 

        # Update the previews
        self._update_all_previews()

    def _update_print_template_editor_content(self):
        # This method is now only for switching radio buttons
        self._update_template_text_and_previews()

    def _update_template_font(self):
        """
        Updates the font of the text editor and preview based on user selection.
        Only updates if the widgets exist (they're in the Print Templates tab).
        """
        # Check if the template text widget exists
        if not hasattr(self, 'print_template_text') or not self.print_template_text:
            return
            
        font_family = self.print_template_font_family_var.get()
        font_size = self.print_template_font_size_var.get()
        font_weight = "bold" if self.print_template_font_bold_var.get() else "normal"
        font_tuple = (font_family, font_size, font_weight)

        self.print_template_text.config(font=font_tuple)
        # The PDF preview canvas font is set differently
        self._update_print_preview()
        
    def _insert_placeholder_from_event(self, event):
        """
        Helper function to get the selected placeholder from a double-click event.
        """
        try:
            index = self.placeholder_listbox.nearest(event.y)
            self.placeholder_listbox.selection_clear(0, tk.END)
            self.placeholder_listbox.selection_set(index)
            self._insert_placeholder_into_template_btn_click()
        except IndexError:
            pass


    def _insert_placeholder_into_template_btn_click(self):
        """
        Inserts a selected placeholder from the listbox into the text editor.
        """
        selected_indices = self.placeholder_listbox.curselection()
        if not selected_indices:
            self.msg_box.showwarning("No Selection", "Please select a placeholder from the list to insert.")
            return
        
        selected_text = self.placeholder_listbox.get(selected_indices[0])
        placeholder_match = re.search(r"(\{.*\})", selected_text)
        if placeholder_match:
            placeholder = placeholder_match.group(1)
            self.print_template_text.config(state=tk.NORMAL) # Allow editing
            self.print_template_text.insert(tk.INSERT, placeholder)
            self.print_template_text.config(state=tk.DISABLED) # Disable editing
            self._on_template_text_change()

    def _reset_print_format_to_default(self):
        """
        Resets print-related settings only (not templates - admins define templates).
        """
        # Using a non-blocking dialog with a callback
        def on_confirm(result):
            if result:
                # Reset only settings, NOT templates (admin-defined templates are preserved)
                self.print_line_spacing.set(0)
                self.print_encoding_var.set("utf-8")
                self.print_include_barcode_var.set(False)
                self.selected_printer_var.set("")
                self.print_copies_var.set(1)
                
                # Reset Raw Text Print Settings
                self.print_template_font_family_var.set("Courier New")
                self.print_template_font_size_var.set(10)
                self.print_template_font_bold_var.set(False)
                
                # Reset new PDF Print Settings
                self.pdf_print_template_font_family_var.set("Helvetica")
                self.pdf_print_template_font_size_var.set(12)
                self.pdf_print_template_font_bold_var.set(False)
                self.pdf_page_size_var.set("Half Letter")
                self.pdf_orientation_var.set("Portrait")
                
                try:
                    self._update_template_text_and_previews()
                except Exception:
                    pass  # Widget may not exist in this tab
                    
                try:
                    self._update_template_font()
                except Exception:
                    pass  # Widget may not exist in this tab
                    
                self._update_all_previews()
                self.msg_box.showinfo("Reset", "Print format settings reset. Admin-defined templates preserved.")
        
        self.msg_box.askyesno("Reset Print Format", "Reset print format settings to default? (Templates remain as defined by admin)", on_confirm)


    def build_master_data_tab(self, tab: ttk.Frame):
        """
        Builds the GUI for the Master Data management tab.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        categories = ["companies", "trucks", "products", "drivers", "origins", "destinations", "designations", "senders"]
        self.master_entries = {}

        tab.grid_columnconfigure(1, weight=1)
        tab.grid_columnconfigure(4, weight=2)

        for i, category in enumerate(categories):
            ttk.Label(tab, text=category.capitalize() + ":", style="TLabel").grid(row=i, column=0, sticky="w", padx=5, pady=2)

            entry_var = tk.StringVar()
            entry_var.trace_add("write", lambda name, index, mode, var=entry_var: var.set(var.get().upper()))

            entry_widget = ttk.Entry(tab, textvariable=entry_var, style="TEntry")
            entry_widget.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
            self.master_entries[category] = {'entry_var': entry_var, 'entry_widget': entry_widget}

            ttk.Button(tab, text="Add", command=lambda c=category: self.add_master(c), width=8, style="TButton").grid(row=i, column=2, padx=2, pady=2)
            ttk.Button(tab, text="Delete", command=lambda c=category: self.delete_master(c), width=8, style="TButton").grid(row=i, column=3, padx=2, pady=2)

            listbox = tk.Listbox(tab, height=3, exportselection=False, font=("Helvetica", 10), selectmode=tk.SINGLE)
            listbox.grid(row=i, column=4, sticky="nsew", padx=5, pady=2)
            self.master_entries[category]['listbox'] = listbox

            listbox_vsb = ttk.Scrollbar(tab, orient="vertical", command=listbox.yview)
            listbox_vsb.grid(row=i, column=5, sticky="ns")
            listbox.config(yscrollcommand=listbox_vsb.set)

            listbox.bind("<Double-Button-1>", lambda event, c=category: self.populate_master_entry(event, c))
            listbox.bind("<Return>", lambda event, c=category: self.populate_master_entry(event, c))

        self.load_master_data_lists()

    def build_customized_tab(self, tab: ttk.Frame):
        """
        Builds the GUI for customizing the entry form fields and their order.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        field_selection_frame = ttk.LabelFrame(tab, text="Select Entry Form Fields and Order", padding="10")
        field_selection_frame.pack(padx=10, pady=10, fill="both", expand=True)
        field_selection_frame.grid_columnconfigure(0, weight=1)
        field_selection_frame.grid_columnconfigure(2, weight=1)

        ttk.Label(field_selection_frame, text="Available Fields:", style="TLabel").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.available_entry_fields_listbox = tk.Listbox(field_selection_frame, selectmode=tk.SINGLE, height=10, font=("Helvetica", 10))
        self.available_entry_fields_listbox.grid(row=1, column=0, rowspan=4, sticky="nsew", padx=5, pady=5)
        self.available_entry_fields_listbox_vsb = ttk.Scrollbar(field_selection_frame, orient="vertical", command=self.available_entry_fields_listbox.yview)
        self.available_entry_fields_listbox_vsb.grid(row=1, column=1, rowspan=4, sticky="ns")
        self.available_entry_fields_listbox.config(yscrollcommand=self.available_entry_fields_listbox_vsb.set)

        button_col = 1
        ttk.Button(field_selection_frame, text="Add >>", command=self._add_entry_field, style="TButton").grid(row=1, column=button_col, padx=5, pady=2, sticky="s")
        ttk.Button(field_selection_frame, text="<< Remove", command=self._remove_entry_field, style="TButton").grid(row=2, column=button_col, padx=5, pady=2, sticky="n")
        ttk.Button(field_selection_frame, text="Move Up", command=self._move_entry_field_up, style="TButton").grid(row=3, column=button_col, padx=5, pady=2, sticky="s")
        ttk.Button(field_selection_frame, text="Move Down", command=self._move_entry_field_down, style="TButton").grid(row=4, column=button_col, padx=5, pady=2, sticky="n")

        ttk.Label(field_selection_frame, text="Selected Fields (Entry Order):", style="TLabel").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.selected_entry_fields_listbox = tk.Listbox(field_selection_frame, selectmode=tk.SINGLE, height=10, font=("Helvetica", 10))
        self.selected_entry_fields_listbox.grid(row=1, column=2, rowspan=4, sticky="nsew", padx=5, pady=5)
        self.selected_entry_fields_listbox_vsb = ttk.Scrollbar(field_selection_frame, orient="vertical", command=self.selected_entry_fields_listbox.yview)
        self.selected_entry_fields_listbox_vsb.grid(row=1, column=3, rowspan=4, sticky="ns")
        self.selected_entry_fields_listbox.config(yscrollcommand=self.selected_entry_fields_listbox_vsb.set)

        self.selected_entry_fields_listbox.bind("<<ListboxSelect>>", lambda event: self._on_entry_field_selection_changed())

        # ttk.Button(tab, text="Reset Entry Form to Default", command=self._reset_entry_form_to_default, style="TButton").pack(pady=10)
        ttk.Button(tab, text="SAVE ENTRY FORM LAYOUT", command=self._save_entry_form_layout, style="TButton").pack(pady=10)

        # Price Computation Settings Frame (Admin/Operator only)
        self.price_settings_frame = ttk.LabelFrame(tab, text="Price Computation Settings (Admin/Operator Only)", padding="10")
        self.price_settings_frame.pack(padx=10, pady=10, fill="x")
        
        # Check if user is admin or operator and enable/disable accordingly
        is_privileged = hasattr(self, 'current_user_role') and self.current_user_role in ['admin', 'operator']
        
        # Enable price computation (variable already initialized in __init__)
        price_enabled_cb = ttk.Checkbutton(self.price_settings_frame, text="Enable Price Computation", 
                                         variable=self.price_computation_enabled_var,
                                         command=self._on_price_computation_toggled,
                                         state="normal" if is_privileged else "disabled")
        price_enabled_cb.grid(row=0, column=0, columnspan=2, sticky="w", pady=2)
        
        # Base weight and price
        ttk.Label(self.price_settings_frame, text="Base Weight (kg):").grid(row=1, column=0, sticky="e", padx=5, pady=2)
        self.base_weight_var = tk.DoubleVar(value=self.config.get("base_weight", DEFAULT_BASE_WEIGHT))
        base_weight_entry = ttk.Entry(self.price_settings_frame, textvariable=self.base_weight_var, width=10,
                                     state="normal" if is_privileged else "disabled")
        base_weight_entry.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        
        ttk.Label(self.price_settings_frame, text="Base Price:").grid(row=2, column=0, sticky="e", padx=5, pady=2)
        self.base_price_var = tk.DoubleVar(value=self.config.get("base_price", DEFAULT_BASE_PRICE))
        base_price_entry = ttk.Entry(self.price_settings_frame, textvariable=self.base_price_var, width=10,
                                    state="normal" if is_privileged else "disabled")
        base_price_entry.grid(row=2, column=1, sticky="w", padx=5, pady=2)
        
        # Increment weight and price
        ttk.Label(self.price_settings_frame, text="Increment Weight (kg):").grid(row=3, column=0, sticky="e", padx=5, pady=2)
        self.increment_weight_var = tk.DoubleVar(value=self.config.get("increment_weight", DEFAULT_INCREMENT_WEIGHT))
        increment_weight_entry = ttk.Entry(self.price_settings_frame, textvariable=self.increment_weight_var, width=10,
                                          state="normal" if is_privileged else "disabled")
        increment_weight_entry.grid(row=3, column=1, sticky="w", padx=5, pady=2)
        
        ttk.Label(self.price_settings_frame, text="Increment Price:").grid(row=4, column=0, sticky="e", padx=5, pady=2)
        self.increment_price_var = tk.DoubleVar(value=self.config.get("increment_price", DEFAULT_INCREMENT_PRICE))
        increment_price_entry = ttk.Entry(self.price_settings_frame, textvariable=self.increment_price_var, width=10,
                                         state="normal" if is_privileged else "disabled")
        increment_price_entry.grid(row=4, column=1, sticky="w", padx=5, pady=2)
        
        # Save button for price settings
        save_price_btn = ttk.Button(self.price_settings_frame, text="Save Price Settings", 
                                  command=self._save_price_settings,
                                  state="normal" if is_privileged else "disabled")
        save_price_btn.grid(row=5, column=0, columnspan=2, pady=5)

        # Update price computation settings state in case user is already logged in
        self._update_price_settings_state()

        # Add Total Price display for current transaction
        # total_price_frame = ttk.LabelFrame(tab, text="Current Transaction Price", padding="10")
        # total_price_frame.pack(padx=10, pady=10, fill="x")
        
        # ttk.Label(total_price_frame, text="Total Price:", font=("Helvetica", 12, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        # self.transaction_total_price_var = tk.StringVar(value="0")
        # self.transaction_total_price_label = ttk.Label(total_price_frame, textvariable=self.transaction_total_price_var, 
        #                                              font=("Courier", 14, "bold"), foreground="green")
        # self.transaction_total_price_label.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        # Bind weight changes to update transaction price
        self.weight_value.trace_add("write", self._update_transaction_price_display)

        self._update_customized_entry_form_lists()

    def _update_customized_entry_form_lists(self):
        """
        Updates the two listboxes in the entry form customization tab.
        """
        self.available_entry_fields_listbox.delete(0, tk.END)
        self.selected_entry_fields_listbox.delete(0, tk.END)

        selected_keys = {key for _, key in self.selected_entry_fields}

        for display_name, key in self.available_entry_fields:
            if key not in selected_keys:
                self.available_entry_fields_listbox.insert(tk.END, display_name)

        for display_name, _ in self.selected_entry_fields:
            self.selected_entry_fields_listbox.insert(tk.END, display_name)

    def _add_entry_field(self):
        """
        Adds a selected field from the 'available' list to the 'selected' list.
        """
        selected_indices = self.available_entry_fields_listbox.curselection()
        if not selected_indices:
            return

        index = selected_indices[0]
        display_name_to_add = self.available_entry_fields_listbox.get(index)

        field_to_add = next(((d, k) for d, k in self.available_entry_fields if d == display_name_to_add), None)

        if field_to_add and field_to_add not in self.selected_entry_fields:
            self.selected_entry_fields.append(field_to_add)
            self._update_customized_entry_form_lists()
            self.rebuild_entry_form_tab()
            self.save_config()  # Real-time save

    def _remove_entry_field(self):
        """
        Removes a selected field from the 'selected' list.
        """
        selected_indices = self.selected_entry_fields_listbox.curselection()
        if not selected_indices:
            return

        index = selected_indices[0]
        display_name_to_remove = self.selected_entry_fields_listbox.get(index)

        field_to_remove = next(((d, k) for d, k in self.available_entry_fields if d == display_name_to_remove), None)

        if field_to_remove and field_to_remove[1] == "truck_plate":
            self.msg_box.showerror("Cannot Remove", "The 'Truck Plate' field cannot be removed as it is essential for transaction management.")
            return

        if field_to_remove:
            self.selected_entry_fields.remove(field_to_remove)
            self._update_customized_entry_form_lists()
            self.rebuild_entry_form_tab()
            self.save_config()  # Real-time save

    def _move_entry_field_up(self):
        """
        Moves a selected field up in the 'selected' list.
        """
        selected_indices = self.selected_entry_fields_listbox.curselection()
        if not selected_indices:
            return

        index = selected_indices[0]
        if index > 0:
            field = self.selected_entry_fields.pop(index)
            self.selected_entry_fields.insert(index - 1, field)
            self._update_customized_entry_form_lists()
            self.selected_entry_fields_listbox.selection_set(index - 1)
            self.rebuild_entry_form_tab()
            self.save_config()  # Real-time save

    def _move_entry_field_down(self):
        """
        Moves a selected field down in the 'selected' list.
        """
        selected_indices = self.selected_entry_fields_listbox.curselection()
        if not selected_indices:
            return

        index = selected_indices[0]
        if index < len(self.selected_entry_fields) - 1:
            field = self.selected_entry_fields.pop(index)
            self.selected_entry_fields.insert(index + 1, field)
            self._update_customized_entry_form_lists()
            self.selected_entry_fields_listbox.selection_set(index + 1)
            self.rebuild_entry_form_tab()
            self.save_config()  # Real-time save

    def _reset_entry_form_to_default(self):
        """
        Resets the entry form field layout to the original default.
        """
        # Using a non-blocking dialog with a callback
        def on_confirm(result):
            if result:
                self.selected_entry_fields = list(self.default_entry_fields)
                self._update_customized_entry_form_lists()
                self.rebuild_entry_form_tab()
                self.msg_box.showinfo("Reset", "Entry Form layout reset to default. Remember to save settings.")
        
        self.msg_box.askyesno("Reset Entry Form Layout", "Are you sure you want to reset the Entry Form layout to default settings?", on_confirm)


    def _save_entry_form_layout(self):
        """
        Saves the current entry form field layout to the config file.
        """
        self.save_config()
        self.rebuild_entry_form_tab()  # Update transaction details list
        
        # Switch to Entry Form tab to show real-time sync
        if hasattr(self, 'notebook'):
            for i in range(self.notebook.index("end")):
                if self.notebook.tab(i, "text") == "Entry Form":
                    self.notebook.select(i)
                    break
        
        self.msg_box.showinfo("Layout Saved", "Entry Form layout saved and applied immediately.")

    def build_print_settings_tab(self, tab: ttk.Frame):
        """Builds the unified Print Settings tab with all components."""
        for widget in tab.winfo_children():
            widget.destroy()
        
        # Main frame for the tab
        main_frame = ttk.Frame(tab, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Printer Selection Section
        printer_frame = ttk.LabelFrame(main_frame, text="Printer Settings", padding="10")
        printer_frame.pack(fill="x", pady=(0, 10))
        
        # Printer selection
        printer_row = ttk.Frame(printer_frame)
        printer_row.pack(fill="x", pady=2)
        
        ttk.Label(printer_row, text="Printer:").pack(side=tk.LEFT)
        self.printer_combobox = ttk.Combobox(printer_row, textvariable=self.selected_printer_var, width=30)
        self.printer_combobox.pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(printer_row, text="Refresh", command=self._populate_printer_list).pack(side=tk.LEFT, padx=(5, 0))
        
        # Number of copies
        copies_row = ttk.Frame(printer_frame)
        copies_row.pack(fill="x", pady=2)
        ttk.Label(copies_row, text="Number of Copies:").pack(side=tk.LEFT)
        ttk.Spinbox(copies_row, from_=1, to=10, textvariable=self.print_copies_var, width=10).pack(side=tk.LEFT, padx=(5, 0))
        
        # PDF Settings Section
        pdf_frame = ttk.LabelFrame(main_frame, text="PDF Settings", padding="10")
        pdf_frame.pack(fill="x", pady=(0, 10))
        
        # PDF Page Size
        page_size_row = ttk.Frame(pdf_frame)
        page_size_row.pack(fill="x", pady=2)
        ttk.Label(page_size_row, text="Page Size:").pack(side=tk.LEFT)
        page_size_combo = ttk.Combobox(page_size_row, textvariable=self.pdf_page_size_var, values=["Letter", "A6", "Legal", "Half Letter"], width=6)  # Added "Half Letter"
        page_size_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        # PDF Orientation
        orientation_row = ttk.Frame(pdf_frame)
        orientation_row.pack(fill="x", pady=2)
        ttk.Label(orientation_row, text="Orientation:").pack(side=tk.LEFT)
        orientation_combo = ttk.Combobox(orientation_row, textvariable=self.pdf_orientation_var,
                                         values=["Portrait", "Landscape"], width=6)
        orientation_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        # Template Editing Section
        template_frame = ttk.LabelFrame(main_frame, text="Print Templates", padding="10")
        template_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # Template selection buttons
        template_select_frame = ttk.Frame(template_frame)
        template_select_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Radiobutton(template_select_frame, text="One Way Template", 
                       variable=self.current_template_editor_mode_var, value="ONE_WAY").pack(side=tk.LEFT)
        ttk.Radiobutton(template_select_frame, text="Two Way Template", 
                       variable=self.current_template_editor_mode_var, value="TWO_WAY").pack(side=tk.LEFT, padx=(10, 0))
        
        # Template editor
        editor_frame = ttk.Frame(template_frame)
        editor_frame.pack(fill="both", expand=True)
        
        # Text editor for template
        self.print_template_text = tk.Text(editor_frame, wrap="none", height=15)
        self.print_template_text.pack(side=tk.LEFT, fill="both", expand=True)
        
        # Scrollbars for text editor
        h_scroll = ttk.Scrollbar(editor_frame, orient="horizontal", command=self.print_template_text.xview)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        v_scroll = ttk.Scrollbar(editor_frame, orient="vertical", command=self.print_template_text.yview)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.print_template_text.config(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        
        # Bind text changes
        self.print_template_text.bind("<KeyRelease>", lambda event: self._on_template_text_change())
        
        # Preview frame
        preview_frame = ttk.LabelFrame(template_frame, text="Preview")
        preview_frame.pack(fill="both", expand=True, pady=(10, 0))
        
        # Preview text widget
        self.print_preview_text = tk.Text(preview_frame, wrap="word", height=8)
        self.print_preview_text.pack(expand=True, fill="both")
        self.print_preview_text.config(state="disabled")
        
        # Buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill="x", pady=(10, 0))
        
        ttk.Button(buttons_frame, text="Save Settings", command=self._save_print_settings).pack(side=tk.RIGHT, padx=5)
        ttk.Button(buttons_frame, text="Reset to Default", command=self._reset_print_format_to_default).pack(side=tk.RIGHT, padx=5)
        
        # Initialize with current templates
        self._update_template_text_and_previews()

    def build_print_and_ticket_settings_tab(self, tab: ttk.Frame):
        """Builds the Ticket Format tab with Printer/PDF settings on the right."""
        # Only build once - don't destroy/rebuild on each call
        if tab.winfo_children():
            return
        
        # Create main frame with left and right sections
        main_frame = ttk.Frame(tab, padding=5)
        main_frame.pack(fill="both", expand=True)
        
        # LEFT SIDE: Ticket Format
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 5))
        self.build_ticket_output_tab(left_frame)
        
        # RIGHT SIDE: Printer and PDF Settings
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill="both", expand=False, padx=(5, 0))
        self.build_print_settings_right_panel(right_frame)

    def build_print_settings_right_panel(self, panel: ttk.Frame):
        """Builds the right panel with Printer and PDF settings only (no templates)."""
        # Main frame for the panel
        main_frame = ttk.Frame(panel, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Printer Selection Section
        printer_frame = ttk.LabelFrame(main_frame, text="Printer Settings", padding="10")
        printer_frame.pack(fill="x", pady=(0, 10))
        
        # Printer selection
        printer_row = ttk.Frame(printer_frame)
        printer_row.pack(fill="x", pady=2)
        
        ttk.Label(printer_row, text="Printer:").pack(side=tk.LEFT)
        self.printer_combobox = ttk.Combobox(printer_row, textvariable=self.selected_printer_var, width=25)
        self.printer_combobox.pack(side=tk.LEFT, padx=(5, 0), fill="x", expand=True)
        ttk.Button(printer_row, text="Refresh", command=self._populate_printer_list).pack(side=tk.LEFT, padx=(5, 0))
        
        # Populate printer list on initialization
        self._populate_printer_list()
        
        # Number of copies
        copies_row = ttk.Frame(printer_frame)
        copies_row.pack(fill="x", pady=5)
        ttk.Label(copies_row, text="Number of Copies:").pack(side=tk.LEFT)
        ttk.Spinbox(copies_row, from_=1, to=10, textvariable=self.print_copies_var, width=10).pack(side=tk.LEFT, padx=(5, 0))
        
        # PDF Settings Section
        pdf_frame = ttk.LabelFrame(main_frame, text="PDF Settings", padding="10")
        pdf_frame.pack(fill="x", pady=(0, 10))
        
        # PDF Page Size
        page_size_row = ttk.Frame(pdf_frame)
        page_size_row.pack(fill="x", pady=5)
        ttk.Label(page_size_row, text="Page Size:").pack(side=tk.LEFT)
        page_size_combo = ttk.Combobox(page_size_row, textvariable=self.pdf_page_size_var, 
                                       values=["Letter", "A6", "Legal", "Half Letter"], width=12, state="readonly")
        page_size_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        # PDF Orientation
        orientation_row = ttk.Frame(pdf_frame)
        orientation_row.pack(fill="x", pady=5)
        ttk.Label(orientation_row, text="Orientation:").pack(side=tk.LEFT)
        orientation_combo = ttk.Combobox(orientation_row, textvariable=self.pdf_orientation_var,
                                         values=["Portrait", "Landscape"], width=12, state="readonly")
        orientation_combo.pack(side=tk.LEFT, padx=(5, 0))
        
        # Ticket Number Settings Section
        ticket_frame = ttk.LabelFrame(main_frame, text="Ticket Number Settings", padding="10")
        ticket_frame.pack(fill="x", pady=(0, 10))
        
        # Starting Ticket Number
        ticket_number_row = ttk.Frame(ticket_frame)
        ticket_number_row.pack(fill="x", pady=5)
        ttk.Label(ticket_number_row, text="Starting Ticket Number:").pack(side=tk.LEFT)
        ticket_number_spinbox = ttk.Spinbox(ticket_number_row, from_=1, to=99999, 
                                         textvariable=self.starting_ticket_number_var, width=10)
        ticket_number_spinbox.pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(ticket_number_row, text="(Used when database is empty)", 
                 font=("Helvetica", 8), foreground="gray").pack(side=tk.LEFT, padx=(10, 0))
        
        # ============== Optional Fields Selection for One-way Ticket ==============
        optional_one_way_frame = ttk.LabelFrame(main_frame, text="Optional Fields - One-way", padding="5")
        optional_one_way_frame.pack(fill="x", pady=2)
        
        # Get current selection from config
        self.one_way_selected_fields = set(self.config.get("one_way_optional_fields", []))
        
        self.one_way_designation_var = tk.BooleanVar(value="Designation" in self.one_way_selected_fields)
        self.one_way_sender_var = tk.BooleanVar(value="Sender" in self.one_way_selected_fields)
        self.one_way_origin_var = tk.BooleanVar(value="Origin" in self.one_way_selected_fields)
        self.one_way_destination_var = tk.BooleanVar(value="Destination" in self.one_way_selected_fields)
        self.one_way_driver_var = tk.BooleanVar(value="Driver" in self.one_way_selected_fields)
        self.one_way_total_price_var = tk.BooleanVar(value="Total Price" in self.one_way_selected_fields)
        
        # Compact single row for all one-way optional fields
        fields_row = ttk.Frame(optional_one_way_frame)
        fields_row.pack(fill="x", padx=0, pady=0)
        ttk.Checkbutton(fields_row, text="Designation", variable=self.one_way_designation_var, command=self._update_one_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row, text="Sender", variable=self.one_way_sender_var, command=self._update_one_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row, text="Origin", variable=self.one_way_origin_var, command=self._update_one_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row, text="Destination", variable=self.one_way_destination_var, command=self._update_one_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row, text="Driver", variable=self.one_way_driver_var, command=self._update_one_way_optional_fields).pack(side="left", padx=2)
        self.one_way_total_price_cb = ttk.Checkbutton(fields_row, text="Total Price", variable=self.one_way_total_price_var, command=self._update_one_way_optional_fields)
        self.one_way_total_price_cb.pack(side="left", padx=2)

        # ============== Optional Fields Selection for Two-way Ticket ==============
        optional_two_way_frame = ttk.LabelFrame(main_frame, text="Optional Fields - Two-way", padding="5")
        optional_two_way_frame.pack(fill="x", pady=2)
        
        # Get current selection from config
        self.two_way_selected_fields = set(self.config.get("two_way_optional_fields", []))
        
        self.two_way_designation_var = tk.BooleanVar(value="Designation" in self.two_way_selected_fields)
        self.two_way_sender_var = tk.BooleanVar(value="Sender" in self.two_way_selected_fields)
        self.two_way_origin_var = tk.BooleanVar(value="Origin" in self.two_way_selected_fields)
        self.two_way_destination_var = tk.BooleanVar(value="Destination" in self.two_way_selected_fields)
        self.two_way_driver_var = tk.BooleanVar(value="Driver" in self.two_way_selected_fields)
        self.two_way_total_price_var = tk.BooleanVar(value="Total Price" in self.two_way_selected_fields)
        
        # Compact single row for all two-way optional fields
        fields_row2 = ttk.Frame(optional_two_way_frame)
        fields_row2.pack(fill="x", padx=0, pady=0)
        ttk.Checkbutton(fields_row2, text="Designation", variable=self.two_way_designation_var, command=self._update_two_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row2, text="Sender", variable=self.two_way_sender_var, command=self._update_two_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row2, text="Origin", variable=self.two_way_origin_var, command=self._update_two_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row2, text="Destination", variable=self.two_way_destination_var, command=self._update_two_way_optional_fields).pack(side="left", padx=2)
        ttk.Checkbutton(fields_row2, text="Driver", variable=self.two_way_driver_var, command=self._update_two_way_optional_fields).pack(side="left", padx=2)
        self.two_way_total_price_cb = ttk.Checkbutton(fields_row2, text="Total Price", variable=self.two_way_total_price_var, command=self._update_two_way_optional_fields)
        self.two_way_total_price_cb.pack(side="left", padx=2)
        
        # Set initial state of total price checkboxes based on price computation setting
        if hasattr(self, 'price_computation_enabled_var'):
            enabled = self.price_computation_enabled_var.get()
            state = "normal" if enabled else "disabled"
            self.one_way_total_price_cb.config(state=state)
            self.two_way_total_price_cb.config(state=state)
            if not enabled:
                self.one_way_total_price_var.set(False)
                self.two_way_total_price_var.set(False)
        
    def build_settings_container_tab(self, tab: ttk.Frame):
        """Builds the container for all settings sub-tabs."""
        for widget in tab.winfo_children():
            widget.destroy()
        
        self.settings_notebook = ttk.Notebook(tab)
        self.settings_notebook.pack(expand=True, fill="both")
        self.settings_notebook.bind("<<NotebookTabChanged>>", self._on_internal_settings_tab_changed)
        
        # Create individual tabs
        general_settings_tab = ttk.Frame(self.settings_notebook)
        print_ticket_settings_tab = ttk.Frame(self.settings_notebook)  # Combined print and ticket settings
        master_data_tab = ttk.Frame(self.settings_notebook)
        customized_tab = ttk.Frame(self.settings_notebook)
        activation_tab = ttk.Frame(self.settings_notebook)
        
        # Add tabs to notebook
        self.settings_notebook.add(master_data_tab, text="Master Data")
        self.settings_notebook.add(customized_tab, text="Customize Entry Form")
        self.settings_notebook.add(general_settings_tab, text="Comm Ports")
        self.settings_notebook.add(activation_tab, text="Admin Settings")
        self.settings_notebook.add(print_ticket_settings_tab, text="Print & Ticket Settings")
        
        # Build content for each tab
        self.build_general_settings_sub_tab(general_settings_tab)
        self.build_print_and_ticket_settings_tab(print_ticket_settings_tab)  # Combined tab
        self.build_master_data_tab(master_data_tab)
        self.build_customized_tab(customized_tab)
        self.build_activation_tab(activation_tab)
        
    def _populate_placeholders_list(self):
        """Populate the placeholders listbox with available placeholders."""
        self.placeholder_listbox.delete(0, tk.END)
        
        # Common placeholders for both templates
        common_placeholders = [
            "{company}", "{ticket_no}", "{truck_plate}", "{product}", 
            "{designation}", "{sender}", "{origin}", "{destination}",
            "{driver}", "{gross_weight}", "{gross_date}", "{gross_time}",
            "{tare_weight}", "{tare_date}", "{tare_time}", "{net_weight}",
            "{weight_type}", "{timestamp}", "{barcode}"
        ]
        
        # Additional placeholders for two-way templates
        two_way_placeholders = [
            "{gross_weight}", "{gross_date}", "{gross_time}",
            "{tare_weight}", "{tare_date}", "{tare_time}",
            "{net_weight}", "{weight_type}"
        ]
        
        # Note: Buttons are created in the calling context where placeholder_frame is defined
                
        # Add all placeholders to the listbox
        for placeholder in common_placeholders:
            self.placeholder_listbox.insert(tk.END, f"[{placeholder}]")

    def _insert_placeholder_from_list(self, event=None):
        """Insert selected placeholder from listbox into template editor."""
        selected_indices = self.placeholder_listbox.curselection()
        if not selected_indices:
            self.msg_box.showwarning("No Selection", "Please select a placeholder from the list to insert.")
            return
        
        selected_text = self.placeholder_listbox.get(selected_indices[0])
        # Extract actual placeholder from bracket format
        placeholder_match = re.search(r"\[(.*)\]", selected_text)
        if placeholder_match:
            placeholder = placeholder_match.group(1)
            self.print_template_text.config(state=tk.NORMAL)
            self.print_template_text.insert(tk.INSERT, f"{{{placeholder}}}")
            self.print_template_text.config(state=tk.DISABLED)
            self._on_template_text_change()

    def _insert_image_placeholder(self):
        """Insert an image placeholder at cursor position."""
        self.print_template_text.config(state=tk.NORMAL)
        self.print_template_text.insert(tk.INSERT, "{image_placeholder}")
        self.print_template_text.config(state=tk.DISABLED)
        self._on_template_text_change()

    def _preview_print_template(self):
        """Generate a preview of the current template."""
        # This would generate a sample preview based on the current template
        # Implementation depends on how you want to simulate the preview
        pass

    def _save_current_template(self):
        """Save the currently edited template."""
        self._on_template_text_change()
        self.save_config()
        self._update_all_previews()
        self.msg_box.showinfo("Template Saved", "Current template saved successfully.")

    def _update_template_text_and_previews(self):
        """Update the template text editor and refresh all previews."""
        current_mode = self.current_template_editor_mode_var.get()
        template_content = ""
        
        if current_mode == "ONE_WAY":
            template_content = self.one_way_print_template_var.get()
        else:  # "TWO_WAY"
            template_content = self.two_way_print_template_var.get()
        
        # Update the text editor
        self.print_template_text.config(state=tk.NORMAL)
        self.print_template_text.delete("1.0", tk.END)
        self.print_template_text.insert("1.0", template_content)
        self.print_template_text.config(state=tk.DISABLED)
        
        # Update preview
        self._update_print_preview()

    def _update_print_preview(self):
        """Update the preview with sample data."""
        try:
            current_mode = self.current_template_editor_mode_var.get()
            template_str = self.one_way_print_template_var.get() if current_mode == "ONE_WAY" else self.two_way_print_template_var.get()
            
            # Sample data for preview
            date_printed_value = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
            sample_data = {
                "company": "SAMPLE CO.",
                "ticket_no": "12345",
                "truck_plate": "ABC 123",
                "product": "GRAIN",
                "designation": "BULK",
                "sender": "FARMER JOE",
                "origin": "FARM #1",
                "destination": "WAREHOUSE A",
                "driver": "JOHN DOE",
                "gross_weight": "25000.00",
                "gross_date": "2023-10-27",
                "gross_time": "10:30 AM",
                "tare_weight": "15000.00",
                "tare_date": "2023-10-27",
                "tare_time": "11:15 AM",
                "net_weight": "10000.00",
                "weight_type": "ONE WAY WEIGHING",
                "timestamp": "2023-10-27 10:30:00",
                "date_printed": date_printed_value,
                "logged_in_user": self.logged_in_user if self.logged_in_user else "admin"
            }
            
            # Replace placeholders with sample data
            formatted_content = template_str.format(**sample_data)
            
            # Apply line spacing if needed
            line_spacing = self.print_line_spacing.get()
            if line_spacing > 0:
                formatted_content = formatted_content.replace("\n", "\n" * (1 + line_spacing))
            
            # Update preview
            self.print_preview_text.config(state="normal")
            self.print_preview_text.delete("1.0", tk.END)
            self.print_preview_text.insert("1.0", formatted_content)
            self.print_preview_text.config(state="disabled")
            
        except Exception as e:
            self.print_preview_text.config(state="normal")
            self.print_preview_text.delete("1.0", tk.END)
            self.print_preview_text.insert("1.0", f"Preview Error: {str(e)}")
            self.print_preview_text.config(state="disabled")

    def _on_template_text_change(self, event=None):
        """Event handler for changes in the print template text editor."""
        current_text_content = self.print_template_text.get("1.0", tk.END).strip()
        
        # Save to the correct variable based on the selected radio button
        if self.current_template_editor_mode_var.get() == "ONE_WAY":
            self.one_way_print_template_var.set(current_text_content)
        else:  # "TWO_WAY"
            self.two_way_print_template_var.set(current_text_content)
        
        self.print_template_text.edit_modified(False)
        self._debounce_template_updates()

    def _debounce_template_updates(self):
        """Debounces and schedules a function to update the previews and save the templates."""
        self.root.after_cancel(getattr(self, '_template_update_id', ''))
        self._template_update_id = self.root.after(1000, self._update_and_save_templates)

    def _update_and_save_templates(self):
        """Updates the previews and saves the current templates to the database."""
        self._update_all_previews()
        self._save_templates_to_db()

    def _save_templates_to_db(self):
        """Saves the print templates from the in-memory variables to the database."""
        try:
            with sqlite3.connect(DB_FILE) as conn:
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'ONE_WAY'",
                            (self.one_way_print_template_var.get(),))
                conn.execute("UPDATE print_templates SET template_content = ? WHERE template_name = 'TWO_WAY'",
                            (self.two_way_print_template_var.get(),))
                conn.commit()
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Save Error", f"Failed to save print templates to database: {e}")


    def build_print_template_sub_tab(self, tab: ttk.Frame):
        """Builds the Print Template Editor tab."""
        for widget in tab.winfo_children():
            widget.destroy()
        
        main_frame = ttk.Frame(tab, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Left frame for template selection and editor
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 10))
        
        # Right frame for preview and controls
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill="both", expand=True)
        
        # Template selection buttons
        template_frame = ttk.Frame(left_frame)
        template_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Button(template_frame, text="Load ONE WAY Template", 
                   command=lambda: self._load_template("ONE_WAY")).pack(side=tk.LEFT, padx=5)
        ttk.Button(template_frame, text="Load TWO WAY Template", 
                   command=lambda: self._load_template("TWO_WAY")).pack(side=tk.LEFT, padx=5)
        
        # Editor frame
        editor_frame = ttk.LabelFrame(left_frame, text="Template Editor")
        editor_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # Text widget for template editing
        self.print_template_text = tk.Text(editor_frame, wrap="none", undo=True, 
                                           font=("Courier New", 10), height=15)
        self.print_template_text.pack(expand=True, fill="both")
        
        # Scrollbars for editor
        h_scroll = ttk.Scrollbar(editor_frame, orient="horizontal", command=self.print_template_text.xview)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.print_template_text.config(xscrollcommand=h_scroll.set)
        
        v_scroll = ttk.Scrollbar(editor_frame, orient="vertical", command=self.print_template_text.yview)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.print_template_text.config(yscrollcommand=v_scroll.set)
        
        # Bind text changes
        self.print_template_text.bind("<KeyRelease>", lambda event: self._on_template_text_change())
        
        # Preview frame
        preview_frame = ttk.LabelFrame(right_frame, text="Preview")
        preview_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # Preview text widget
        self.print_preview_text = tk.Text(preview_frame, wrap="word", height=10)
        self.print_preview_text.pack(expand=True, fill="both")
        self.print_preview_text.config(state="disabled")
        
        # Font and formatting controls
        font_frame = ttk.LabelFrame(right_frame, text="Font and Formatting", padding="10")
        font_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(font_frame, text="Font:").grid(row=0, column=0, sticky="w")
        self.font_family_combobox = ttk.Combobox(font_frame, textvariable=self.print_template_font_family_var, 
                                                values=self.monospace_fonts, state="readonly", width=15)
        self.font_family_combobox.grid(row=0, column=1, sticky="ew", padx=2)
        self.font_family_combobox.bind("<<ComboboxSelected>>", lambda e: self._update_all_previews())
        
        ttk.Label(font_frame, text="Size:").grid(row=1, column=0, sticky="w")
        self.font_size_spinbox = ttk.Spinbox(font_frame, from_=8, to=24, 
                                            textvariable=self.print_template_font_size_var, width=5)
        self.font_size_spinbox.grid(row=1, column=1, sticky="w", padx=2)
        self.font_size_spinbox.bind("<Return>", lambda e: self._update_all_previews())
        
        ttk.Checkbutton(font_frame, text="Bold", variable=self.print_template_font_bold_var,
                       command=self._update_all_previews).grid(row=2, column=0, columnspan=2, sticky="w")
        
        # Action buttons
        button_frame = ttk.Frame(right_frame)
        button_frame.pack(fill="x", pady=(10, 0))
        
        ttk.Button(button_frame, text="Save Template", command=self._save_current_template).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Reset to Default", command=self._reset_print_format_to_default).pack(side=tk.LEFT, padx=5)
        
        # Initialize with default templates
        self._update_template_text_and_previews()

    def _load_template(self, template_type):
        """Load a predefined template into the editor."""
        if template_type == "ONE_WAY":
            self.template_text_widget.delete("1.0", tk.END)
            self.template_text_widget.insert("1.0", self._get_one_way_template())
        elif template_type == "TWO_WAY":
            self.template_text_widget.delete("1.0", tk.END)
            self.template_text_widget.insert("1.0", self._get_two_way_template())

    def _save_current_template(self):
        """Save the current template content to the database."""
        try:
            template_content = self.template_text_widget.get("1.0", tk.END)
            template_name = "CUSTOM_TEMPLATE_" + datetime.now().strftime("%Y%m%d_%H%M%S")
            
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                # Save the template content to the database
                cursor.execute("""
                    INSERT OR REPLACE INTO print_templates (template_name, template_content) 
                    VALUES (?, ?)
                """, (template_name, template_content))
                conn.commit()
            
            self.msg_box.showinfo("Success", "Template saved successfully!")
            
            # Update the preview after saving
            self._update_all_previews()
            
        except Exception as e:
            self.msg_box.showerror("Save Error", f"Failed to save template: {e}")

    def _reset_print_format_to_default(self):
        """Reset the print format to default."""
        self.template_text_widget.delete("1.0", tk.END)
        self.template_text_widget.insert("1.0", self._get_default_template())
        
        # Update preview after reset
        self._update_all_previews()

    def _get_default_template(self):
        """Get the default template based on current mode."""
        if self.current_template_editor_mode_var.get() == "ONE_WAY":
            return self._get_one_way_template()
        else:
            return self._get_two_way_template()
            
    def _initialize_database_tables(self):
        """Ensures required database tables exist."""
        try:
            with sqlite3.connect(DB_FILE) as conn:
                # Create print_templates table if it doesn't exist
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS print_templates (
                        template_name TEXT PRIMARY KEY,
                        template_content TEXT NOT NULL
                    )
                """)
                
                # Insert empty templates if they don't exist (admin will define them)
                conn.execute("""
                    INSERT OR IGNORE INTO print_templates (template_name, template_content) 
                    VALUES (?, ?)
                """, ("ONE_WAY", ""))
                
                conn.execute("""
                    INSERT OR IGNORE INTO print_templates (template_name, template_content) 
                    VALUES (?, ?)
                """, ("TWO_WAY", ""))
                
                conn.commit()
        except sqlite3.Error as e:
            self.msg_box.showerror("Database Error", f"Failed to initialize database tables: {e}")
      
    
    def _on_entry_field_selection_changed(self):
        # A placeholder function for future use
        pass

    def _on_internal_settings_tab_changed(self, event: tk.Event):
        """
        Handles actions to perform when a user switches tabs in the settings notebook.
        """
        selected_tab_widget = event.widget
        selected_tab_text = selected_tab_widget.tab(selected_tab_widget.select(), "text")

        if selected_tab_text == "Print Templates":
            self._update_template_text_and_previews()
        elif selected_tab_text == "Master Data":
            self.load_master_data_lists()
        elif selected_tab_text == "Print Settings":
            self._populate_printer_list()
        elif selected_tab_text == "Print & Ticket Settings":
            # Populate printer list when accessing this combined tab
            self._populate_printer_list()
    
    def build_print_template_sub_tab(self, tab: ttk.Frame):
        """
        Builds the Print Template Editor tab.
        """
        for widget in tab.winfo_children():
            widget.destroy()

        main_frame = ttk.Frame(tab, padding="10")
        main_frame.pack(fill="both", expand=True)

        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 10))

        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill="both", expand=True, padx=(10, 0))

        # --- Left Frame: Editor ---
        editor_frame = ttk.LabelFrame(left_frame, text="Print Template Editor", padding="10")
        editor_frame.pack(fill="both", expand=True)

        radio_frame = ttk.Frame(editor_frame)
        radio_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 5))
        
        ttk.Radiobutton(radio_frame, text="One Way Weighing", variable=self.current_template_editor_mode_var,
                        value="ONE_WAY", command=self._update_print_template_editor_content).pack(side=tk.LEFT, padx=5)
        
        self.two_way_template_radio = ttk.Radiobutton(radio_frame, text="Two Way Weighing", variable=self.current_template_editor_mode_var,
                        value="TWO_WAY", command=self._update_print_template_editor_content)
        self.two_way_template_radio.pack(side=tk.LEFT, padx=5)

        self.print_template_text = tk.Text(editor_frame, wrap="none", undo=True, font=("Courier New", 10), height=15)
        self.print_template_text.pack(expand=True, fill="both")
        self.print_template_text.bind("<KeyRelease>", lambda event: self._on_template_text_change())

        h_scroll = ttk.Scrollbar(editor_frame, orient="horizontal", command=self.print_template_text.xview)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.print_template_text.config(xscrollcommand=h_scroll.set)

        v_scroll = ttk.Scrollbar(editor_frame, orient="vertical", command=self.print_template_text.yview)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.print_template_text.config(yscrollcommand=v_scroll.set)

        # --- Right Frame: Placeholders, Preview, and Controls ---
        font_frame = ttk.LabelFrame(right_frame, text="Font and Formatting", padding="10")
        font_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(font_frame, text="Font:").grid(row=0, column=0, sticky="w")
        ttk.Combobox(font_frame, textvariable=self.print_template_font_family_var, values=self.monospace_fonts,
                     state="readonly", width=15, style="TCombobox").grid(row=0, column=1, sticky="ew", padx=2)

        ttk.Label(font_frame, text="Size:").grid(row=1, column=0, sticky="w")
        ttk.Spinbox(font_frame, from_=6, to=20, textvariable=self.print_template_font_size_var, width=5,
                    command=self._update_template_font).grid(row=1, column=1, sticky="w", padx=2)
        self.print_template_font_size_var.trace_add("write", lambda name, index, mode: self._update_template_font())
        
        ttk.Checkbutton(font_frame, text="Bold", variable=self.print_template_font_bold_var,
                        command=self._update_template_font).grid(row=1, column=2, sticky="w", padx=2)

        font_frame.grid_columnconfigure(1, weight=1)

        options_frame = ttk.LabelFrame(right_frame, text="Print Options", padding="10")
        options_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(options_frame, text="Line Spacing:").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(options_frame, from_=0, to=3, textvariable=self.print_line_spacing, width=5,
                    command=self._debounce_template_updates).grid(row=0, column=1, sticky="ew", padx=2)
        self.print_line_spacing.trace_add("write", lambda name, index, mode: self._debounce_template_updates())

        ttk.Label(options_frame, text="Encoding:").grid(row=1, column=0, sticky="w")
        ttk.Combobox(options_frame, textvariable=self.print_encoding_var,
                     values=["utf-8", "cp437", "shift_jis"],
                     state="readonly", style="TCombobox").grid(row=1, column=1, sticky="ew", padx=2)

        ttk.Checkbutton(options_frame, text="Include Barcode", variable=self.print_include_barcode_var,
                        command=self._debounce_template_updates).grid(row=2, column=0, columnspan=2, sticky="w")

        options_frame.grid_columnconfigure(1, weight=1)
        
        placeholder_frame = ttk.LabelFrame(right_frame, text="Available Placeholders", padding="10")
        placeholder_frame.pack(fill="both", expand=True, pady=(0, 10))

        self.placeholder_listbox = tk.Listbox(placeholder_frame, exportselection=False, font=("Courier New", 10))
        self.placeholder_listbox.pack(side=tk.LEFT, fill="both", expand=True)

        for name, placeholder in self.available_print_placeholders:
            self.placeholder_listbox.insert(tk.END, f"{name}: {placeholder}")
        
        placeholder_vsb = ttk.Scrollbar(placeholder_frame, orient="vertical", command=self.placeholder_listbox.yview)
        placeholder_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.placeholder_listbox.config(yscrollcommand=placeholder_vsb.set)
        
        self.placeholder_listbox.bind("<Double-1>", self._insert_placeholder_from_event)
        
        ttk.Button(placeholder_frame, text="Insert Selected", command=self._insert_placeholder_into_template_btn_click, style="TButton").pack(fill=tk.X, pady=(5,0))

        # --- Right Frame: Preview and Controls ---
        preview_frame = ttk.LabelFrame(right_frame, text="Template Preview (Raw Text)", padding="10")
        preview_frame.pack(fill="both", expand=True)

        self.print_preview_text = tk.Text(preview_frame, wrap="none", state="disabled", font=("Courier New", 10), height=10)
        self.print_preview_text.pack(expand=True, fill="both")
        
        preview_h_scroll = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.print_preview_text.xview)
        preview_h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.print_preview_text.config(xscrollcommand=preview_h_scroll.set)

        preview_v_scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=self.print_preview_text.yview)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.print_preview_text.config(yscrollcommand=preview_v_scroll.set)

        button_frame = ttk.Frame(right_frame)
        button_frame.pack(fill="x", pady=10)
        
        self.test_print_button = ttk.Button(button_frame, text="Test Print", command=self._test_print, style="TButton")
        self.test_print_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))
        
        ttk.Button(button_frame, text="Reset to Default", command=self._reset_print_format_to_default, style="TButton").pack(side=tk.RIGHT, expand=True, fill=tk.X, padx=(5, 0))
        
        self._update_template_text_and_previews()
# Part 12: Emulator, Big Display, and Main Execution

    def _update_emulator_button_states(self):
        """
        Updates the enabled/disabled state of emulator buttons and its status label.
        """
        if hasattr(self, '_emulator_connect_button') and hasattr(self, '_emulator_disconnect_button') and hasattr(self, '_emulator_status_var'):
            if self._emulator_serial_port and self._emulator_serial_port.is_open:
                self._emulator_connect_button.config(state=tk.DISABLED)
                self._emulator_disconnect_button.config(state=tk.NORMAL)
                self._emulator_start_send_button.config(state=tk.NORMAL if not self._emulator_sending_data and self.is_app_activated else tk.DISABLED)
                self._emulator_stop_send_button.config(state=tk.NORMAL if self._emulator_sending_data else tk.DISABLED)
                self._emulator_port_entry.config(state=tk.DISABLED)
                self._emulator_baud_entry.config(state=tk.DISABLED)
                self._emulator_status_var.set(f"Emulator Status: Connected to {self._emulator_port_entry_var.get()}")
                if self._emulator_sending_data:
                    self._emulator_status_var.set(f"Emulator Status: Sending data to {self._emulator_port_entry_var.get()}...")
            else:
                self._emulator_connect_button.config(state=tk.NORMAL)
                self._emulator_disconnect_button.config(state=tk.DISABLED)
                self._emulator_start_send_button.config(state=tk.DISABLED)
                self._emulator_stop_send_button.config(state=tk.DISABLED)
                self._emulator_port_entry.config(state=tk.NORMAL)
                self._emulator_baud_entry.config(state=tk.NORMAL)
                self._emulator_status_var.set("Emulator Status: Disconnected")
                self._emulator_sending_data = False

    def _emulator_connect_serial(self, port: str, baud_rate: int):
        """
        Connects to the emulator serial port.
        """
        if self.serial_running:
            self.stop_serial_connection()
            self.msg_box.showinfo("Info", "Main scale connection stopped to prioritize emulator connection.")

        if self._emulator_serial_port and self._emulator_serial_port.is_open:
            self._emulator_disconnect_serial()

        if not port or not baud_rate:
            self.msg_box.showerror("Error", "Please select an Emulator COM Port and Baud Rate.")
            return

        try:
            baud_rate = int(baud_rate)
            self._emulator_serial_port = serial.Serial(port, baudrate=baud_rate, timeout=0.01)
            self.big_display_connected = True
            if self.current_edit_transaction_id is None and self.is_app_activated:
                self.root.after(0, lambda: self._update_weight_display_and_status("0.00", "Emulator", "blue"))
        except ValueError:
            self.msg_box.showerror("Error", "Emulator Baud Rate must be a number.")
        except serial.SerialException as e:
            self.msg_box.showerror("Emulator Serial Error", f"Could not open emulator serial port {port}: {e}\n\n"
                                 "Ensure the port is correct and not in use by another application.")
            if self.current_edit_transaction_id is None:
                self.root.after(0, lambda: self._update_weight_display_and_status("ERROR", "Error", "red"))
        except Exception as e:
            self.msg_box.showerror("Emulator Error", f"An unexpected error occurred: {e}")
            if self.current_edit_transaction_id is None:
                self.root.after(0, lambda: self._update_weight_display_and_status("ERROR", "Error", "red"))
        finally:
            self._update_emulator_button_states()

    def _start_emulator_connection_from_settings(self):
        """
        Connects to the emulator from the settings tab.
        """
        port = self._emulator_port_entry_var.get()
        baud_rate = self._emulator_baud_entry_var.get()
        self._emulator_connect_serial(port, baud_rate)

    def _emulator_disconnect_serial(self):
        """
        Disconnects from the emulator serial port.
        """
        if self._emulator_sending_data:
            self._emulator_sending_data = False
        if self._emulator_serial_port and self._emulator_serial_port.is_open:
            self._emulator_serial_port.close()
        self._emulator_serial_port = None
        if self.current_edit_transaction_id is None:
            self._update_weight_display_and_status(f"%.{self.decimal_places}f" % 0.00, "No Source", "gray")
        self._update_emulator_button_states()

    def _emulator_send_data(self, data_str: str):
        """
        Sends a single line of data to the emulator port.
        """
        if self._emulator_serial_port and self._emulator_serial_port.is_open:
            try:
                self._emulator_serial_port.write(f"{data_str}\r\n".encode('utf-8'))
                format_string = f"%.{self.decimal_places}f"
                # Update the main weight display (weight_value) with the emulator's sent weight
                self.root.after(0, lambda w=float(data_str): self.weight_value.set(format_string % w))
                if self.current_edit_transaction_id is None and "Loaded Gross" not in self.weight_source_status.get() and "EXPIRED" not in self.weight_source_status.get():
                    self.root.after(0, lambda: self._update_weight_display_and_status(self.weight_value.get(), "Emulator", "blue"))
            except serial.SerialException as e:
                self.root.after(0, self._emulator_disconnect_serial)
            except ValueError:
                self.root.after(0, lambda: self.msg_box.showerror("Emulator Send Error", f"Emulator tried to send non-numeric data: {data_str}"))
            except Exception as e:
                self.root.after(0, lambda: self.msg_box.showerror("Emulator Send Error", f"An unexpected error occurred during emulator send: {e}"))
        else:
            pass

    def _emulator_continuous_send_loop(self):
        """
        Threaded loop for continuously sending weight data from the emulator.
        """
        while self._emulator_sending_data and self._emulator_serial_port and self._emulator_serial_port.is_open:
            weight_str = self._emulator_weight_to_send_var.get().strip()
            try:
                float(weight_str)
                self._emulator_send_data(weight_str)
            except ValueError:
                self._emulator_sending_data = False
                self.root.after(0, lambda: self.msg_box.showwarning("Emulator Input Error", "Invalid weight value for emulator. Stopping continuous send."))

            try:
                interval = float(self._emulator_send_interval_var.get())
                if interval <= 0:
                    interval = 0.1
            except ValueError:
                interval = 0.1

            time.sleep(interval)
        self.root.after(0, self._update_emulator_button_states)

    def _emulator_start_continuous_sending(self):
        """
        Starts the continuous sending thread for the emulator.
        """
        if not self.is_app_activated:
            self.msg_box.showerror("Activation Required", "Please activate the application to use the emulator.")
            return

        if not self._emulator_serial_port or not self._emulator_serial_port.is_open:
            self.msg_box.showwarning("Emulator Not Connected", "Please connect the emulator to a serial port first.")
            return

        try:
            float(self._emulator_weight_to_send_var.get())
            interval = float(self._emulator_send_interval_var.get())
            if interval <= 0:
                raise ValueError("Interval must be positive.")
        except ValueError as e:
            self.msg_box.showerror("Invalid Input", f"Please enter valid numbers for emulator weight and interval: {e}")
            return
        
        if self.serial_running:
            self.msg_box.showwarning("Conflict", "Main scale is active. Please stop the main scale connection before starting the emulator.")
            return

        if not self._emulator_sending_data:
            self._emulator_sending_data = True
            self._emulator_send_thread = threading.Thread(target=self._emulator_continuous_send_loop, daemon=True)
            self._emulator_send_thread.start()
            self._update_emulator_button_states()

    def _emulator_stop_continuous_sending(self):
        """
        Stops the continuous sending from the emulator.
        """
        if self._emulator_sending_data:
            self._emulator_sending_data = False
            self._update_emulator_button_states()
    
    def _update_big_display_button_states(self):
        """
        Updates the state of the Big Display connect/disconnect buttons
        based on whether the required fields are filled and the connection status.
        """
        # Check if the necessary UI elements exist before trying to access them
        # This prevents AttributeError if called before GUI is fully built
        if not hasattr(self, 'big_display_port_entry') or not hasattr(self, 'big_display_baud_entry'):
            # Log a warning or just return if needed, but don't crash
            # print("Warning: Big Display UI elements not initialized yet.")
            return
            
    def _connect_big_display(self):
        """Connects to the big display serial port."""
        if self.big_display_connected:
            return
        
        port = self.big_display_port_var.get()
        baud_rate = self.big_display_baud_var.get()
        
        if not port or not baud_rate:
            self.msg_box.showerror("Error", "Please select a Big Display COM Port and Baud Rate.")
            return
            
        try:
            baud_rate = int(baud_rate)
            print(f"Attempting to connect to {port} at {baud_rate} baud")
            self.big_display_serial_port = serial.Serial(port, baudrate=baud_rate, timeout=0)
            
            # Set the connected flag
            self.big_display_connected = True
            
            # Update status label
            self.root.after(0, lambda: self.big_display_status_var.set("Big Display Status: Connected"))
            
            # Start sending data loop
            self.big_display_send_thread = threading.Thread(target=self._big_display_continuous_send_loop, daemon=True)
            self.big_display_send_thread.start()
            
            logging.info(f"SUCCESS: Connected to Big Display at {port} baud {baud_rate}")
            
            # Explicitly update button states
            self._update_big_display_button_states()
            
        except ValueError as e:
            self.msg_box.showerror("Big Display Error", f"Baud Rate must be a number: {e}")
        except serial.SerialException as e:
            self.msg_box.showerror("Big Display Error", f"Could not open big display serial port {port}: {e}")
            logging.error(f"Serial Exception connecting to Big Display: {e}")
        except Exception as e:
            self.msg_box.showerror("Big Display Error", f"An unexpected error occurred: {e}")
            logging.error(f"Unexpected error connecting to Big Display: {e}")
    
    def _update_big_display_button_states(self):
        """Updates the state of the Big Display connect/disconnect buttons based on whether the required fields are filled and the connection status."""
        # Check if the necessary UI elements exist
        if not hasattr(self, 'big_display_connect_button') or not hasattr(self, 'big_display_disconnect_button'):
            return

        # Debug output
        print(f"Update Big Display Button States - Connected: {self.big_display_connected}, Port: {self.big_display_port_var.get()}, Baud: {self.big_display_baud_var.get()}")

        # Enable/disable buttons based on connection status and field validity
        if self.big_display_connected and self.big_display_port_var.get() and self.big_display_baud_var.get():
            self.big_display_connect_button.config(state=tk.DISABLED)
            self.big_display_disconnect_button.config(state=tk.NORMAL)
        else:
            self.big_display_connect_button.config(state=tk.NORMAL)
            self.big_display_disconnect_button.config(state=tk.DISABLED)
        
    # Add this new method to handle continuous sending
    def _big_display_continuous_send_loop(self):
        """Continuously sends weight data to the big display."""
        while self.big_display_connected and self.big_display_serial_port and self.big_display_serial_port.is_open:
            try:
                # Get the current weight value from the main display
                current_weight = self.weight_value.get()
                
                # Convert to float to check if it's valid
                try:
                    weight_float = float(current_weight)
                    # Format as 6-digit zero-padded string
                    weight_str = f"{int(weight_float):06d}"
                    # Reverse the digits
                    reversed_digits = weight_str[::-1]
                    # Create payload
                    payload = f"={reversed_digits}"
                    
                    # Strip whitespace
                    payload = payload.strip()
                    
                    # Send data
                    bytes_written = self.big_display_serial_port.write(payload.encode('ascii'))
                    self.big_display_serial_port.flush()
                    
                    if bytes_written > 0:
                        logging.info(f"Sent to Big Display: '{payload}' ({bytes_written} bytes)")
                    else:
                        logging.warning("No bytes written to Big Display")
                        
                except ValueError:
                    # If weight is not a valid number, send zeros
                    payload = "=000000"
                    bytes_written = self.big_display_serial_port.write(payload.encode('ascii'))
                    self.big_display_serial_port.flush()
                    logging.debug("Sent zero padding to Big Display")
                    
            except serial.SerialException as e:
                logging.error(f"Serial error sending to Big Display: {e}")
                self.root.after(0, self._disconnect_big_display)
            except Exception as e:
                logging.error(f"Unexpected error in Big Display loop: {e}")
                
            # Small delay to prevent excessive CPU usage
            time.sleep(0.3)

    def _disconnect_big_display(self):
        """Disconnects from the big display serial port."""
        self.big_display_connected = False
        if self.big_display_serial_port and self.big_display_serial_port.is_open:
            try:
                self.big_display_serial_port.close()
                self.big_display_serial_port = None
            except Exception as e:
                logging.error(f"Error closing Big Display port: {e}")
        
        # Stop the sending thread
        if hasattr(self, 'big_display_send_thread') and self.big_display_send_thread:
            self.big_display_send_thread.join(timeout=1)
        
        self.root.after(0, lambda: self.big_display_status_var.set("Big Display Status: Disconnected"))
        self._update_big_display_button_states()

    def _send_weight_to_big_display(self, weight: float, tare_date: str, tare_time: str):
        """
        Sends formatted weight data to the big display.
        """
        if self.big_display_connected and self.big_display_serial_port and self.big_display_serial_port.is_open:
            current_data_tuple = (weight, tare_date, tare_time)
            if self.last_sent_weight == current_data_tuple:
                return
            try:
                output_format = self.big_display_data_format_var.get()
                data_to_send = output_format.format(weight=weight, tare_date=tare_date, tare_time=tare_time)
                data_to_send = data_to_send.replace("\\r", "\r").replace("\\n", "\n")

                bytes_to_send = data_to_send.encode('ascii', errors='ignore')

                self.big_display_serial_port.write(bytes_to_send)
                self.last_sent_weight = current_data_tuple
            except serial.SerialException as e:
                self.root.after(0, self._disconnect_big_display)
            except Exception as e:
                pass
        else:
            pass

    def _auto_refresh_ui_elements(self):
        """
        A background thread to periodically refresh UI elements that need constant updates.
        """
        while True:
            self.root.after(0, self._update_emulator_button_states)
            self.root.after(0, self._update_big_display_button_states)
            self.root.after(0, self._update_activation_status_label_on_gui)
            self.root.after(0, self._update_action_button_states)
            self.root.after(0, lambda: self.auto_time_var.set(datetime.now().strftime("%I:%M:%S %p")))

            try:
                time.sleep(1)
            except Exception as e:
                break

    # Camera Management Methods
    def _refresh_camera_list(self):
        """Refresh the list of available cameras."""
        try:
            # Check if camera_device_combobox exists before proceeding
            if not hasattr(self, 'camera_device_combobox'):
                return  # Silently return if combobox doesn't exist yet
                
            if self.camera_manager:
                available_cameras = self.camera_manager.get_available_cameras()
                self.camera_device_combobox['values'] = available_cameras
                
                # If current selection is not in the list, clear it
                current_selection = self.camera_device_var.get()
                if current_selection and current_selection not in available_cameras:
                    self.camera_device_var.set("")
                    
                # Don't show info message to avoid OK button before login
                # self.msg_box.showinfo("Camera List", f"Found {len(available_cameras)} camera(s)")
            else:
                self.msg_box.showerror("Camera Error", "Camera manager not initialized")
        except Exception as e:
            self.msg_box.showerror("Camera Error", f"Failed to refresh camera list: {e}")
    
    def _toggle_camera_connection(self):
        """Toggle camera connection on/off."""
        try:
            if not self.camera_manager:
                self.msg_box.showerror("Camera Error", "Camera manager not initialized")
                return
            
            if self.camera_connected:
                # Disconnect camera
                self.camera_manager.disconnect_camera()
                self.camera_connected = False
                self.camera_status_var.set("Camera Status: Disconnected")
                self.camera_connect_button.config(text="📷 Connect")
                
                # Automatically switch to uploaded image if available
                if self.camera_uploaded_image_path and os.path.exists(self.camera_uploaded_image_path):
                    self.camera_use_image_var.set(True)
                    self._on_camera_image_toggle()
                    self.camera_status_var.set("Camera Status: Using Uploaded Image (Camera Disconnected)")
                else:
                    self.camera_status_var.set("Camera Status: Disconnected")
            else:
                # Connect camera
                selected_camera = self.camera_device_var.get()
                if not selected_camera:
                    self.msg_box.showerror("Camera Error", "Please select a camera device")
                    return
                
                # Extract camera index from selection (e.g., "Camera 0" -> 0)
                try:
                    camera_index = int(selected_camera.split()[-1])
                except (ValueError, IndexError):
                    self.msg_box.showerror("Camera Error", "Invalid camera selection")
                    return
                
                if self.camera_manager.connect_camera(camera_index):
                    self.camera_connected = True
                    self.camera_status_var.set("Camera Status: Connected")
                    self.camera_connect_button.config(text="🔌 Disconnect")
                    
                    # Automatically switch to camera if using uploaded image
                    if self.camera_use_image_var.get():
                        self.camera_use_image_var.set(False)
                        self._on_camera_image_toggle()
                    
                    # Set mirror mode
                    self.camera_manager.set_mirror_mode(self.camera_mirror_mode_var.get())
                    
                    # Start capture
                    self.camera_manager.start_capture()
                else:
                    self.msg_box.showerror("Camera Error", "Failed to connect to camera")
                    
        except Exception as e:
            self.msg_box.showerror("Camera Error", f"Error toggling camera connection: {e}")
    
    def _on_camera_image_toggle(self):
        """Handle camera image toggle switch."""
        if self.camera_use_image_var.get():
            # Load the uploaded image if path exists
            if self.camera_uploaded_image_path:
                self._load_uploaded_image()
            # Update status
            if self.camera_connected:
                self.camera_status_var.set("Camera Status: Using Uploaded Image (Camera Connected)")
            else:
                self.camera_status_var.set("Camera Status: Using Uploaded Image")
        else:
            # Clear uploaded image
            self.camera_uploaded_image = None
            # Update status back to camera status
            if self.camera_connected:
                self.camera_status_var.set("Camera Status: Connected")
            else:
                self.camera_status_var.set("Camera Status: Disconnected")
        
        # Auto-save camera settings
        self.save_config()
    
    def _on_camera_mirror_toggle(self):
        """Handle camera mirror mode toggle."""
        if self.camera_manager:
            self.camera_manager.set_mirror_mode(self.camera_mirror_mode_var.get())
        
        # If using uploaded image, reload it with mirror mode applied
        if self.camera_use_image_var.get() and self.camera_uploaded_image:
            self._load_uploaded_image()
        
        # Auto-save camera settings
        self.save_config()
    
    def _browse_camera_image(self):
        """Browse for an image file to upload."""
        from tkinter import filedialog
        
        file_types = [
            ("Image Files", "*.jpg *.jpeg *.png *.bmp *.gif *.tiff"),
            ("JPEG Files", "*.jpg *.jpeg"),
            ("PNG Files", "*.png"),
            ("Bitmap Files", "*.bmp"),
            ("All Files", "*.*")
        ]
        
        filename = filedialog.askopenfilename(
            title="Select Camera Image",
            filetypes=file_types
        )
        
        if filename:
            self.camera_image_path_var.set(filename)
            self.camera_uploaded_image_path = filename
            self._load_uploaded_image()
            # Auto-save camera settings
            self.save_config()
    
    def _load_uploaded_image(self):
        """Load the uploaded image and display it."""
        try:
            if self.camera_uploaded_image_path and os.path.exists(self.camera_uploaded_image_path):
                # Load image using PIL
                self.camera_uploaded_image = Image.open(self.camera_uploaded_image_path)
                
                # Resize image to fit camera display area (640x480)
                self.camera_uploaded_image = self.camera_uploaded_image.resize((640, 480), Image.Resampling.LANCZOS)
                
                # Apply mirror mode if enabled
                if self.camera_mirror_mode_var.get():
                    self.camera_uploaded_image = self.camera_uploaded_image.transpose(Image.FLIP_LEFT_RIGHT)
                
                # Convert to PhotoImage and display
                photo = ImageTk.PhotoImage(self.camera_uploaded_image)
                self._update_camera_frame(photo, self.camera_uploaded_image)
                
                # Don't show success message to avoid OK button during operation
                # self.msg_box.showinfo("Success", "Camera image loaded successfully!")
            else:
                self.msg_box.showerror("Error", "Image file not found")
                self.camera_uploaded_image = None
        except Exception as e:
            self.msg_box.showerror("Error", f"Failed to load image: {e}")
            self.camera_uploaded_image = None
    
    def _update_camera_frame(self, photo, pil_image=None):
        """Update camera frame display.
        
        Args:
            photo: PhotoImage object from camera
            pil_image: PIL Image object for resizing
        """
        try:
            if hasattr(self, 'camera_label') and self.camera_label:
                self.camera_label.update_idletasks()
                label_width = self.camera_label.winfo_width()
                label_height = self.camera_label.winfo_height()

                if label_width > 1 and label_height > 1 and pil_image:
                    resized_image = pil_image.resize((label_width, label_height), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(resized_image)
                
                self.camera_label.config(image=photo)
                self.camera_label.image = photo  # Keep a reference
        except Exception as e:
            logging.error(f"Error updating camera frame: {e}")

    def _sync_camera_feed_width(self):
        if not hasattr(self, 'camera_feed_frame') or not hasattr(self, 'save_button'):
            return

        def apply_width():
            if not self.camera_feed_frame.winfo_exists() or not self.save_button.winfo_exists():
                return
            self.camera_feed_frame.update_idletasks()
            self.save_button.update_idletasks()
            target_width = self.save_button.winfo_width()
            if target_width > 1:
                self.camera_feed_frame.grid_propagate(False)
                self.camera_feed_frame.configure(width=target_width)
                self.camera_label.configure(width=target_width)

        self.root.after(0, apply_width)
        self.root.after(250, apply_width)
    
    def cleanup_camera(self):
        """Clean up camera resources."""
        try:
            if self.camera_manager:
                self.camera_manager.cleanup()
                self.camera_connected = False
                self.camera_status_var.set("Camera Status: Disconnected")
        except Exception as e:
            logging.error(f"Error cleaning up camera: {e}")

    def on_closing(self):
        """
        Saves the config and gracefully closes all serial connections before the application exits.
        The template content is now saved directly from the UI, so we only need to save the config.
        """
        # Only attempt to save config if the main GUI elements are initialized
        # This prevents errors if the app is closed before login completes
        if hasattr(self, 'pdf_print_template_font_family_var'):
            self.config["pdf_print_template_font_family"] = self.pdf_print_template_font_family_var.get()
            self.config["pdf_print_template_font_size"] = self.pdf_print_template_font_size_var.get()
            self.config["pdf_print_template_font_bold"] = self.pdf_print_template_font_bold_var.get()
            # NEW: Save PDF page size and orientation on closing
            self.config["pdf_page_size"] = self.pdf_page_size_var.get()
            self.config["pdf_orientation"] = self.pdf_orientation_var.get()
            # Comprehensive save of all settings
            self.save_config()

        self.stop_serial_connection()
        self._emulator_disconnect_serial()
        self._disconnect_big_display()
        
        # Clean up camera
        self.cleanup_camera()

        self.root.destroy()

    # Utility methods for enhanced search functionality
    def _debounced_pending_search(self, event=None):
        """Debounced search for pending records to avoid excessive database queries."""
        if hasattr(self, '_pending_search_timer'):
            self.root.after_cancel(self._pending_search_timer)
        
        # Set a timer to perform search after 500ms of inactivity
        self._pending_search_timer = self.root.after(500, lambda: self.load_pending(self.pending_search_query_var.get()))
    
    def _debounced_transaction_search(self, event=None):
        """Debounced search for transactions to avoid excessive database queries."""
        if hasattr(self, '_transaction_search_timer'):
            self.root.after_cancel(self._transaction_search_timer)
        
        # Set a timer to perform search after 500ms of inactivity
        self._transaction_search_timer = self.root.after(500, lambda: self.load_transactions(self.search_query_var.get()))
    
    def _debounced_report_search(self, event=None):
        """Debounced search for reports to avoid excessive database queries."""
        if hasattr(self, '_report_search_timer'):
            self.root.after_cancel(self._report_search_timer)
        
        # Set a timer to perform search after 500ms of inactivity
        self._report_search_timer = self.root.after(500, lambda: self._search_report_transactions())
    
    def _show_date_picker(self, date_var):
        """Show a simple date picker dialog."""
        try:
            from tkinter import simpledialog
            
            # Get current date or default to today
            current_value = date_var.get()
            if current_value:
                try:
                    current_date = datetime.strptime(current_value, "%Y-%m-%d")
                except ValueError:
                    current_date = datetime.now()
            else:
                current_date = datetime.now()
            
            # Create a simple date input dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Date")
            dialog.geometry("350x250")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Center the dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
            y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
            dialog.geometry(f"+{x}+{y}")
            
            # Add style
            dialog.configure(bg="#f0f0f0")
            
            ttk.Label(dialog, text="Select Date:", font=("Arial", 12, "bold")).pack(pady=15)
            
            # Date input
            date_frame = ttk.Frame(dialog)
            date_frame.pack(pady=15)
            
            ttk.Label(date_frame, text="Year:", font=("Arial", 10)).grid(row=0, column=0, padx=5)
            year_var = tk.StringVar(value=str(current_date.year))
            year_spin = ttk.Spinbox(date_frame, from_=2020, to=2030, textvariable=year_var, width=10, font=("Arial", 10))
            year_spin.grid(row=0, column=1, padx=5)
            
            ttk.Label(date_frame, text="Month:", font=("Arial", 10)).grid(row=0, column=2, padx=5)
            month_var = tk.StringVar(value=str(current_date.month))
            month_spin = ttk.Spinbox(date_frame, from_=1, to=12, textvariable=month_var, width=8, font=("Arial", 10))
            month_spin.grid(row=0, column=3, padx=5)
            
            ttk.Label(date_frame, text="Day:", font=("Arial", 10)).grid(row=0, column=4, padx=5)
            day_var = tk.StringVar(value=str(current_date.day))
            day_spin = ttk.Spinbox(date_frame, from_=1, to=31, textvariable=day_var, width=8, font=("Arial", 10))
            day_spin.grid(row=0, column=5, padx=5)
            
            # Quick select buttons
            quick_frame = ttk.Frame(dialog)
            quick_frame.pack(pady=10)
            
            def today_date():
                today = datetime.now()
                year_var.set(str(today.year))
                month_var.set(str(today.month))
                day_var.set(str(today.day))
            
            def yesterday_date():
                yesterday = datetime.now() - timedelta(days=1)
                year_var.set(str(yesterday.year))
                month_var.set(str(yesterday.month))
                day_var.set(str(yesterday.day))
            
            def tomorrow_date():
                tomorrow = datetime.now() + timedelta(days=1)
                year_var.set(str(tomorrow.year))
                month_var.set(str(tomorrow.month))
                day_var.set(str(tomorrow.day))
            
            ttk.Button(quick_frame, text="Today", command=today_date, width=10).pack(side=tk.LEFT, padx=5)
            ttk.Button(quick_frame, text="Yesterday", command=yesterday_date, width=10).pack(side=tk.LEFT, padx=5)
            ttk.Button(quick_frame, text="Tomorrow", command=tomorrow_date, width=10).pack(side=tk.LEFT, padx=5)
            
            # Buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=20)
            
            def apply_date():
                try:
                    year = int(year_var.get())
                    month = int(month_var.get())
                    day = int(day_var.get())
                    selected_date = datetime(year, month, day)
                    date_var.set(selected_date.strftime("%Y-%m-%d"))
                    dialog.destroy()
                except ValueError:
                    self.msg_box.showerror("Invalid Date", "Please enter a valid date.")
            
            ttk.Button(button_frame, text="Apply", command=apply_date, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Cancel", command=dialog.destroy, width=12).pack(side=tk.LEFT, padx=5)
            
            # Bind Enter key to apply
            dialog.bind('<Return>', lambda event: apply_date())
            dialog.bind('<Escape>', lambda event: dialog.destroy())
            
            # Set focus to year spinbox
            year_spin.focus_set()
            
        except Exception as e:
            # Fallback: just set today's date
            date_var.set(datetime.now().strftime("%Y-%m-%d"))
    
    def _get_current_month_range(self):
        """Helper function to get the start and end dates of the current month."""
        today = datetime.now()
        month_start = today.replace(day=1)
        next_month = month_start.replace(day=28) + timedelta(days=4)
        month_end = next_month - timedelta(days=next_month.day)
        return month_start, month_end
    
    def _set_quick_date_range(self, range_type, tab_type):
        """Set quick date ranges for filters."""
        today = datetime.now()
        from_date_var = None
        to_date_var = None
        
        if tab_type == "pending":
            from_date_var = self.pending_from_date_var
            to_date_var = self.pending_to_date_var
        elif tab_type == "transaction":
            from_date_var = self.from_date_var
            to_date_var = self.to_date_var
        elif tab_type == "report":
            from_date_var = self.report_from_date_var
            to_date_var = self.report_to_date_var
        
        if not from_date_var or not to_date_var:
            return
        
        if range_type == "today":
            from_date_var.set(today.strftime("%Y-%m-%d"))
            to_date_var.set(today.strftime("%Y-%m-%d"))
        elif range_type == "week":
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            from_date_var.set(week_start.strftime("%Y-%m-%d"))
            to_date_var.set(week_end.strftime("%Y-%m-%d"))
        elif range_type == "month":
            month_start = today.replace(day=1)
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
            from_date_var.set(month_start.strftime("%Y-%m-%d"))
            to_date_var.set(month_end.strftime("%Y-%m-%d"))
        
        # Apply filters after setting dates
        if tab_type == "pending":
            self._apply_pending_filters()
        elif tab_type == "transaction":
            self._apply_transaction_filters()
        elif tab_type == "report":
            self._search_report_transactions()

def _get_app_icon_path() -> str:
    """
    Get the path to the application icon.
    Returns the absolute path to app_icon.ico or None if not found.
    """
    # List of possible locations for the icon
    possible_paths = [
        # Relative to the script location (most reliable)
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "app_icon.ico"),
        # Current working directory
        os.path.join(os.getcwd(), "assets", "app_icon.ico"),
        # Parallel assets folder (for development)
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets", "app_icon.ico"),
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return os.path.abspath(path)
    
    return None


if __name__ == "__main__":
    root = tk.Tk()
    
    # Set app icon IMMEDIATELY after creating root window
    # This ensures the icon is displayed in the title bar and taskbar
    icon_path = _get_app_icon_path()
    if icon_path:
        try:
            root.iconbitmap(default=icon_path)
            print(f"[ICON] Icon set successfully: {icon_path}")
        except Exception as e:
            print(f"[ICON] Failed to set iconbitmap: {e}")
            try:
                # Fallback: try iconphoto with PIL
                icon_image = Image.open(icon_path)
                icon_photo = ImageTk.PhotoImage(icon_image)
                root.iconphoto(True, icon_photo)
                print(f"[ICON] Icon set via iconphoto fallback")
            except Exception as e2:
                print(f"[ICON] Failed to set icon via iconphoto: {e2}")
    else:
        print(f"[ICON] Warning: Could not find icon file in any expected location")
    
    app = TruckScaleApp(root)
    root.mainloop()
